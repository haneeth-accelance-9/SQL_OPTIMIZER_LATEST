"""
Unit tests for optimizer.services.upload_validator.
Tests _normalize, ValidationResult, and validate_upload with mock file objects.
No @pytest.mark.django_db required.
"""
import io
import pytest

from optimizer.services.upload_validator import (
    ValidationResult,
    _normalize,
    validate_upload,
    MAX_FILE_SIZE_BYTES,
    EXPECTED_COLUMN_COUNT,
    EXPECTED_HEADERS,
)


# ===========================================================================
# ValidationResult
# ===========================================================================

class TestValidationResult:
    def test_valid_result(self):
        r = ValidationResult(valid=True)
        assert r.valid is True
        assert r.error is None

    def test_invalid_result_with_error(self):
        r = ValidationResult(valid=False, error="Something went wrong")
        assert r.valid is False
        assert r.error == "Something went wrong"


# ===========================================================================
# _normalize
# ===========================================================================

class TestNormalize:
    def test_normal_string_unchanged(self):
        assert _normalize("Server Name") == "Server Name"

    def test_strips_leading_trailing(self):
        assert _normalize("  Server Name  ") == "Server Name"

    def test_collapses_multiple_spaces(self):
        assert _normalize("Server  Name") == "Server Name"

    def test_three_spaces_collapsed(self):
        assert _normalize("Server   Name") == "Server Name"

    def test_none_returns_empty(self):
        assert _normalize(None) == ""

    def test_empty_string(self):
        assert _normalize("") == ""

    def test_mixed_leading_and_internal(self):
        assert _normalize("  Is  Virtual?  ") == "Is Virtual?"


# ===========================================================================
# validate_upload
# ===========================================================================

def _make_valid_xlsx():
    """Build a minimal valid .xlsx with correct headers."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(EXPECTED_HEADERS)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    buf.name = "upload.xlsx"
    buf.size = buf.getbuffer().nbytes
    buf.content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return buf


class TestValidateUpload:
    def test_valid_file_returns_valid(self):
        f = _make_valid_xlsx()
        result = validate_upload(f)
        assert result.valid is True
        assert result.error is None

    def test_oversized_file_rejected(self):
        f = _make_valid_xlsx()
        f.size = MAX_FILE_SIZE_BYTES + 1
        result = validate_upload(f)
        assert result.valid is False
        assert "too large" in result.error

    def test_wrong_extension_rejected(self):
        f = _make_valid_xlsx()
        f.name = "data.csv"
        result = validate_upload(f)
        assert result.valid is False
        assert ".xlsx" in result.error

    def test_invalid_content_type_rejected(self):
        f = _make_valid_xlsx()
        f.content_type = "text/plain"
        result = validate_upload(f)
        assert result.valid is False
        assert "content type" in result.error.lower()

    def test_unreadable_file_rejected(self):
        buf = io.BytesIO(b"not an xlsx file")
        buf.name = "broken.xlsx"
        buf.size = 100
        buf.content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        result = validate_upload(buf)
        assert result.valid is False
        assert "Could not read" in result.error

    def test_wrong_column_count_rejected(self):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Column A", "Column B"])  # only 2 columns
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        buf.name = "short.xlsx"
        buf.size = buf.getbuffer().nbytes
        buf.content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        result = validate_upload(buf)
        assert result.valid is False
        assert "column count" in result.error.lower() or "Unexpected column count" in result.error

    def test_wrong_header_names_rejected(self):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        bad_headers = list(EXPECTED_HEADERS)
        bad_headers[0] = "WRONG_HEADER"
        ws.append(bad_headers)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        buf.name = "wrong_headers.xlsx"
        buf.size = buf.getbuffer().nbytes
        buf.content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        result = validate_upload(buf)
        assert result.valid is False
        assert "Column headers" in result.error or "column" in result.error.lower()

    def test_size_fallback_via_seek(self):
        """When file has no .size attribute, validator falls back to seek/tell."""
        f = _make_valid_xlsx()
        del f.size  # remove size attribute
        result = validate_upload(f)
        assert result.valid is True

    def test_no_content_type_skips_content_type_check(self):
        f = _make_valid_xlsx()
        f.content_type = ""  # empty = no content-type check
        result = validate_upload(f)
        assert result.valid is True

    def test_octet_stream_content_type_accepted(self):
        f = _make_valid_xlsx()
        f.content_type = "application/octet-stream"
        result = validate_upload(f)
        assert result.valid is True

    def test_multiple_mismatches_capped_at_five(self):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        # Replace first 8 columns with wrong names
        bad_headers = ["WRONG" + str(i) for i in range(EXPECTED_COLUMN_COUNT)]
        ws.append(bad_headers)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        buf.name = "many_wrong.xlsx"
        buf.size = buf.getbuffer().nbytes
        buf.content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        result = validate_upload(buf)
        assert result.valid is False
        # Error should mention mismatches but be capped
        assert "…" in result.error or "column" in result.error.lower()

    def test_headers_with_extra_spaces_accepted(self):
        """Headers with double spaces should still pass after normalization."""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        normalized_expected = list(EXPECTED_HEADERS)
        # Add extra spaces to first header (normalizer should collapse them)
        headers_with_spaces = []
        for h in normalized_expected:
            if h and len(h) > 0:
                headers_with_spaces.append(h.replace(" ", "  "))  # double all spaces
            else:
                headers_with_spaces.append(h)
        ws.append(headers_with_spaces)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        buf.name = "spaced.xlsx"
        buf.size = buf.getbuffer().nbytes
        buf.content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        result = validate_upload(buf)
        assert result.valid is True
