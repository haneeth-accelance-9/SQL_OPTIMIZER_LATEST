"""
Extra tests for optimizer.services.report_export pure functions.
No Django DB required.
"""
import io
import re
from datetime import datetime
from unittest.mock import patch

import pytest

from optimizer.services.report_export import (
    DEFAULT_REPORT_TITLE,
    EURO_SYMBOL,
    _build_template_blocks,
    _extract_executive_summary,
    _format_count,
    _format_generated_at,
    _markdown_to_plain,
    _ordinal,
    _parse_report_blocks,
    _remove_hidden_report_rules,
    _to_float,
    _to_int,
    build_report_markdown,
    export_xlsx,
    format_currency,
    normalize_report_content_text,
    normalize_report_currency_text,
)


# ===========================================================================
# _ordinal
# ===========================================================================

def test_ordinal_1():
    assert _ordinal(1) == "1st"

def test_ordinal_2():
    assert _ordinal(2) == "2nd"

def test_ordinal_3():
    assert _ordinal(3) == "3rd"

def test_ordinal_4():
    assert _ordinal(4) == "4th"

def test_ordinal_11():
    assert _ordinal(11) == "11th"

def test_ordinal_12():
    assert _ordinal(12) == "12th"

def test_ordinal_13():
    assert _ordinal(13) == "13th"

def test_ordinal_21():
    assert _ordinal(21) == "21st"

def test_ordinal_22():
    assert _ordinal(22) == "22nd"


# ===========================================================================
# _to_int / _to_float / _format_count
# ===========================================================================

def test_to_int_with_valid_string():
    assert _to_int("42") == 42

def test_to_int_with_none_returns_zero():
    assert _to_int(None) == 0

def test_to_int_with_garbage_returns_zero():
    assert _to_int("abc") == 0

def test_to_int_with_zero_string():
    assert _to_int("0") == 0

def test_to_float_with_valid_string():
    assert _to_float("3.14") == pytest.approx(3.14)

def test_to_float_with_none_returns_zero():
    assert _to_float(None) == pytest.approx(0.0)

def test_to_float_with_invalid():
    assert _to_float("xyz") == pytest.approx(0.0)

def test_format_count_formats_with_commas():
    assert _format_count(1234567) == "1,234,567"

def test_format_count_with_zero():
    assert _format_count(0) == "0"

def test_format_count_with_string_number():
    assert _format_count("100") == "100"


# ===========================================================================
# format_currency
# ===========================================================================

def test_format_currency_small_value():
    result = format_currency(345.6)
    assert "345,60" in result
    assert EURO_SYMBOL in result

def test_format_currency_large_value_uses_dots_as_thousands():
    result = format_currency(1234567.89)
    assert "1.234.567,89" in result
    assert EURO_SYMBOL in result

def test_format_currency_zero():
    result = format_currency(0)
    assert "0,00" in result
    assert EURO_SYMBOL in result

def test_format_currency_with_string_number():
    result = format_currency("100.5")
    assert "100,50" in result

def test_format_currency_large_round_number():
    result = format_currency(1000000)
    assert "1.000.000,00" in result


# ===========================================================================
# normalize_report_currency_text
# ===========================================================================

def test_normalize_report_currency_text_replaces_dollar_with_euro():
    text = "Total cost: $1,234.56"
    result = normalize_report_currency_text(text)
    assert "$" not in result
    assert EURO_SYMBOL in result

def test_normalize_report_currency_text_no_change_when_no_dollar():
    text = "Some plain text without money symbols"
    result = normalize_report_currency_text(text)
    assert result == text


# ===========================================================================
# normalize_report_content_text
# ===========================================================================

def test_normalize_report_content_text_strips_multiple_blank_lines():
    text = "Line one.\n\n\n\nLine two."
    result = normalize_report_content_text(text)
    assert "\n\n\n" not in result

def test_normalize_report_content_text_converts_backtick_rule_ids_in_table():
    text = "| `uc_1_1_azure_byol_to_payg` |\n| some value |"
    result = normalize_report_content_text(text)
    assert "`uc_1_1_azure_byol_to_payg`" not in result
    assert "Uc 1 1 Azure Byol To Payg" in result

def test_normalize_report_content_text_converts_backtick_rule_ids_in_heading():
    text = "### `uc_1_2_retired_devices`"
    result = normalize_report_content_text(text)
    assert "`uc_1_2_retired_devices`" not in result

def test_normalize_report_content_text_strips_html_tags():
    text = "Total: <span id='val'>42</span> devices"
    result = normalize_report_content_text(text)
    assert "<span" not in result
    assert "42" in result


# ===========================================================================
# _remove_hidden_report_rules
# ===========================================================================

def test_remove_hidden_report_rules_removes_lifecycle_risk_flags_table_row():
    text = "| Uc 3 5 Lifecycle Risk Flags | 5 | High |\n| Other Rule | 1 | Low |"
    result = _remove_hidden_report_rules(text)
    assert "Lifecycle Risk Flags" not in result
    assert "Other Rule" in result

def test_remove_hidden_report_rules_removes_physical_system_review_section():
    text = "### Uc 3 6 Physical System Review\n\nSome detail here.\n\n### Next Section"
    result = _remove_hidden_report_rules(text)
    assert "Physical System Review" not in result

def test_remove_hidden_report_rules_preserves_unrelated_content():
    text = "### Executive Summary\n\nThis is important content.\n\n### Recommendations"
    result = _remove_hidden_report_rules(text)
    assert "Executive Summary" in result
    assert "Recommendations" in result

def test_remove_hidden_report_rules_removes_lifecycle_risk_flags_section():
    text = "### Lifecycle Risk Flags (Human Review)\n\nDetails here.\n\n## Next"
    result = _remove_hidden_report_rules(text)
    assert "Lifecycle Risk Flags" not in result

def test_remove_hidden_report_rules_removes_physical_systems_require_review():
    text = "| Physical Systems Require Review | 3 | Medium |\n| Rule X | 2 | Low |"
    result = _remove_hidden_report_rules(text)
    assert "Physical Systems Require Review" not in result
    assert "Rule X" in result


# ===========================================================================
# _parse_report_blocks
# ===========================================================================

def test_parse_report_blocks_extracts_title_from_h1():
    text = "# My Report Title\n\n## Section One\n\nSome text."
    title, blocks = _parse_report_blocks(text)
    assert title == "My Report Title"

def test_parse_report_blocks_extracts_section_blocks():
    text = "# Title\n\n## Section One\n\n## Section Two"
    _, blocks = _parse_report_blocks(text)
    section_texts = [b["text"] for b in blocks if b["kind"] == "section"]
    assert "Section One" in section_texts
    assert "Section Two" in section_texts

def test_parse_report_blocks_extracts_bullet_blocks_dash():
    text = "# Title\n\n- First bullet\n- Second bullet"
    _, blocks = _parse_report_blocks(text)
    bullets = [b["text"] for b in blocks if b["kind"] == "bullet"]
    assert "First bullet" in bullets

def test_parse_report_blocks_extracts_bullet_blocks_asterisk():
    text = "# Title\n\n* Asterisk bullet"
    _, blocks = _parse_report_blocks(text)
    bullets = [b["text"] for b in blocks if b["kind"] == "bullet"]
    assert "Asterisk bullet" in bullets

def test_parse_report_blocks_extracts_table_rows_and_skips_separator():
    text = "# Title\n\n| Rule | Count |\n|------|-------|\n| Rule A | 5 |"
    _, blocks = _parse_report_blocks(text)
    table_rows = [b for b in blocks if b["kind"] == "table_row"]
    assert len(table_rows) == 2  # header + data (separator skipped)
    assert table_rows[0]["cells"] == ["Rule", "Count"]

def test_parse_report_blocks_falls_back_to_default_title_when_no_h1():
    text = "## Section Only\n\nNo H1 here."
    title, _ = _parse_report_blocks(text)
    assert title == DEFAULT_REPORT_TITLE

def test_parse_report_blocks_extracts_subsection_blocks():
    text = "# Title\n\n## Section\n\n### Subsection A"
    _, blocks = _parse_report_blocks(text)
    subsections = [b["text"] for b in blocks if b["kind"] == "subsection"]
    assert "Subsection A" in subsections

def test_parse_report_blocks_extracts_numbered_list():
    text = "# Title\n\n1. First item\n2. Second item"
    _, blocks = _parse_report_blocks(text)
    numbered = [b for b in blocks if b["kind"] == "numbered"]
    assert len(numbered) == 2
    assert numbered[0]["text"] == "First item"
    assert numbered[0]["label"] == "1."

def test_parse_report_blocks_extracts_paragraph():
    text = "# Title\n\nThis is a paragraph of plain text."
    _, blocks = _parse_report_blocks(text)
    paragraphs = [b for b in blocks if b["kind"] == "paragraph"]
    assert any("paragraph of plain text" in b["text"] for b in paragraphs)

def test_parse_report_blocks_extracts_horizontal_rule():
    text = "# Title\n\nSome text.\n\n---\n\nMore text."
    _, blocks = _parse_report_blocks(text)
    rules = [b for b in blocks if b["kind"] == "rule"]
    assert len(rules) == 1


# ===========================================================================
# _markdown_to_plain
# ===========================================================================

def test_markdown_to_plain_strips_bold_markers():
    result = _markdown_to_plain("**Bold text** and normal")
    assert "**" not in result
    assert "Bold text" in result

def test_markdown_to_plain_strips_heading_markers():
    result = _markdown_to_plain("## Section Header")
    assert "##" not in result
    assert "Section Header" in result

def test_markdown_to_plain_strips_backtick_code():
    result = _markdown_to_plain("Use `some_code` here")
    assert "`" not in result
    assert "some_code" in result

def test_markdown_to_plain_strips_italic_markers():
    result = _markdown_to_plain("*italicized* text")
    assert "*" not in result
    assert "italicized" in result


# ===========================================================================
# _extract_executive_summary
# ===========================================================================

def test_extract_executive_summary_returns_paragraph_from_section():
    text = "# Report\n\n## Executive Summary\n\nThis is the executive summary paragraph.\n\n## Other Section"
    result = _extract_executive_summary(text, {})
    assert "executive summary paragraph" in result.lower()

def test_extract_executive_summary_falls_back_to_default_when_section_missing():
    text = "# Report\n\n## Other Section\n\nSome content."
    context = {
        "total_demand_quantity": 100,
        "total_license_cost": 5000.0,
        "azure_payg_count": 3,
        "retired_count": 2,
    }
    result = _extract_executive_summary(text, context)
    assert isinstance(result, str)
    assert len(result) > 0


# ===========================================================================
# _build_template_blocks
# ===========================================================================

def test_build_template_blocks_returns_default_report_title():
    title, _ = _build_template_blocks("## Executive Summary\nSummary", {})
    assert title == DEFAULT_REPORT_TITLE

def test_build_template_blocks_includes_various_block_kinds():
    _, blocks = _build_template_blocks(
        "## Executive Summary\nSummary",
        {
            "azure_payg_count": 5,
            "retired_count": 3,
            "total_demand_quantity": 1000,
            "total_license_cost": 500000.0,
            "total_savings": 10000.0,
            "azure_payg_savings": 8000.0,
            "retired_devices_savings": 2000.0,
        },
    )
    kinds = {b["kind"] for b in blocks}
    assert "section" in kinds
    assert "bullet" in kinds

def test_build_template_blocks_includes_savings_bullets():
    _, blocks = _build_template_blocks(
        "## Executive Summary\nSummary",
        {
            "azure_payg_count": 2,
            "retired_count": 1,
            "total_demand_quantity": 500,
            "total_license_cost": 50000.0,
            "total_savings": 5000.0,
            "azure_payg_savings": 4000.0,
            "retired_devices_savings": 1000.0,
        },
    )
    bullet_texts = [b["text"] for b in blocks if b["kind"] == "bullet"]
    assert any("savings" in t.lower() for t in bullet_texts)

def test_build_template_blocks_formats_currency_with_euro():
    _, blocks = _build_template_blocks(
        "## Executive Summary\nSummary",
        {
            "azure_payg_count": 2,
            "retired_count": 1,
            "total_demand_quantity": 1000,
            "total_license_cost": 999999.99,
            "total_savings": 12345.67,
            "azure_payg_savings": 10000.0,
            "retired_devices_savings": 2345.67,
        },
    )
    all_texts = " ".join(b.get("text", "") for b in blocks)
    assert EURO_SYMBOL in all_texts

def test_build_template_blocks_with_empty_context_does_not_raise():
    title, blocks = _build_template_blocks("## Executive Summary\nSummary", {})
    assert isinstance(title, str)
    assert isinstance(blocks, list)


# ===========================================================================
# build_report_markdown
# ===========================================================================

def test_build_report_markdown_produces_string_starting_with_h1():
    md = build_report_markdown("## Executive Summary\nSummary", {
        "azure_payg_count": 0,
        "retired_count": 0,
        "total_demand_quantity": 100,
        "total_license_cost": 1000.0,
    })
    assert md.startswith("# ")

def test_build_report_markdown_includes_cost_section():
    md = build_report_markdown("## Executive Summary\nSummary", {
        "azure_payg_count": 5,
        "retired_count": 2,
        "total_demand_quantity": 200,
        "total_license_cost": 20000.0,
        "total_savings": 500.0,
        "azure_payg_savings": 400.0,
        "retired_devices_savings": 100.0,
    })
    assert "Cost" in md

def test_build_report_markdown_without_context_parses_raw_text():
    raw = "# Custom Title\n\n## Executive Summary\n\nThis is a summary.\n\n## Recommendations\n\n- Do something"
    md = build_report_markdown(raw)
    assert "Executive Summary" in md


# ===========================================================================
# _format_generated_at
# ===========================================================================

def test_format_generated_at_uses_provided_datetime():
    dt = datetime(2025, 6, 15, 14, 30)
    result = _format_generated_at(dt)
    assert "June 2025" in result
    assert "15th" in result

def test_format_generated_at_without_argument_returns_string():
    result = _format_generated_at()
    assert isinstance(result, str)
    assert len(result) > 5

def test_format_generated_at_first_of_month():
    dt = datetime(2025, 1, 1, 9, 0)
    result = _format_generated_at(dt)
    assert "1st" in result
    assert "January 2025" in result


# ===========================================================================
# export_xlsx
# ===========================================================================

def test_export_xlsx_returns_bytes():
    content = export_xlsx("## Section\n\nSome report content.")
    assert isinstance(content, bytes)
    assert len(content) > 0

def test_export_xlsx_produces_readable_workbook():
    from openpyxl import load_workbook
    content = export_xlsx(
        "## Executive Summary\nSummary",
        report_context={
            "azure_payg_count": 3,
            "retired_count": 1,
            "total_demand_quantity": 500,
            "total_license_cost": 50000.0,
            "total_savings": 1000.0,
            "azure_payg_savings": 800.0,
            "retired_devices_savings": 200.0,
        },
    )
    wb = load_workbook(io.BytesIO(content))
    assert "Report" in wb.sheetnames

def test_export_xlsx_title_cell_contains_report_title():
    from openpyxl import load_workbook
    content = export_xlsx("## Executive Summary\nSummary")
    wb = load_workbook(io.BytesIO(content))
    sheet = wb["Report"]
    title_cell = sheet["A1"].value
    assert title_cell is not None
    assert DEFAULT_REPORT_TITLE in str(title_cell)

def test_export_xlsx_with_savings_data_in_output():
    from openpyxl import load_workbook
    content = export_xlsx(
        "## Executive Summary\nSummary",
        report_context={
            "azure_payg_count": 10,
            "retired_count": 4,
            "total_demand_quantity": 1000,
            "total_license_cost": 100000.0,
            "total_savings": 9999.99,
            "azure_payg_savings": 7777.77,
            "retired_devices_savings": 2222.22,
        },
    )
    wb = load_workbook(io.BytesIO(content))
    sheet = wb["Report"]
    all_values = [
        str(sheet.cell(row=r, column=1).value or "")
        for r in range(1, sheet.max_row + 1)
    ]
    joined = " ".join(all_values)
    assert "Savings" in joined

def test_export_xlsx_table_rows_written_across_columns():
    from openpyxl import load_workbook
    text = "# Title\n\n| Rule | Count | Status |\n|------|-------|--------|\n| Rule A | 5 | Active |"
    content = export_xlsx(text)
    wb = load_workbook(io.BytesIO(content))
    sheet = wb["Report"]
    found_rule_a = any(
        any("Rule A" in str(c or "") for c in row)
        for row in sheet.iter_rows(values_only=True)
        if row
    )
    assert found_rule_a
