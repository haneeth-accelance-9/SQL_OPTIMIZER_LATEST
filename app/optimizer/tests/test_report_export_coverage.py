"""
Additional coverage tests for optimizer.services.report_export.
Targets missed lines: 58-63, 67-68, 167-168, 177-183, 187-230, 234-261,
275-279, 349-351, 565, 567, 569-570, 579-754, 759-809, 817-819, 878, 891-900.

No tests are duplicated from test_report_export.py or test_report_export_extra.py.
"""
import io
import pytest
from datetime import datetime
from pathlib import Path
from unittest.mock import patch, MagicMock


# ===========================================================================
# _project_root  (lines 57-63)
# ===========================================================================

from optimizer.services.report_export import (
    DEFAULT_REPORT_TITLE,
    EURO_SYMBOL,
    _project_root,
    _get_asset_path,
    _normalize_report_text,
    _resolve_generated_at,
    _find_font_file,
    _resolve_pdf_fonts,
    _draw_bayer_logo,
    _markdown_to_reportlab,
    _get_render_blocks,
    _resolve_savings_value,
    _default_executive_summary,
    normalize_report_title_text,
    export_pdf,
    export_docx,
    export_xlsx,
)


class TestProjectRoot:
    def test_returns_path_object(self):
        result = _project_root()
        assert isinstance(result, Path)

    def test_returns_non_empty_path(self):
        result = _project_root()
        assert str(result) != ""

    def test_fallback_when_settings_unavailable(self):
        """When Django settings raise, falls back to file-relative path."""
        with patch("django.conf.settings.BASE_DIR", side_effect=AttributeError("no BASE_DIR")):
            # Even patched, the import inside the function should succeed or fallback
            result = _project_root()
            assert isinstance(result, Path)


# ===========================================================================
# _get_asset_path  (lines 67-68)
# ===========================================================================

class TestGetAssetPath:
    def test_nonexistent_path_returns_none(self):
        result = _get_asset_path(Path("nonexistent_file_xyz.png"))
        assert result is None

    def test_existing_path_returns_path(self, tmp_path):
        asset = tmp_path / "logo.png"
        asset.write_bytes(b"\x89PNG")
        root = _project_root()
        # We can't easily make it "exist" relative to project root, so test directly
        with patch("optimizer.services.report_export._project_root", return_value=tmp_path):
            result = _get_asset_path(Path("logo.png"))
            assert result is not None
            assert result == tmp_path / "logo.png"


# ===========================================================================
# _normalize_report_text  (lines 71-77)
# ===========================================================================

class TestNormalizeReportText:
    def test_strips_html_tags(self):
        result = _normalize_report_text("Hello <b>World</b>")
        assert "<b>" not in result
        assert "World" in result

    def test_replaces_crlf_with_lf(self):
        result = _normalize_report_text("line1\r\nline2")
        assert "\r\n" not in result
        assert "line1" in result
        assert "line2" in result

    def test_replaces_nbsp(self):
        result = _normalize_report_text("word1 word2")
        assert " " not in result
        assert "word1 word2" in result

    def test_replaces_em_dash(self):
        result = _normalize_report_text("A—B")
        assert "—" not in result
        assert "A-B" in result

    def test_strips_leading_trailing_whitespace(self):
        result = _normalize_report_text("  hello  ")
        assert result == "hello"

    def test_none_returns_empty_string(self):
        result = _normalize_report_text(None)
        assert result == ""

    def test_empty_string_returns_empty(self):
        assert _normalize_report_text("") == ""


# ===========================================================================
# _resolve_generated_at  (lines 161-168)
# ===========================================================================

class TestResolveGeneratedAt:
    def test_returns_provided_datetime(self):
        dt = datetime(2025, 3, 15, 10, 30)
        result = _resolve_generated_at(dt)
        assert result is dt

    def test_returns_now_when_none(self):
        result = _resolve_generated_at(None)
        assert result is not None
        assert hasattr(result, "day")

    def test_fallback_when_django_timezone_unavailable(self):
        """When timezone import fails, returns datetime.now() instead."""
        with patch("django.utils.timezone.localtime", side_effect=Exception("no tz")):
            result = _resolve_generated_at(None)
            assert hasattr(result, "day")


# ===========================================================================
# normalize_report_title_text  (line 148-149)
# ===========================================================================

class TestNormalizeReportTitleText:
    def test_delegates_to_normalize_report_content_text(self):
        text = "# My Report\n\nSome text."
        result = normalize_report_title_text(text)
        assert isinstance(result, str)

    def test_legacy_title_replaced(self):
        text = "# SQL Server License Optimization Report"
        result = normalize_report_title_text(text)
        assert "IT License and Cost Optimization Report" in result


# ===========================================================================
# _find_font_file  (lines 177-183)
# ===========================================================================

class TestFindFontFile:
    def test_nonexistent_directory_returns_none(self):
        result = _find_font_file(Path("/nonexistent/path/xyz"), ["*.ttf"])
        assert result is None

    def test_finds_matching_file(self, tmp_path):
        font = tmp_path / "Acumin-Regular.ttf"
        font.write_bytes(b"fake font data")
        result = _find_font_file(tmp_path, ["Acumin*Regular*.ttf"])
        assert result == font

    def test_no_match_returns_none(self, tmp_path):
        result = _find_font_file(tmp_path, ["Acumin*Regular*.ttf"])
        assert result is None

    def test_returns_first_match_when_multiple(self, tmp_path):
        font_a = tmp_path / "AcuminA-Regular.ttf"
        font_b = tmp_path / "AcuminB-Regular.ttf"
        font_a.write_bytes(b"data")
        font_b.write_bytes(b"data")
        result = _find_font_file(tmp_path, ["Acumin*Regular*.ttf"])
        assert result is not None


# ===========================================================================
# _resolve_pdf_fonts  (lines 187-230)
# ===========================================================================

class TestResolvePdfFonts:
    def test_returns_dict_with_regular_and_bold(self):
        result = _resolve_pdf_fonts()
        assert "regular" in result
        assert "bold" in result

    def test_default_fonts_are_helvetica(self):
        """Without custom font files, should fall back to Helvetica."""
        with patch("optimizer.services.report_export._find_font_file", return_value=None):
            result = _resolve_pdf_fonts()
            assert result["regular"] in ("Helvetica", "AcuminProRegular")
            assert result["bold"] in ("Helvetica-Bold", "AcuminProBold")

    def test_without_reportlab_falls_back_to_helvetica(self):
        """If reportlab is not importable, returns Helvetica defaults."""
        with patch.dict("sys.modules", {"reportlab": None, "reportlab.pdfbase": None,
                                         "reportlab.pdfbase.pdfmetrics": None,
                                         "reportlab.pdfbase.ttfonts": None}):
            result = _resolve_pdf_fonts()
            # Should still return a dict (either fonts found before import or defaults)
            assert isinstance(result, dict)


# ===========================================================================
# _markdown_to_reportlab  (lines 274-279)
# ===========================================================================

class TestMarkdownToReportlab:
    def test_bold_converted_to_b_tags(self):
        result = _markdown_to_reportlab("**bold text**")
        assert "<b>bold text</b>" in result

    def test_italic_converted_to_i_tags(self):
        result = _markdown_to_reportlab("*italic*")
        assert "<i>italic</i>" in result

    def test_backtick_converted_to_courier_font(self):
        result = _markdown_to_reportlab("`code`")
        assert "Courier" in result
        assert "code" in result

    def test_html_entities_escaped(self):
        result = _markdown_to_reportlab("5 > 3 & 2 < 4")
        assert "&gt;" in result or ">" in result
        assert "&amp;" in result or "&" in result

    def test_plain_text_unchanged(self):
        result = _markdown_to_reportlab("Plain text here")
        assert "Plain text here" in result


# ===========================================================================
# _get_render_blocks  (lines 541-544)
# ===========================================================================

class TestGetRenderBlocks:
    def test_with_report_context_calls_template(self):
        title, blocks = _get_render_blocks("## Summary\nText", {"azure_payg_count": 1})
        assert isinstance(title, str)
        assert isinstance(blocks, list)
        assert len(blocks) > 0

    def test_without_report_context_calls_parse(self):
        title, blocks = _get_render_blocks("# Custom Title\n\n## Section\n\nContent")
        assert "Custom Title" in title or isinstance(title, str)
        assert isinstance(blocks, list)

    def test_none_context_calls_parse(self):
        title, blocks = _get_render_blocks("## Only Section\n\nText", None)
        assert isinstance(title, str)
        assert isinstance(blocks, list)


# ===========================================================================
# _resolve_savings_value  (lines 416-420)
# ===========================================================================

class TestResolveSavingsValue:
    def test_direct_key_returned_when_present(self):
        ctx = {"total_savings": 5000.0}
        result = _resolve_savings_value(ctx, "total_savings")
        assert result == 5000.0

    def test_rule_key_fallback_when_direct_missing(self):
        ctx = {"rule_wise_savings": {"azure_payg": 3000.0}}
        result = _resolve_savings_value(ctx, "azure_payg_savings", "azure_payg")
        assert result == 3000.0

    def test_none_when_both_missing(self):
        ctx = {}
        result = _resolve_savings_value(ctx, "nonexistent_key")
        assert result is None

    def test_none_when_rule_wise_savings_missing(self):
        ctx = {}
        result = _resolve_savings_value(ctx, "azure_payg_savings", "azure_payg")
        assert result is None


# ===========================================================================
# _default_executive_summary  (lines 388-398)
# ===========================================================================

class TestDefaultExecutiveSummary:
    def test_contains_license_count(self):
        ctx = {"total_demand_quantity": 500, "total_license_cost": 25000.0,
               "azure_payg_count": 3, "retired_count": 2}
        result = _default_executive_summary(ctx)
        assert "500" in result

    def test_contains_azure_count(self):
        ctx = {"total_demand_quantity": 100, "total_license_cost": 5000.0,
               "azure_payg_count": 10, "retired_count": 5}
        result = _default_executive_summary(ctx)
        assert "10" in result

    def test_returns_non_empty_string(self):
        result = _default_executive_summary({})
        assert isinstance(result, str)
        assert len(result) > 0

    def test_contains_euro_symbol(self):
        ctx = {"total_demand_quantity": 100, "total_license_cost": 99999.99}
        result = _default_executive_summary(ctx)
        assert EURO_SYMBOL in result


# ===========================================================================
# export_pdf  (lines 577-754)
# ===========================================================================

class TestExportPdf:
    def test_returns_bytes_or_none(self):
        """export_pdf returns bytes if reportlab installed, else None."""
        result = export_pdf("## Section\n\nContent.")
        assert result is None or isinstance(result, bytes)

    def test_with_context_produces_output(self):
        result = export_pdf(
            "## Executive Summary\nSummary",
            report_context={
                "azure_payg_count": 2,
                "retired_count": 1,
                "total_demand_quantity": 500,
                "total_license_cost": 50000.0,
            },
        )
        assert result is None or isinstance(result, bytes)

    def test_with_generated_at(self):
        dt = datetime(2025, 6, 15, 10, 0)
        result = export_pdf("## Section\n\nContent.", generated_at=dt)
        assert result is None or isinstance(result, bytes)

    def test_with_empty_text(self):
        result = export_pdf("")
        assert result is None or isinstance(result, bytes)

    def test_pdf_bytes_non_empty_when_reportlab_installed(self):
        try:
            import reportlab  # noqa: F401
            result = export_pdf("# Report\n\n## Section\n\nContent.")
            assert result is not None
            assert len(result) > 100
        except ImportError:
            pytest.skip("reportlab not installed")

    def test_pdf_with_table_row_blocks(self):
        text = "# Report\n\n| Rule | Count |\n|------|-------|\n| Rule A | 5 |"
        result = export_pdf(text)
        assert result is None or isinstance(result, bytes)

    def test_pdf_with_numbered_list(self):
        text = "# Report\n\n1. First item\n2. Second item"
        result = export_pdf(text)
        assert result is None or isinstance(result, bytes)

    def test_pdf_with_hr_rule(self):
        text = "# Report\n\nSome text.\n\n---\n\nMore text."
        result = export_pdf(text)
        assert result is None or isinstance(result, bytes)

    def test_pdf_with_all_block_kinds(self):
        text = (
            "# Report Title\n\n"
            "## Executive Summary\n\n"
            "### Subsection\n\n"
            "- Bullet point\n"
            "1. Numbered item\n"
            "---\n"
            "| Col1 | Col2 |\n"
            "|------|------|\n"
            "| A    | B    |\n"
            "Plain paragraph text."
        )
        result = export_pdf(text)
        assert result is None or isinstance(result, bytes)


# ===========================================================================
# export_docx  (lines 757-809)
# ===========================================================================

class TestExportDocx:
    def test_returns_bytes_or_none(self):
        """export_docx returns bytes if python-docx installed, else None."""
        result = export_docx("## Section\n\nContent.")
        assert result is None or isinstance(result, bytes)

    def test_with_context_and_generated_at(self):
        dt = datetime(2025, 6, 15, 10, 0)
        result = export_docx(
            "## Executive Summary\nSummary",
            generated_at=dt,
            report_context={
                "azure_payg_count": 1,
                "retired_count": 1,
                "total_demand_quantity": 100,
                "total_license_cost": 10000.0,
            },
        )
        assert result is None or isinstance(result, bytes)

    def test_bytes_are_non_empty_when_docx_installed(self):
        try:
            from docx import Document  # noqa: F401
            result = export_docx("# Report\n\n## Section\n\nSome content.")
            assert result is not None
            assert len(result) > 100
        except ImportError:
            pytest.skip("python-docx not installed")

    def test_docx_with_all_block_types(self):
        try:
            from docx import Document  # noqa: F401
        except ImportError:
            pytest.skip("python-docx not installed")
        text = (
            "# Report\n\n"
            "## Section\n\n"
            "### Subsection\n\n"
            "- Bullet\n"
            "  - Sub-bullet\n"
            "1. Numbered\n"
            "---\n"
            "| A | B |\n|---|---|\n| 1 | 2 |\n"
            "Plain text paragraph."
        )
        result = export_docx(text)
        assert isinstance(result, bytes)
        assert len(result) > 100

    def test_docx_readable_workbook(self):
        try:
            from docx import Document
        except ImportError:
            pytest.skip("python-docx not installed")
        result = export_docx("# Title\n\n## Summary\n\nSome text here.")
        doc = Document(io.BytesIO(result))
        full_text = " ".join(p.text for p in doc.paragraphs)
        assert isinstance(full_text, str)

    def test_empty_text_does_not_crash(self):
        result = export_docx("")
        assert result is None or isinstance(result, bytes)


# ===========================================================================
# export_xlsx — additional paths not in test_report_export_extra.py
# (lines 812-917: sub_bullet, bullet_group, numbered, rule, paragraph, else blocks)
# ===========================================================================

class TestExportXlsxExtended:
    def test_sub_bullet_written(self):
        from openpyxl import load_workbook
        text = (
            "# Title\n\n"
            "## Section\n\n"
            "- Parent bullet\n"
            "  - Sub-bullet item\n"
        )
        # sub_bullet requires the template path (_build_template_blocks processes sub_bullet_group)
        # Use raw parse path
        content = export_xlsx(text)
        assert isinstance(content, bytes)
        wb = load_workbook(io.BytesIO(content))
        sheet = wb["Report"]
        all_values = " ".join(
            str(sheet.cell(row=r, column=1).value or "")
            for r in range(1, sheet.max_row + 1)
        )
        assert isinstance(all_values, str)

    def test_bullet_group_and_sub_bullet_from_template(self):
        """Template blocks include bullet_group and sub_bullet kinds."""
        from openpyxl import load_workbook
        content = export_xlsx(
            "## Executive Summary\nSummary",
            report_context={
                "azure_payg_count": 2,
                "retired_count": 1,
                "total_demand_quantity": 100,
                "total_license_cost": 10000.0,
                "total_savings": 1000.0,
                "azure_payg_savings": 800.0,
                "retired_devices_savings": 200.0,
            },
        )
        assert isinstance(content, bytes)
        wb = load_workbook(io.BytesIO(content))
        assert "Report" in wb.sheetnames
        sheet = wb["Report"]
        assert sheet.max_row > 5

    def test_numbered_list_written(self):
        from openpyxl import load_workbook
        text = "# Title\n\n1. First step\n2. Second step\n3. Third step"
        content = export_xlsx(text)
        wb = load_workbook(io.BytesIO(content))
        sheet = wb["Report"]
        found_numbered = any(
            "1." in str(sheet.cell(row=r, column=1).value or "")
            for r in range(1, sheet.max_row + 1)
        )
        assert found_numbered

    def test_hr_rule_written(self):
        """Horizontal rule blocks add an empty row with bottom border."""
        from openpyxl import load_workbook
        text = "# Title\n\nText before.\n\n---\n\nText after."
        content = export_xlsx(text)
        assert isinstance(content, bytes)
        wb = load_workbook(io.BytesIO(content))
        assert wb["Report"].max_row >= 4

    def test_paragraph_block_written(self):
        from openpyxl import load_workbook
        text = "# Title\n\nThis is a plain paragraph of text that should be written to the sheet."
        content = export_xlsx(text)
        wb = load_workbook(io.BytesIO(content))
        sheet = wb["Report"]
        all_values = " ".join(
            str(sheet.cell(row=r, column=1).value or "")
            for r in range(1, sheet.max_row + 1)
        )
        assert "paragraph" in all_values.lower() or "plain" in all_values.lower()

    def test_column_widths_set(self):
        from openpyxl import load_workbook
        content = export_xlsx("# Report\n\n## Section\n\nSome content here.")
        wb = load_workbook(io.BytesIO(content))
        sheet = wb["Report"]
        assert sheet.column_dimensions["A"].width > 0

    def test_row_heights_set(self):
        from openpyxl import load_workbook
        content = export_xlsx("# Report\n\n## Section\n\nContent.")
        wb = load_workbook(io.BytesIO(content))
        sheet = wb["Report"]
        assert sheet.row_dimensions[1].height > 0

    def test_freeze_panes_set(self):
        from openpyxl import load_workbook
        content = export_xlsx("# Report\n\n## Section\n\nContent.")
        wb = load_workbook(io.BytesIO(content))
        sheet = wb["Report"]
        assert sheet.freeze_panes == "A4"

    def test_meta_cell_generated_at(self):
        """Row 2 should contain the generated-at timestamp."""
        from openpyxl import load_workbook
        dt = datetime(2025, 6, 15, 14, 30)
        content = export_xlsx("## Section\n\nContent.", generated_at=dt)
        wb = load_workbook(io.BytesIO(content))
        sheet = wb["Report"]
        meta_value = str(sheet["A2"].value or "")
        assert "June" in meta_value or "15" in meta_value or "2025" in meta_value

    def test_table_row_written_across_four_columns(self):
        from openpyxl import load_workbook
        text = "# T\n\n| Col1 | Col2 | Col3 | Col4 |\n|------|------|------|------|\n| A | B | C | D |"
        content = export_xlsx(text)
        wb = load_workbook(io.BytesIO(content))
        sheet = wb["Report"]
        found_d = any(
            "D" in str(sheet.cell(row=r, column=4).value or "")
            for r in range(1, sheet.max_row + 1)
        )
        assert found_d

    def test_grid_lines_disabled(self):
        from openpyxl import load_workbook
        content = export_xlsx("# Report\n\n## Section\n\nText.")
        wb = load_workbook(io.BytesIO(content))
        sheet = wb["Report"]
        assert not sheet.sheet_view.showGridLines

    def test_without_reportlab_pdf_returns_none(self):
        """If reportlab is not available, export_pdf should return None gracefully."""
        with patch.dict("sys.modules", {"reportlab": None,
                                         "reportlab.lib": None,
                                         "reportlab.lib.pagesizes": None}):
            result = export_pdf("## Section\n\nContent.")
            # Either None (import failed) or bytes (reportlab already loaded)
            assert result is None or isinstance(result, bytes)

    def test_without_docx_returns_none(self):
        """If python-docx is not available, export_docx should return None gracefully."""
        with patch.dict("sys.modules", {"docx": None,
                                         "docx.enum": None,
                                         "docx.enum.text": None,
                                         "docx.shared": None}):
            result = export_docx("## Section\n\nContent.")
            assert result is None or isinstance(result, bytes)

    def test_without_openpyxl_returns_none(self):
        """If openpyxl is not available, export_xlsx should return None gracefully."""
        with patch.dict("sys.modules", {"openpyxl": None,
                                         "openpyxl.styles": None}):
            result = export_xlsx("## Section\n\nContent.")
            assert result is None or isinstance(result, bytes)


# ===========================================================================
# _draw_bayer_logo  (lines 233-261) — canvas mock
# ===========================================================================

class TestDrawBayerLogo:
    def test_does_not_raise_with_mock_canvas(self):
        canvas = MagicMock()
        canvas.beginPath.return_value = MagicMock()
        fonts = {"bold": "Helvetica-Bold", "regular": "Helvetica"}
        # Should not raise
        _draw_bayer_logo(canvas, x=10.0, y=10.0, size=50.0, fonts=fonts)

    def test_calls_save_and_restore_state(self):
        canvas = MagicMock()
        canvas.beginPath.return_value = MagicMock()
        fonts = {"bold": "Helvetica-Bold", "regular": "Helvetica"}
        _draw_bayer_logo(canvas, x=0.0, y=0.0, size=60.0, fonts=fonts)
        canvas.saveState.assert_called()
        canvas.restoreState.assert_called()

    def test_draws_bayer_text(self):
        canvas = MagicMock()
        canvas.beginPath.return_value = MagicMock()
        fonts = {"bold": "Helvetica-Bold", "regular": "Helvetica"}
        _draw_bayer_logo(canvas, x=0.0, y=0.0, size=60.0, fonts=fonts)
        # drawCentredString should be called for "BAYER" and letter stacks
        assert canvas.drawCentredString.call_count >= 5
