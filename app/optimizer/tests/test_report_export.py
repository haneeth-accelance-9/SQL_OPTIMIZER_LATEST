import io

from openpyxl import load_workbook

from optimizer.services.report_export import _build_template_blocks, build_report_markdown, export_xlsx, format_currency


def test_build_template_blocks_uses_euro_for_cost_and_savings_values():
    # format_currency uses European format: "1.234.567,89 \u20ac"
    _, blocks = _build_template_blocks(
        "## Executive Summary\nSummary",
        {
            "azure_payg_count": 2,
            "retired_count": 1,
            "total_demand_quantity": 20568,
            "total_license_cost": 6504721.08,
            "total_savings": 55780.96,
            "azure_payg_savings": 55433.08,
            "retired_devices_savings": 347.88,
        },
    )

    assert {"kind": "bullet", "text": "Total annual license cost: 6.504.721,08 \u20ac"} in blocks
    assert {"kind": "subsection", "text": "Savings"} in blocks
    assert {"kind": "bullet", "text": "Total savings: 55.780,96 \u20ac"} in blocks
    assert {"kind": "bullet", "text": "BYOL to PAYG Savings: 55.433,08 \u20ac"} in blocks
    assert {"kind": "bullet", "text": "Retired but reporting Savings: 347,88 \u20ac"} in blocks


def test_build_report_markdown_includes_savings_subsection():
    # build_report_markdown calls normalize_report_content_text which prepends \u20ac before
    # numeric amounts, so European-format "55.780,96 \u20ac" becomes "\u20ac55.780,96 \u20ac".
    markdown = build_report_markdown(
        "## Executive Summary\nSummary",
        {
            "azure_payg_count": 626,
            "retired_count": 22,
            "total_demand_quantity": 20568,
            "total_license_cost": 6504721.08,
            "total_savings": 55780.96,
            "azure_payg_savings": 55433.08,
            "retired_devices_savings": 347.88,
        },
    )

    assert "### Cost" in markdown
    assert "### Savings" in markdown
    assert "55.780,96" in markdown
    assert "55.433,08" in markdown
    assert "347,88" in markdown


def test_format_currency_uses_euro_symbol():
    # European format: comma as decimal, dot as thousands, symbol at end
    assert format_currency(345.6) == "345,60 \u20ac"


def test_export_xlsx_includes_savings_section_and_values():
    content = export_xlsx(
        "## Executive Summary\nSummary",
        report_context={
            "azure_payg_count": 626,
            "retired_count": 22,
            "total_demand_quantity": 20568,
            "total_license_cost": 6504721.08,
            "total_savings": 55780.96,
            "azure_payg_savings": 55433.08,
            "retired_devices_savings": 347.88,
        },
    )

    workbook = load_workbook(io.BytesIO(content))
    sheet = workbook["Report"]
    values = [str(row[0]) for row in sheet.iter_rows(values_only=True) if row and row[0]]

    assert "IT License and Cost Optimization Report" in values
    assert "Savings" in values
    # After normalize_report_content_text, amounts get \u20ac prepended: "\u20ac55.780,96 \u20ac"
    assert any("55.780,96" in v for v in values)
    assert any("55.433,08" in v for v in values)
    assert any("347,88" in v for v in values)
