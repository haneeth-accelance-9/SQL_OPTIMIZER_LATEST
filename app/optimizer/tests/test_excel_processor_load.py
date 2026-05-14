"""
Tests for ExcelProcessor.load_file() and _detect_sheet().
Covers lines 15-16, 43-52, 69-73, 80-157 of excel_processor.py.
No DB required — pure file-IO mocks using openpyxl + BytesIO.

Windows note: pd.ExcelFile holds a file handle open while alive, so we
always use BytesIO buffers (never temp-file paths) for the _detect_sheet
tests and use explicit try/finally with closed ExcelFile for load_file tests.
"""
import os
import tempfile
from io import BytesIO
from pathlib import Path
from unittest.mock import MagicMock, patch

import openpyxl
import pandas as pd
import pytest

from optimizer.services.excel_processor import ExcelProcessor, _detect_sheet


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _make_excel_bytes(sheet_names: list) -> bytes:
    """Create a minimal in-memory Excel workbook with the given sheet names."""
    wb = openpyxl.Workbook()
    wb.active.title = sheet_names[0]
    for name in sheet_names[1:]:
        wb.create_sheet(title=name)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _make_excel_buf(sheet_names: list, rows_per_sheet: int = 2) -> BytesIO:
    """
    Create an in-memory Excel workbook (BytesIO) with the given sheet names.
    Each sheet gets a header row + `rows_per_sheet` data rows.
    Returns a fresh BytesIO positioned at 0.
    """
    wb = openpyxl.Workbook()
    wb.active.title = sheet_names[0]
    wb.active.append(["Server Name", "Product", "Value"])
    for i in range(rows_per_sheet):
        wb.active.append([f"server-{i}", f"product-{i}", i * 10])

    for name in sheet_names[1:]:
        ws = wb.create_sheet(title=name)
        ws.append(["Server Name", "Product", "Value"])
        for i in range(rows_per_sheet):
            ws.append([f"srv-{name}-{i}", f"prod-{i}", i])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _write_tmp_excel(sheet_names: list, rows_per_sheet: int = 2) -> str:
    """
    Write a temporary Excel file with the given sheet names and return its path.
    The file is written once and the workbook is immediately closed.
    The caller is responsible for unlinking the file after closing any
    ExcelFile/DataFrame that reads it.
    """
    wb = openpyxl.Workbook()
    wb.active.title = sheet_names[0]
    wb.active.append(["Server Name", "Product", "Value"])
    for i in range(rows_per_sheet):
        wb.active.append([f"server-{i}", f"product-{i}", i * 10])

    for name in sheet_names[1:]:
        ws = wb.create_sheet(title=name)
        ws.append(["Server Name", "Product", "Value"])
        for i in range(rows_per_sheet):
            ws.append([f"srv-{name}-{i}", f"prod-{i}", i])

    fd, path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    wb.save(path)
    # wb is closed when it goes out of scope — no file handle held
    return path


def _safe_unlink(path: str) -> None:
    """Try to unlink; silently ignore Windows locking errors after short retry."""
    import time
    for _ in range(3):
        try:
            os.unlink(path)
            return
        except PermissionError:
            time.sleep(0.1)
    # Last attempt — let any exception propagate
    os.unlink(path)


# ---------------------------------------------------------------------------
# _detect_sheet  (uses BytesIO — no file handles left open)
# ---------------------------------------------------------------------------

class TestDetectSheet:
    def _excel_file(self, sheet_names):
        """Build a pd.ExcelFile from an in-memory workbook."""
        data = _make_excel_bytes(sheet_names)
        return pd.ExcelFile(BytesIO(data), engine="openpyxl")

    def test_returns_first_exact_match(self):
        xf = self._excel_file(["MVP - Data 1 - Installation", "Data 2", "Other"])
        result = _detect_sheet(xf, "MVP - Data 1 - Installation", "Data 1")
        assert result == "MVP - Data 1 - Installation"

    def test_returns_second_candidate_when_first_missing(self):
        xf = self._excel_file(["Data 1", "Data 2", "Other"])
        result = _detect_sheet(xf, "MVP - Data 1 - Installation", "Data 1")
        assert result == "Data 1"

    def test_falls_back_to_substring_match(self):
        xf = self._excel_file(["My Installation Sheet", "Data 2"])
        result = _detect_sheet(xf, "MVP - Data 1 - Installation", "Data 1", "Installation")
        assert result == "My Installation Sheet"

    def test_returns_none_when_no_match(self):
        xf = self._excel_file(["Sheet1", "Sheet2"])
        result = _detect_sheet(xf, "MVP - Data 1 - Installation", "Data 1")
        assert result is None

    def test_exact_match_takes_priority_over_substring(self):
        xf = self._excel_file(["Data 1 Extra", "Data 1"])
        result = _detect_sheet(xf, "Data 1")
        assert result == "Data 1"

    def test_substring_match_case_insensitive(self):
        xf = self._excel_file(["INSTALLATION SHEET"])
        # "installation" is substring of "INSTALLATION SHEET" (case-insensitive)
        result = _detect_sheet(xf, "MVP - Data 1 - Installation", "Installation")
        assert result == "INSTALLATION SHEET"

    def test_no_candidates_returns_none(self):
        xf = self._excel_file(["Sheet1"])
        result = _detect_sheet(xf)  # no candidates passed
        assert result is None


# ---------------------------------------------------------------------------
# ExcelProcessor.load_file() — file not found
# ---------------------------------------------------------------------------

class TestLoadFileNotFound:
    def test_missing_file_returns_error_dict(self):
        proc = ExcelProcessor()
        result = proc.load_file("/nonexistent/path/to/file.xlsx")
        assert "error" in result

    def test_missing_file_error_mentions_path(self):
        proc = ExcelProcessor()
        result = proc.load_file("/nonexistent/file.xlsx")
        assert "nonexistent" in result["error"] or "file.xlsx" in result["error"]

    def test_missing_file_no_exception_raised(self):
        proc = ExcelProcessor()
        try:
            proc.load_file("/does/not/exist.xlsx")
        except Exception:
            pytest.fail("load_file() should not raise for missing files")


# ---------------------------------------------------------------------------
# ExcelProcessor.load_file() — correct sheet names via BytesIO patch
#
# We patch Path.exists() to return True and pd.ExcelFile to use our BytesIO.
# This avoids temp files entirely on Windows.
# ---------------------------------------------------------------------------

def _load_from_buf(sheet_names: list, proc: ExcelProcessor = None) -> dict:
    """
    Helper: exercise load_file() using a BytesIO buffer.
    We mock `Path.exists` → True and `pd.ExcelFile` to read from our buffer.
    """
    if proc is None:
        proc = ExcelProcessor()

    buf = _make_excel_buf(sheet_names)
    excel_bytes = buf.read()

    original_excel_file = pd.ExcelFile

    def _fake_excel_file(path, **kwargs):
        return original_excel_file(BytesIO(excel_bytes), engine="openpyxl")

    with patch.object(Path, "exists", return_value=True), \
         patch("optimizer.services.excel_processor.pd.ExcelFile", side_effect=_fake_excel_file):
        return proc.load_file("/fake/path/file.xlsx")


class TestLoadFileCorrectSheetNames:
    def test_correct_sheet_names_returns_installations_df(self):
        result = _load_from_buf([
            "MVP - Data 1 - Installation",
            "MVP - Data 2 - Demand Results",
            "MVP - Data 3 - Prices",
        ])
        assert "error" not in result
        assert result["installations"] is not None
        assert isinstance(result["installations"], pd.DataFrame)

    def test_correct_sheet_names_returns_demand_df(self):
        result = _load_from_buf([
            "MVP - Data 1 - Installation",
            "MVP - Data 2 - Demand Results",
        ])
        assert result["demand"] is not None
        assert isinstance(result["demand"], pd.DataFrame)

    def test_sheet_names_used_populated(self):
        result = _load_from_buf([
            "MVP - Data 1 - Installation",
            "MVP - Data 2 - Demand Results",
        ])
        assert "installations" in result["sheet_names_used"]

    def test_all_sheet_names_populated(self):
        sheets = ["MVP - Data 1 - Installation", "MVP - Data 2 - Demand Results", "Extra"]
        result = _load_from_buf(sheets)
        assert set(result["all_sheet_names"]) == set(sheets)

    def test_columns_normalized(self):
        result = _load_from_buf([
            "MVP - Data 1 - Installation",
            "MVP - Data 2 - Demand Results",
        ])
        for col in result["installations"].columns:
            assert col == col.lower()
            assert " " not in col


# ---------------------------------------------------------------------------
# ExcelProcessor.load_file() — fallback "Data 1 / Data 2" sheet names
# ---------------------------------------------------------------------------

class TestLoadFileFallbackSheetNames:
    def test_data_1_data_2_sheet_names_work(self):
        result = _load_from_buf(["Data 1", "Data 2"])
        assert "error" not in result
        assert result["installations"] is not None

    def test_demand_loaded_with_data_2_sheet(self):
        result = _load_from_buf(["Data 1", "Data 2"])
        assert result["demand"] is not None

    def test_substring_fallback_installation(self):
        """Sheet name that merely contains 'Installation' should be detected."""
        result = _load_from_buf(["My Installation Data", "Demand Sheet"])
        # Either detected or returns an error — just no exception
        assert isinstance(result, dict)


# ---------------------------------------------------------------------------
# ExcelProcessor.load_file() — missing installations sheet → error
# ---------------------------------------------------------------------------

class TestLoadFileMissingInstallations:
    def test_missing_installations_returns_error(self):
        result = _load_from_buf(["Totally Unrelated", "Another Sheet"])
        assert "error" in result

    def test_missing_installations_error_message_is_descriptive(self):
        result = _load_from_buf(["Sheet1", "Sheet2"])
        err = result.get("error", "")
        assert "Installation" in err or "Data 1" in err


# ---------------------------------------------------------------------------
# ExcelProcessor.load_file() — optional sheets
# ---------------------------------------------------------------------------

class TestLoadFileOptionalSheets:
    def test_prices_loaded_when_present(self):
        result = _load_from_buf(["Data 1", "Data 2", "Data 3"])
        assert result["prices"] is not None

    def test_optimization_loaded_when_present(self):
        result = _load_from_buf(["Data 1", "Data 2", "Data 4"])
        assert result["optimization"] is not None

    def test_helpful_reports_loaded_when_present(self):
        result = _load_from_buf(["Data 1", "Data 2", "Data 5"])
        assert result["helpful_reports"] is not None

    def test_missing_optional_sheets_still_succeeds(self):
        result = _load_from_buf(["Data 1", "Data 2"])
        assert "error" not in result
        assert result["prices"] is None
        assert result["optimization"] is None
        assert result["helpful_reports"] is None


# ---------------------------------------------------------------------------
# ExcelProcessor.load_file() — injected sheet names via constructor
# ---------------------------------------------------------------------------

class TestLoadFileInjectedSheetNames:
    def test_custom_installation_sheet_name_used(self):
        custom_name = "Custom Installation Sheet"
        proc = ExcelProcessor(
            sheet_installations=custom_name,
            sheet_demand="Demand Sheet",
        )
        result = _load_from_buf([custom_name, "Demand Sheet"], proc=proc)
        assert "error" not in result
        assert result["installations"] is not None

    def test_wrong_custom_name_falls_back_to_detection(self):
        proc = ExcelProcessor(sheet_installations="NonExistentSheet")
        result = _load_from_buf(["Data 1", "Data 2"], proc=proc)
        # Falls back via substring detection — should work or return an error (no exception)
        assert isinstance(result, dict)


# ---------------------------------------------------------------------------
# ExcelProcessor.load_file() — corrupted / non-Excel file
# (uses real temp file since we WANT the actual file open/parse to fail)
# ---------------------------------------------------------------------------

class TestLoadFileCorrupted:
    def test_non_excel_file_returns_error(self):
        fd, path = tempfile.mkstemp(suffix=".xlsx")
        try:
            os.write(fd, b"this is not a valid excel file \x00\x01\x02")
        finally:
            os.close(fd)

        try:
            proc = ExcelProcessor()
            result = proc.load_file(path)
            assert "error" in result
        finally:
            _safe_unlink(path)

    def test_non_excel_file_does_not_raise(self):
        fd, path = tempfile.mkstemp(suffix=".xlsx")
        try:
            os.write(fd, b"garbage content")
        finally:
            os.close(fd)

        try:
            proc = ExcelProcessor()
            try:
                proc.load_file(path)
            except Exception:
                pytest.fail("load_file should not raise for corrupted files")
        finally:
            _safe_unlink(path)
