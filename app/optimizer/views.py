"""
Views for SQL License Optimizer: upload, process, results, dashboard, report.
All optimizer views require authentication (enterprise security).
Uses AnalysisSession for persistence and TTL; session stores only analysis_id.
"""
import logging
import os
import re
import uuid


def _eu_currency(value):
    """Format a number as European currency: 1.234.567,89 €"""
    try:
        number = float(value)
    except (TypeError, ValueError):
        return str(value)
    formatted = f"{number:,.2f}"
    formatted = formatted.replace(",", "X").replace(".", ",").replace("X", ".")
    return f"{formatted} €"
from django.conf import settings
from django.shortcuts import render, redirect
from django.urls import reverse
from django.http import HttpResponse, JsonResponse
from django.views.decorators.http import require_GET, require_http_methods
from django.contrib.auth import logout as auth_logout
from django.views.decorators.csrf import csrf_protect
from django.contrib.auth.decorators import login_required
from django.contrib.auth.views import LoginView
from django.contrib.auth.forms import AuthenticationForm
from django.contrib import messages
from django.utils import timezone

import pandas as pd
from optimizer.models import AnalysisSession, UserProfile
from optimizer.forms import SignUpForm, UserProfileForm
from optimizer.services.analysis_service import run_analysis, build_dashboard_context
from optimizer.services.analysis_logs import build_analysis_summary_metrics, get_user_analysis_logs
from optimizer.services.report_export import (
    export_pdf,
    export_docx,
    export_xlsx,
    normalize_report_content_text,
    build_report_markdown,
)

try:
    from optimizer.services.plotly_charts import get_all_plotly_specs
except ImportError:
    get_all_plotly_specs = None


def _make_json_serializable(obj):
    """Convert numpy/pandas types to native Python so result_data passes SQLite JSON_VALID."""
    if isinstance(obj, dict):
        return {k: _make_json_serializable(v) for k, v in obj.items()}
    if isinstance(obj, (list, tuple)):
        return [_make_json_serializable(v) for v in obj]
    if hasattr(obj, "tolist") and not isinstance(obj, (str, bytes)):  # numpy array
        return [_make_json_serializable(v) for v in obj]
    if hasattr(obj, "item"):  # numpy scalar (0-d array)
        try:
            return obj.item()
        except (ValueError, AttributeError):
            return [_make_json_serializable(v) for v in obj]
    try:
        if pd.isna(obj):
            return None
    except (ValueError, TypeError):
        pass
    if isinstance(obj, (int, float, str, bool, type(None))):
        return obj
    if isinstance(obj, (pd.Timestamp,)):
        return obj.isoformat()
    return str(obj)

try:
    from optimizer.services.chart_generator import generate_all_charts
except ImportError:
    generate_all_charts = None

logger = logging.getLogger(__name__)

REPORT_FORMAT_ALIASES = {
    "excel": "xlsx",
}
ALLOWED_REPORT_FORMATS = frozenset({"pdf", "docx", "xlsx", *REPORT_FORMAT_ALIASES.keys()})
ALLOWED_RULE_IDS = frozenset({"rule1", "rule2"})
PAYG_ZONE_BREAKDOWN_LABELS = ["Public Cloud", "Private Cloud AVS"]
POST_LOGIN_RESULTS_QUERY = "rs3_workload=ALL&rs3_filter=PROD_CPU_Optimization&rs3_page=1"
POST_LOGIN_RESULTS_FRAGMENT = "dashboard"


def _build_profile_initials(user) -> str:
    """Build a short initials badge from the user's name, falling back to username."""
    first = (getattr(user, "first_name", "") or "").strip()
    last = (getattr(user, "last_name", "") or "").strip()
    if first or last:
        initials = f"{first[:1]}{last[:1]}".strip()
    else:
        username = (getattr(user, "username", "") or "").strip()
        parts = [part for part in re.split(r"[\s._-]+", username) if part]
        initials = "".join(part[:1] for part in parts[:2]) or username[:2]
    return (initials or "U").upper()


def _build_profile_context(user, profile, form=None):
    """Build the display context expected by the profile page template."""
    full_name = f"{(user.first_name or '').strip()} {(user.last_name or '').strip()}".strip()
    display_name = full_name or user.username
    return {
        "title": "Profile",
        "profile_form": form or UserProfileForm(instance=profile, user=user),
        "profile_display_name": display_name,
        "profile_initials": _build_profile_initials(user),
        "profile_email": user.email or "Not provided",
        "profile_username": user.username,
        "profile_team_name": profile.team_name or "Not provided",
        "profile_image_url": profile.image_url or "",
        "profile_first_name": user.first_name or "Not provided",
        "profile_last_name": user.last_name or "Not provided",
    }


def _get_or_create_user_profile(user):
    """Return the persisted profile row for the authenticated user."""
    profile, _ = UserProfile.objects.get_or_create(user=user)
    return profile


def _build_post_login_redirect_url() -> str:
    """Return the default authenticated landing page for login flows."""
    return f"{reverse('optimizer:dashboard')}#{POST_LOGIN_RESULTS_FRAGMENT}"




def _get_analysis_record(request):
    """Load the current persisted analysis record and enforce ownership/TTL checks."""
    analysis_id = request.session.get("optimizer_analysis_id")
    if not analysis_id:
        return None, redirect("optimizer:dashboard")
    analysis = AnalysisSession.objects.filter(pk=analysis_id).first()
    if not analysis:
        messages.info(request, "Analysis not found. Please upload a new file.")
        return None, redirect("optimizer:dashboard")
    if analysis.user_id and analysis.user_id != request.user.id:
        return None, redirect("optimizer:dashboard")
    ttl = getattr(settings, "OPTIMIZER_ANALYSIS_TTL_SECONDS", 86400)
    if ttl > 0 and analysis.created_at:
        from datetime import timedelta
        if timezone.now() - analysis.created_at > timedelta(seconds=ttl):
            messages.info(request, "This analysis has expired. Please upload a new file.")
            return None, redirect("optimizer:dashboard")
    return analysis, None


def _normalize_analysis_context(analysis):
    """Return a safe, normalized analysis context from persisted result data."""
    raw_context = analysis.result_data if isinstance(analysis.result_data, dict) else {}
    context = dict(raw_context)
    context["rule_results"] = context.get("rule_results") or {}
    context["license_metrics"] = context.get("license_metrics") or {}
    context["file_name"] = context.get("file_name") or analysis.file_name or ""
    context["sheet_names_used"] = context.get("sheet_names_used") or {}
    context["total_devices_analyzed"] = int(context.get("total_devices_analyzed", 0) or 0)
    context["report_text"] = context.get("report_text") or ""
    context["report_used_fallback"] = bool(context.get("report_used_fallback", False))
    context["cost_reduction_ai_recommendations"] = context.get("cost_reduction_ai_recommendations") or ""

    dashboard_defaults = build_dashboard_context(context)
    context["rule_wise_savings"] = context.get("rule_wise_savings") or dashboard_defaults.get("rule_wise_savings", {})
    context["scenario_wise_savings"] = context.get("scenario_wise_savings") or dashboard_defaults.get("scenario_wise_savings", {})
    context["total_savings"] = float(context.get("total_savings", dashboard_defaults.get("total_savings", 0)) or 0)
    return context


def _get_analysis_context(request):
    """
    Load analysis context from DB by session's analysis_id. Check TTL and ownership.
    Returns (context dict, None) or (None, redirect_response).
    """
    analysis, redir = _get_analysis_record(request)
    if redir is not None:
        return None, redir
    raw_context = analysis.result_data if isinstance(analysis.result_data, dict) else {}
    if not raw_context:
        return None, redirect("optimizer:dashboard")
    return _normalize_analysis_context(analysis), None


def _build_report_render_context(context):
    dashboard_context = build_dashboard_context(context)
    return {
        "azure_payg_count": context.get("rule_results", {}).get("azure_payg_count", 0),
        "retired_count": context.get("rule_results", {}).get("retired_count", 0),
        "total_demand_quantity": context.get("license_metrics", {}).get("total_demand_quantity", 0),
        "total_license_cost": context.get("license_metrics", {}).get("total_license_cost", 0),
        "by_product": context.get("license_metrics", {}).get("by_product", []),
        "rule_wise_savings": dashboard_context.get("rule_wise_savings", {}),
        "total_savings": dashboard_context.get("total_savings", 0),
        "azure_payg_savings": dashboard_context.get("azure_payg_savings", 0),
        "retired_devices_savings": dashboard_context.get("retired_devices_savings", 0),
    }


def _build_rightsizing_filter_funnel() -> dict:
    """
    Compute detailed filter-funnel counts for all UC3 rules (UC3.1–3.5 + Physical).
    Called once per report page load to power the Live Metrics tab.
    Returns a flat dict of intermediate and final counts.
    """
    try:
        from optimizer.services.db_analysis_service import _build_rightsizing_df
        from optimizer.rules.rightsizing import (
            NON_PROD_ENVS, COL_CRITICALITY,
            ALL_CRITICAL_VALS, CRITICAL_VALS, LC_CRITICAL_VALS,
            find_cpu_rightsizing_optimizations, find_ram_rightsizing_optimizations,
            find_criticality_cpu_downsize_optimizations, find_criticality_cpu_upsize_optimizations,
            find_criticality_ram_downsize_optimizations, find_criticality_ram_upsize_optimizations,
            find_lifecycle_risk_flags,
        )
        df = _build_rightsizing_df()
    except Exception:
        return {}

    if df.empty:
        return {}

    total_input = len(df)

    prod_mask  = ~df["Environment"].isin(NON_PROD_ENVS) if "Environment" in df.columns else pd.Series([True] * len(df), index=df.index)
    prod_df    = df[prod_mask]
    nonprod_df = df[~prod_mask]
    prod_total    = len(prod_df)
    nonprod_total = len(nonprod_df)

    # ── UC3.1 CPU ──────────────────────────────────────────────────────────────
    cpu_prod_f1   = prod_df[prod_df["Avg_CPU_12m"].fillna(100) < 15]
    cpu_prod_f2   = cpu_prod_f1[cpu_prod_f1["Peak_CPU_12m"].fillna(100) <= 70]
    cpu_prod_f3   = cpu_prod_f2[cpu_prod_f2["Current_vCPU"].fillna(0) >= 4]
    cpu_nonprod_f1 = nonprod_df[nonprod_df["Avg_CPU_12m"].fillna(100) < 25]
    cpu_nonprod_f2 = cpu_nonprod_f1[cpu_nonprod_f1["Peak_CPU_12m"].fillna(100) <= 80]
    cpu_nonprod_f3 = cpu_nonprod_f2[cpu_nonprod_f2["Current_vCPU"].fillna(0) >= 4]

    cpu_final_df = find_cpu_rightsizing_optimizations(df)
    if not cpu_final_df.empty:
        cpu_final_df = cpu_final_df[
            cpu_final_df["Recommended_vCPU"].notna() &
            (cpu_final_df["Recommended_vCPU"] != cpu_final_df["Current_vCPU"])
        ]
        cpu_final_df = cpu_final_df[
            (cpu_final_df["Current_vCPU"] >= 4) &
            (cpu_final_df["Recommended_vCPU"] >= 4)
        ]
    if not cpu_final_df.empty and "Env_Type" in cpu_final_df.columns:
        cpu_prod_final    = int((cpu_final_df["Env_Type"] == "PROD").sum())
        cpu_nonprod_final = int((cpu_final_df["Env_Type"] == "NON-PROD").sum())
    else:
        cpu_prod_final    = len(cpu_prod_f3)
        cpu_nonprod_final = len(cpu_nonprod_f3)

    # ── UC3.2 RAM ──────────────────────────────────────────────────────────────
    ram_prod_f1   = prod_df[prod_df["Avg_FreeMem_12m"].fillna(0) >= 35]
    ram_prod_f2   = ram_prod_f1[ram_prod_f1["Min_FreeMem_12m"].fillna(0) >= 20]
    ram_prod_f3   = ram_prod_f2[ram_prod_f2["Current_RAM_GiB"].fillna(0) > 8]
    ram_nonprod_f1 = nonprod_df[nonprod_df["Avg_FreeMem_12m"].fillna(0) >= 30]
    ram_nonprod_f2 = ram_nonprod_f1[ram_nonprod_f1["Min_FreeMem_12m"].fillna(0) >= 15]
    ram_nonprod_f3 = ram_nonprod_f2[ram_nonprod_f2["Current_RAM_GiB"].fillna(0) > 4]

    ram_final_df = find_ram_rightsizing_optimizations(df)
    if not ram_final_df.empty and "Env_Type" in ram_final_df.columns:
        ram_prod_final    = int((ram_final_df["Env_Type"] == "PROD").sum())
        ram_nonprod_final = int((ram_final_df["Env_Type"] == "NON-PROD").sum())
    else:
        ram_prod_final    = len(ram_prod_f3)
        ram_nonprod_final = len(ram_nonprod_f3)

    # ── UC3.3 / UC3.4 Criticality ─────────────────────────────────────────────
    has_criticality = COL_CRITICALITY in df.columns
    if has_criticality:
        crit_all_mask  = df[COL_CRITICALITY].isin(ALL_CRITICAL_VALS)
        crit_bus_mask  = df[COL_CRITICALITY].isin(CRITICAL_VALS)
        crit_all_count = int(crit_all_mask.sum())
        crit_bus_count = int(crit_bus_mask.sum())
        crit_cpu_dn_f2    = int(df[crit_all_mask]["Avg_CPU_12m"].fillna(100).lt(10).sum())
        crit_cpu_up_f2    = int(df[crit_bus_mask]["Avg_CPU_12m"].fillna(0).gt(80).sum())
        crit_ram_dn_f2    = int(df[crit_all_mask]["Avg_FreeMem_12m"].fillna(0).gt(80).sum())
        crit_ram_up_f2    = int(df[crit_bus_mask]["Avg_FreeMem_12m"].fillna(100).lt(20).sum())
        crit_cpu_dn_final = len(find_criticality_cpu_downsize_optimizations(df))
        crit_cpu_up_final = len(find_criticality_cpu_upsize_optimizations(df))
        crit_ram_dn_final = len(find_criticality_ram_downsize_optimizations(df))
        crit_ram_up_final = len(find_criticality_ram_upsize_optimizations(df))

        # Combined + post-rule filters matching compute_rightsizing_metrics()
        _crit_cpu_combined = pd.concat(
            [find_criticality_cpu_downsize_optimizations(df),
             find_criticality_cpu_upsize_optimizations(df)],
            ignore_index=True,
        )
        if not _crit_cpu_combined.empty:
            _crit_cpu_combined = _crit_cpu_combined[
                _crit_cpu_combined["Recommended_vCPU"].notna() &
                (_crit_cpu_combined["Recommended_vCPU"] != _crit_cpu_combined["Current_vCPU"])
            ]
            _crit_cpu_combined = _crit_cpu_combined[
                (_crit_cpu_combined["Current_vCPU"] >= 4) &
                (_crit_cpu_combined["Recommended_vCPU"] >= 4)
            ]
        crit_cpu_filtered_final = len(_crit_cpu_combined)

        _crit_ram_combined = pd.concat(
            [find_criticality_ram_downsize_optimizations(df),
             find_criticality_ram_upsize_optimizations(df)],
            ignore_index=True,
        )
        if not _crit_ram_combined.empty:
            _crit_ram_combined = _crit_ram_combined[
                _crit_ram_combined["Recommended_RAM_GiB"].notna() &
                (_crit_ram_combined["Recommended_RAM_GiB"] != _crit_ram_combined["Current_RAM_GiB"])
            ]
            _crit_ram_combined = _crit_ram_combined[
                (_crit_ram_combined["Current_RAM_GiB"] >= 8) &
                (_crit_ram_combined["Recommended_RAM_GiB"] >= 8)
            ]
        crit_ram_filtered_final = len(_crit_ram_combined)
    else:
        crit_all_count = crit_bus_count = 0
        crit_cpu_dn_f2 = crit_cpu_up_f2 = crit_ram_dn_f2 = crit_ram_up_f2 = 0
        crit_cpu_dn_final = crit_cpu_up_final = crit_ram_dn_final = crit_ram_up_final = 0
        crit_cpu_filtered_final = crit_ram_filtered_final = 0

    # ── UC3.5 Lifecycle ────────────────────────────────────────────────────────
    if has_criticality:
        lc_f1_df    = df[df[COL_CRITICALITY].isin(LC_CRITICAL_VALS)]
        lc_f2_df    = lc_f1_df[lc_f1_df["Peak_CPU_12m"].fillna(0) > 95]
        lc_f3_df    = lc_f2_df[lc_f2_df["Min_FreeMem_12m"].fillna(100) < 5]
        lc_f1_count = len(lc_f1_df)
        lc_f2_count = len(lc_f2_df)
        lc_f3_count = len(lc_f3_df)
        lc_final    = len(find_lifecycle_risk_flags(df))
    else:
        lc_f1_count = lc_f2_count = lc_f3_count = lc_final = 0

    # ── Physical ───────────────────────────────────────────────────────────────
    virt_col = next((c for c in ("Is Virtual?", "is_virtual", "IsVirtual") if c in df.columns), None)
    if virt_col:
        phys_mask      = df[virt_col].astype(str).str.strip().str.lower() == "false"
        physical_count = int(phys_mask.sum())
        virtual_count  = total_input - physical_count
    else:
        physical_count = virtual_count = 0

    return {
        "total_input":    total_input,
        "prod_total":     prod_total,
        "nonprod_total":  nonprod_total,
        # UC3.1 CPU
        "cpu_prod_f1":    len(cpu_prod_f1),
        "cpu_prod_f2":    len(cpu_prod_f2),
        "cpu_prod_f3":    len(cpu_prod_f3),
        "cpu_prod_final": cpu_prod_final,
        "cpu_nonprod_f1":    len(cpu_nonprod_f1),
        "cpu_nonprod_f2":    len(cpu_nonprod_f2),
        "cpu_nonprod_f3":    len(cpu_nonprod_f3),
        "cpu_nonprod_final": cpu_nonprod_final,
        "cpu_final": cpu_prod_final + cpu_nonprod_final,
        # UC3.2 RAM
        "ram_prod_f1":       len(ram_prod_f1),
        "ram_prod_f2":       len(ram_prod_f2),
        "ram_prod_eligible": len(ram_prod_f3),
        "ram_prod_final":    ram_prod_final,
        "ram_nonprod_f1":       len(ram_nonprod_f1),
        "ram_nonprod_f2":       len(ram_nonprod_f2),
        "ram_nonprod_eligible": len(ram_nonprod_f3),
        "ram_nonprod_final":    ram_nonprod_final,
        "ram_final": ram_prod_final + ram_nonprod_final,
        # UC3.3 Crit CPU
        "crit_all_count":     crit_all_count,
        "crit_bus_count":     crit_bus_count,
        "crit_cpu_dn_f2":     crit_cpu_dn_f2,
        "crit_cpu_up_f2":     crit_cpu_up_f2,
        "crit_cpu_dn_final":  crit_cpu_dn_final,
        "crit_cpu_up_final":  crit_cpu_up_final,
        "crit_cpu_final":     crit_cpu_filtered_final,
        # UC3.4 Crit RAM
        "crit_ram_dn_f2":    crit_ram_dn_f2,
        "crit_ram_up_f2":    crit_ram_up_f2,
        "crit_ram_dn_final": crit_ram_dn_final,
        "crit_ram_up_final": crit_ram_up_final,
        "crit_ram_final":    crit_ram_filtered_final,
        # UC3.5 Lifecycle
        "lc_f1":    lc_f1_count,
        "lc_f2":    lc_f2_count,
        "lc_f3":    lc_f3_count,
        "lc_final": lc_final,
        # Physical
        "physical_count": physical_count,
        "virtual_count":  virtual_count,
    }


def _get_db_context_for_report():
    """
    Build a fully-populated report context from live DB data.
    Generates report text (AI if enabled, fallback otherwise) and flattens
    savings keys so _build_report_render_context works correctly.
    """
    from optimizer.services.db_analysis_service import compute_live_db_metrics
    from optimizer.services.ai_report_generator import generate_report_text, get_fallback_report

    context = compute_live_db_metrics()

    # Flatten savings so _build_report_render_context and agent payload builder can read them
    dash = build_dashboard_context(context)
    context["azure_payg_savings"] = dash.get("azure_payg_savings", 0)
    context["retired_devices_savings"] = dash.get("retired_devices_savings", 0)
    context["rightsizing_savings"] = dash.get("rightsizing_savings", 0)

    rr = context.get("rule_results", {})
    lm = context.get("license_metrics", {})
    rs = context.get("rightsizing", {})
    report_context = {
        "azure_payg_count": rr.get("azure_payg_count", 0),
        "retired_count": rr.get("retired_count", 0),
        "total_demand_quantity": lm.get("total_demand_quantity", 0),
        "total_license_cost": lm.get("total_license_cost", 0),
        "by_product": lm.get("by_product", []),
        "demand_row_count": lm.get("demand_row_count", 0),
        # UC3 counts for report narrative
        "cpu_count":          rs.get("cpu_count", 0),
        "ram_count":          rs.get("ram_count", 0),
        "cpu_prod_count":     rs.get("cpu_prod_count", 0),
        "cpu_nonprod_count":  rs.get("cpu_nonprod_count", 0),
        "ram_prod_count":     rs.get("ram_prod_count", 0),
        "ram_nonprod_count":  rs.get("ram_nonprod_count", 0),
        "crit_cpu_count":     rs.get("crit_cpu_count", 0),
        "crit_ram_count":     rs.get("crit_ram_count", 0),
        "lifecycle_count":    rs.get("lifecycle_count", 0),
        "physical_count":     rs.get("physical_count", 0),
        "rightsizing_savings": context.get("rightsizing_savings", 0),
    }

    report_text = None
    if getattr(settings, "OPTIMIZER_AI_REPORT_ENABLED", True):
        try:
            report_text = generate_report_text(report_context)
        except Exception as e:
            logger.warning("AI report generation failed: %s", e)

    used_fallback = not bool(report_text)
    if not report_text:
        report_text = get_fallback_report(report_context)

    # Filter funnel for Live Metrics tab
    context["filter_funnel"] = _build_rightsizing_filter_funnel()

    context["report_text"] = report_text or ""
    context["report_used_fallback"] = used_fallback
    context["title"] = "IT License and Cost Optimization Report"
    context["data_source"] = "database"
    return context


def _get_page_number(request, param_name, default=1):
    """Parse a positive integer query param with a safe fallback."""
    try:
        return max(1, int(request.GET.get(param_name, default)))
    except (TypeError, ValueError):
        return default


def _format_metric_label(metric_name):
    metric_name = str(metric_name or "").strip()
    if not metric_name:
        return ""
    if metric_name == "database_size_mib":
        return "Database Size Mmib"
    return " ".join(part.capitalize() for part in metric_name.split("_") if part)


RS3_CPU_OPTIMIZATION_COLUMNS = [
    "server_name",
    "product_family",
    "product_name",
    "product_description",
    "Env_Type",
    "Avg_CPU_12m",
    "Peak_CPU_12m",
    "Current_vCPU",
    "Potential_vCPU_Reduction",
]

RS3_CPU_RIGHTSIZING_COLUMNS = [
    "server_name",
    "product_family",
    "product_name",
    "product_description",
    "Env_Type",
    "Avg_CPU_12m",
    "Peak_CPU_12m",
    "Current_vCPU",
    "Recommended_vCPU",
    "CPU_Recommendation",
]

RS3_RAM_OPTIMIZATION_COLUMNS = [
    "server_name",
    "product_family",
    "product_name",
    "product_description",
    "Env_Type",
    "Avg_FreeMem_12m",
    "Min_FreeMem_12m",
    "Current_RAM_GiB",
    "Potential_RAM_Reduction_GiB",
]

RS3_RAM_RIGHTSIZING_COLUMNS = [
    "server_name",
    "product_family",
    "product_name",
    "product_description",
    "Env_Type",
    "Avg_FreeMem_12m",
    "Min_FreeMem_12m",
    "Current_RAM_GiB",
    "Recommended_RAM_GiB",
    "RAM_Recommendation",
]

RS3_CPU_RECOMMENDATION_COLUMNS = [
    "server_name",
    "product_family",
    "product_name",
    "product_description",
    "Env_Type",
    "Current_vCPU",
    "Recommended_vCPU",
    "CPU_Recommendation",
]

RS3_RAM_RECOMMENDATION_COLUMNS = [
    "server_name",
    "product_family",
    "product_name",
    "product_description",
    "Env_Type",
    "Current_RAM_GiB",
    "Recommended_RAM_GiB",
    "RAM_Recommendation",
]

RS3_WORKLOAD_DEFAULT = "ALL"
RS3_DEFAULT_FILTER_BY_WORKLOAD = {
    "CPU": "PROD_CPU_Rightsizing",
    "RAM": "PROD_RAM_Rightsizing",
}
RS3_SCREEN_FILTER_OPTIONS = {
    "CPU": [
        "PROD_CPU_Rightsizing",
        "NONPROD_CPU_Rightsizing",
    ],
    "RAM": [
        "PROD_RAM_Rightsizing",
        "NONPROD_RAM_Rightsizing",
    ],
}

RS3_CPU_FILTER_ALIASES = {
    "PROD_CPU_Optimization": "PROD_CPU_Rightsizing",
    "PROD_CPU_Recommendation": "PROD_CPU_Rightsizing",
    "NONPROD_CPU_Optimization": "NONPROD_CPU_Rightsizing",
    "NONPROD_CPU_Recommendation": "NONPROD_CPU_Rightsizing",
}
RS3_RAM_FILTER_ALIASES = {
    "PROD_RAM_Optimization": "PROD_RAM_Rightsizing",
    "PROD_RAM_Recommendation": "PROD_RAM_Rightsizing",
    "NONPROD_RAM_Optimization": "NONPROD_RAM_Rightsizing",
    "NONPROD_RAM_Recommendation": "NONPROD_RAM_Rightsizing",
}
RS3_DOWNLOAD_SHEET_KEYS = tuple(
    RS3_SCREEN_FILTER_OPTIONS["CPU"] + RS3_SCREEN_FILTER_OPTIONS["RAM"]
)
RS3_API_HOSTING_ZONE_OPTIONS = [
    "Functional Site",
    "none",
    "Private Cloud",
    "Private Cloud AVS",
    "Public Cloud",
    "Remote Site",
]
RS3_API_INSTALLED_STATUS_USU_OPTIONS = ["Installed", "Retired"]
RS3_API_DEFAULT_PAGE_SIZE = 25
RS3_API_MAX_PAGE_SIZE = 200
RS3_API_CPU_COLUMNS = [
    {"key": "server_name", "label": "Server Name"},
    {"key": "product_family", "label": "Product Family"},
    {"key": "product_group", "label": "Product Group"},
    {"key": "product_description", "label": "Product Description"},
    {"key": "product_name", "label": "Product Name"},
    {"key": "env_type", "label": "Env Type"},
    {"key": "avg_cpu_12m", "label": "Avg CPU 12m"},
    {"key": "peak_cpu_12m", "label": "Peak CPU 12m"},
    {"key": "current_vcpu", "label": "Current VCPU"},
    {"key": "recommended_vcpu", "label": "Recommended VCPU"},
    {"key": "cpu_recommendation", "label": "CPU Recommendation"},
    {"key": "cost_savings_eur", "label": "Cost Savings Eur"},
]
RS3_API_RAM_COLUMNS = [
    {"key": "server_name", "label": "Server Name"},
    {"key": "product_family", "label": "Product Family"},
    {"key": "product_group", "label": "Product Group"},
    {"key": "product_description", "label": "Product Description"},
    {"key": "product_name", "label": "Product Name"},
    {"key": "env_type", "label": "Env Type"},
    {"key": "avg_free_mem_12m", "label": "Avg Freemem 12m"},
    {"key": "min_free_mem_12m", "label": "Min Freemem 12m"},
    {"key": "current_ram_gib", "label": "Current RAM GiB"},
    {"key": "recommended_ram_gib", "label": "Recommended RAM GiB"},
    {"key": "ram_recommendation", "label": "RAM Recommendation"},
    {"key": "cost_savings_eur", "label": "Cost Savings Eur"},
]
RS3_API_CPU_SORT_FIELD_MAP = {
    "server_name": "server_name",
    "environment": "Environment",
    "env_type": "Env_Type",
    "hosting_zone": "hosting_zone",
    "installed_status_usu": "installed_status_usu",
    "avg_cpu_12m": "Avg_CPU_12m",
    "peak_cpu_12m": "Peak_CPU_12m",
    "current_vcpu": "Current_vCPU",
    "recommended_vcpu": "Recommended_vCPU",
    "potential_vcpu_reduction": "Potential_vCPU_Reduction",
    "cpu_recommendation": "CPU_Recommendation",
    "cost_savings_eur": "Cost_Savings_EUR",
}
RS3_API_RAM_SORT_FIELD_MAP = {
    "server_name": "server_name",
    "environment": "Environment",
    "env_type": "Env_Type",
    "hosting_zone": "hosting_zone",
    "installed_status_usu": "installed_status_usu",
    "avg_free_mem_12m": "Avg_FreeMem_12m",
    "min_free_mem_12m": "Min_FreeMem_12m",
    "current_ram_gib": "Current_RAM_GiB",
    "recommended_ram_gib": "Recommended_RAM_GiB",
    "potential_ram_reduction_gib": "Potential_RAM_Reduction_GiB",
    "ram_recommendation": "RAM_Recommendation",
    "cost_savings_eur": "Cost_Savings_EUR",
}
RS3_API_NUMERIC_SORT_FIELDS = {
    "avg_cpu_12m",
    "peak_cpu_12m",
    "current_vcpu",
    "recommended_vcpu",
    "potential_vcpu_reduction",
    "avg_free_mem_12m",
    "min_free_mem_12m",
    "current_ram_gib",
    "recommended_ram_gib",
    "potential_ram_reduction_gib",
    "cost_savings_eur",
}


def _is_rs3_recommendation_filter(filter_value):
    return str(filter_value or "").endswith("_Recommendation")


def _normalize_rs3_filter_value(workload, filter_value):
    normalized_workload = str(workload or RS3_WORKLOAD_DEFAULT).upper()
    value = str(filter_value or "").strip()
    if normalized_workload == "CPU":
        return RS3_CPU_FILTER_ALIASES.get(value, value)
    if normalized_workload == "RAM":
        return RS3_RAM_FILTER_ALIASES.get(value, value)
    return value


def _get_rs3_filter_field(filter_value):
    if str(filter_value or "").endswith("_Rightsizing"):
        return "Env_Type"
    return "Recommendation_Type" if _is_rs3_recommendation_filter(filter_value) else "Optimization_Type"


def _filter_rs3_records(records, filter_value):
    if not filter_value:
        return list(records or [])
    if str(filter_value or "").endswith("_Rightsizing"):
        env_type = "NON-PROD" if str(filter_value).startswith("NONPROD_") else "PROD"
        return [
            record
            for record in (records or [])
            if str(record.get("Env_Type") or "") == env_type
        ]
    filter_field = _get_rs3_filter_field(filter_value)
    return [
        record
        for record in (records or [])
        if str(record.get(filter_field) or "") == str(filter_value)
    ]


def _get_rs3_columns(workload, filter_value):
    normalized_workload = str(workload or RS3_WORKLOAD_DEFAULT).upper()
    if normalized_workload == "RAM":
        if str(filter_value or "").endswith("_Rightsizing"):
            return RS3_RAM_RIGHTSIZING_COLUMNS
        return (
            RS3_RAM_RECOMMENDATION_COLUMNS
            if _is_rs3_recommendation_filter(filter_value)
            else RS3_RAM_OPTIMIZATION_COLUMNS
        )
    if str(filter_value or "").endswith("_Rightsizing"):
        return RS3_CPU_RIGHTSIZING_COLUMNS
    return (
        RS3_CPU_RECOMMENDATION_COLUMNS
        if _is_rs3_recommendation_filter(filter_value)
        else RS3_CPU_OPTIMIZATION_COLUMNS
    )


def _get_rs3_filter_options(rs, workload):
    normalized_workload = str(workload or RS3_WORKLOAD_DEFAULT).upper()
    if normalized_workload == "CPU":
        return list(RS3_SCREEN_FILTER_OPTIONS["CPU"])
    if normalized_workload == "RAM":
        return list(RS3_SCREEN_FILTER_OPTIONS["RAM"])
    generic_options = (rs.get("screen_filter_options") or {}).get(normalized_workload)
    if generic_options:
        return list(generic_options)
    return list(rs.get("cpu_filter_options") or RS3_SCREEN_FILTER_OPTIONS["CPU"])


def _get_rs3_default_filter(rs, workload):
    normalized_workload = str(workload or RS3_WORKLOAD_DEFAULT).upper()
    defaults = rs.get("default_filter_by_workload") or RS3_DEFAULT_FILTER_BY_WORKLOAD
    fallback = defaults.get(normalized_workload) or RS3_DEFAULT_FILTER_BY_WORKLOAD.get(normalized_workload, "")
    options = _get_rs3_filter_options(rs, normalized_workload)
    fallback = _normalize_rs3_filter_value(normalized_workload, fallback)
    if fallback in options:
        return fallback
    return options[0] if options else fallback


def _get_rs3_summary(rs, workload, filter_value, records):
    normalized_workload = str(workload or RS3_WORKLOAD_DEFAULT).upper()
    filter_value = _normalize_rs3_filter_value(normalized_workload, filter_value)
    summaries = ((rs.get("screen_summaries") or {}).get(normalized_workload) or {})
    summary = summaries.get(filter_value)
    if summary:
        return summary

    selected = _filter_rs3_records(records, filter_value)
    reduction_key = "Potential_RAM_Reduction_GiB" if normalized_workload == "RAM" else "Potential_vCPU_Reduction"
    reduction_total = 0.0
    for record in selected:
        try:
            reduction_total += float(record.get(reduction_key) or 0)
        except (TypeError, ValueError):
            continue
    return {
        "count": len(selected),
        "prod_count": sum(str(record.get("Env_Type") or "") == "PROD" for record in selected),
        "nonprod_count": sum(str(record.get("Env_Type") or "") == "NON-PROD" for record in selected),
        "reduction_total": round(reduction_total, 1),
    }


def _get_rs3_workload_for_filter(filter_value):
    normalized = str(filter_value or "").upper()
    return "RAM" if "_RAM_" in normalized else "CPU"


def _format_rs3_sheet_label(filter_value):
    normalized_filter = RS3_CPU_FILTER_ALIASES.get(str(filter_value or ""), str(filter_value or ""))
    parts = [segment.capitalize() for segment in str(normalized_filter or "").split("_") if segment]
    return " ".join(parts)


def _build_rs3_download_sheet_options(rs):
    screen_options = (rs.get("screen_filter_options") or {}) if isinstance(rs, dict) else {}
    options = []
    for workload in ("CPU", "RAM"):
        filter_values = screen_options.get(workload) or RS3_SCREEN_FILTER_OPTIONS[workload]
        for filter_value in filter_values:
            options.append({
                "value": filter_value,
                "label": _format_rs3_sheet_label(filter_value),
                "workload": workload,
            })
    return options


def _build_rs3_download_dataframe(rightsizing, filter_value):
    workload = _get_rs3_workload_for_filter(filter_value)
    if workload == "RAM":
        source_records = (rightsizing.get("ram_optimizations") or rightsizing.get("ram_candidates") or [])
    else:
        source_records = (rightsizing.get("cpu_optimizations") or rightsizing.get("cpu_candidates") or [])

    columns = _get_rs3_columns(workload, filter_value)
    filtered_records = _filter_rs3_records(source_records, filter_value)
    rows = [
        {column: record.get(column) for column in columns}
        for record in filtered_records
    ]
    return pd.DataFrame(rows, columns=columns)


def _parse_rs3_multi_value_query_param(request, param_name):
    values = []
    for raw_value in request.GET.getlist(param_name):
        if raw_value is None:
            continue
        parts = [part.strip() for part in str(raw_value).split(",")]
        values.extend(part for part in parts if part)
    return values


def _normalize_rs3_hosting_zone_value(value):
    normalized = str(value or "").strip()
    return normalized or "none"


def _normalize_rs3_installed_status_value(value):
    return str(value or "").strip()


def _canonicalize_rs3_filter_values(values, allowed_options, normalizer):
    allowed_map = {str(option).strip().lower(): option for option in allowed_options}
    canonical = []
    invalid = []
    for value in values:
        normalized = normalizer(value)
        canonical_value = allowed_map.get(str(normalized).lower())
        if canonical_value is None:
            invalid.append(normalized)
            continue
        if canonical_value not in canonical:
            canonical.append(canonical_value)
    return canonical, invalid


def _filter_rs3_api_records(records, hosting_zones=None, installed_statuses=None):
    hosting_zone_filter = set(hosting_zones or [])
    installed_status_filter = set(installed_statuses or [])
    filtered = []
    for record in records or []:
        record_hosting_zone = _normalize_rs3_hosting_zone_value(record.get("hosting_zone"))
        record_installed_status = _normalize_rs3_installed_status_value(record.get("installed_status_usu"))
        if hosting_zone_filter and record_hosting_zone not in hosting_zone_filter:
            continue
        if installed_status_filter and record_installed_status not in installed_status_filter:
            continue
        filtered.append(record)
    return filtered


def _coerce_float(value):
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


def _get_rs3_api_sort_field(workload):
    return "Potential_RAM_Reduction_GiB" if str(workload or "").upper() == "RAM" else "Potential_vCPU_Reduction"


def _get_rs3_api_sort_field_map(workload):
    return RS3_API_RAM_SORT_FIELD_MAP if str(workload or "").upper() == "RAM" else RS3_API_CPU_SORT_FIELD_MAP


def _get_rs3_api_sort_params(request, workload):
    sort_order = str(request.GET.get("sort_order") or "desc").strip().lower()
    if sort_order not in {"asc", "desc"}:
        sort_order = "desc"
    requested_sort_field = str(request.GET.get("sort_field") or "").strip().lower()
    sort_field_map = _get_rs3_api_sort_field_map(workload)
    default_sort_field = "potential_ram_reduction_gib" if str(workload or "").upper() == "RAM" else "potential_vcpu_reduction"
    if requested_sort_field not in sort_field_map:
        requested_sort_field = default_sort_field
    return requested_sort_field, sort_order


def _get_rs3_api_sort_value(record, record_field, sort_field):
    value = record.get(record_field)
    if str(sort_field or "").strip().lower() in RS3_API_NUMERIC_SORT_FIELDS:
        numeric_value = _coerce_float(value)
        return (0, numeric_value)
    return (1, str(value or "").lower())


def _sort_rs3_api_records(records, workload, sort_field=None, sort_order="desc"):
    sort_field_map = _get_rs3_api_sort_field_map(workload)
    normalized_sort_field = str(sort_field or "").strip().lower()
    record_field = sort_field_map.get(normalized_sort_field, _get_rs3_api_sort_field(workload))
    reverse = str(sort_order or "desc").lower() == "desc"
    return sorted(
        records or [],
        key=lambda record: (_get_rs3_api_sort_value(record, record_field, normalized_sort_field), str(record.get("server_name") or "").lower()),
        reverse=reverse,
    )


def _build_rs3_api_summary(records, workload):
    reduction_key = _get_rs3_api_sort_field(workload)
    reduction_total = 0.0
    savings_total = 0.0
    for record in records or []:
        reduction_total += _coerce_float(record.get(reduction_key))
        savings_total += _coerce_float(record.get("Cost_Savings_EUR"))
    return {
        "count": len(records or []),
        "prod_count": sum(str(record.get("Env_Type") or "") == "PROD" for record in (records or [])),
        "nonprod_count": sum(str(record.get("Env_Type") or "") == "NON-PROD" for record in (records or [])),
        "reduction_total": round(reduction_total, 1),
        "savings_eur": round(savings_total, 2),
    }


def _enrich_rs3_api_records_with_cost_savings(records, workload, rightsizing_meta=None):
    from optimizer.services.db_analysis_service import _calculate_cpu_rightsizing_costs_eur

    rightsizing_meta = rightsizing_meta or {}
    normalized_workload = str(workload or "").upper()
    avg_ram_cost = _coerce_float(rightsizing_meta.get("avg_cost_per_gib_eur"))
    enriched = []
    for record in records or []:
        next_record = dict(record)
        if normalized_workload == "RAM":
            existing_savings = next_record.get("Cost_Savings_EUR")
            if existing_savings not in (None, ""):
                savings = _coerce_float(existing_savings)
            else:
                reduction = _coerce_float(next_record.get("Potential_RAM_Reduction_GiB"))
                savings = reduction * avg_ram_cost if avg_ram_cost > 0 else 0.0
            next_record["Cost_Savings_EUR"] = round(savings, 2)
        else:
            actual, recommended_cost, savings = _calculate_cpu_rightsizing_costs_eur(
                next_record.get("product_edition"),
                eff_quantity=next_record.get("eff_quantity"),
                recommended_vcpu=next_record.get("Recommended_vCPU"),
                reduction=next_record.get("Potential_vCPU_Reduction"),
            )
            next_record["Actual_Line_Cost"] = actual
            next_record["Recommended_Line_Cost"] = recommended_cost
            next_record["Cost_Savings_EUR"] = savings
        enriched.append(next_record)
    return enriched


def _get_rs3_api_page_size(request):
    try:
        page_size = int(request.GET.get("page_size", RS3_API_DEFAULT_PAGE_SIZE))
    except (TypeError, ValueError):
        page_size = RS3_API_DEFAULT_PAGE_SIZE
    return max(1, min(page_size, RS3_API_MAX_PAGE_SIZE))


def _get_rs3_api_columns(workload):
    return RS3_API_RAM_COLUMNS if str(workload or "").upper() == "RAM" else RS3_API_CPU_COLUMNS


def _serialize_rs3_api_record(record, workload):
    serialized = {
        "server_name": record.get("server_name"),
        "product_family": record.get("product_family") or "",
        "product_group": record.get("product_group") or "",
        "product_description": record.get("product_description") or "",
        "product_name": record.get("product_name") or "",
        "environment": record.get("Environment"),
        "env_type": record.get("Env_Type"),
        "hosting_zone": _normalize_rs3_hosting_zone_value(record.get("hosting_zone")),
        "installed_status_usu": _normalize_rs3_installed_status_value(record.get("installed_status_usu")),
        "is_virtual": record.get("is_virtual"),
        "optimization_type": record.get("Optimization_Type"),
        "recommendation_type": record.get("Recommendation_Type"),
    }
    if str(workload or "").upper() == "RAM":
        serialized.update({
            "avg_free_mem_12m": record.get("Avg_FreeMem_12m"),
            "min_free_mem_12m": record.get("Min_FreeMem_12m"),
            "current_ram_gib": record.get("Current_RAM_GiB"),
            "recommended_ram_gib": record.get("Recommended_RAM_GiB"),
            "potential_ram_reduction_gib": record.get("Potential_RAM_Reduction_GiB"),
            "ram_recommendation": record.get("RAM_Recommendation"),
            "cost_savings_eur": record.get("Cost_Savings_EUR"),
        })
    else:
        serialized.update({
            "avg_cpu_12m": record.get("Avg_CPU_12m"),
            "peak_cpu_12m": record.get("Peak_CPU_12m"),
            "current_vcpu": record.get("Current_vCPU"),
            "recommended_vcpu": record.get("Recommended_vCPU"),
            "potential_vcpu_reduction": record.get("Potential_vCPU_Reduction"),
            "cpu_recommendation": record.get("CPU_Recommendation"),
            "cost_savings_eur": record.get("Cost_Savings_EUR"),
        })
    return _make_json_serializable(serialized)


def _format_rs3_api_screen_label(filter_value):
    normalized = str(filter_value or "").strip().upper()
    mapping = {
        # Primary (new) filter keys
        "PROD_CPU_RIGHTSIZING": "PROD CPU Right-Sizing",
        "NONPROD_CPU_RIGHTSIZING": "Nonprod CPU Right-Sizing",
        "PROD_RAM_RIGHTSIZING": "PROD RAM Right-Sizing",
        "NONPROD_RAM_RIGHTSIZING": "Nonprod RAM Right-Sizing",
        # Legacy aliases (kept for backward compatibility)
        "PROD_CPU_OPTIMIZATION": "PROD CPU Right-Sizing",
        "NONPROD_CPU_OPTIMIZATION": "Nonprod CPU Right-Sizing",
        "PROD_RAM_OPTIMIZATION": "PROD RAM Right-Sizing",
        "NONPROD_RAM_OPTIMIZATION": "Nonprod RAM Right-Sizing",
    }
    return mapping.get(normalized, _format_rs3_sheet_label(filter_value))


def _build_table_rows(records, columns):
    """Project record dicts into table row arrays using the requested column order."""
    return [[record.get(column) for column in columns] for record in records]


def _validate_excel_upload(file_path, original_filename):
    """
    Validate saved file by magic bytes. Returns (True, None) if valid, (False, error_message) otherwise.
    - .xlsx: PK (ZIP)
    - .xls: D0 CF 11 E0 (OLE)
    """
    try:
        with open(file_path, "rb") as f:
            header = f.read(8)
        name_lower = (original_filename or "").lower()
        if name_lower.endswith(".xlsx"):
            if not header.startswith(b"PK"):
                return False, "File is not a valid .xlsx (expected ZIP format)."
        elif name_lower.endswith(".xls"):
            if not header[:8].startswith(b"\xD0\xCF\x11\xE0"):
                return False, "File is not a valid .xls (expected OLE format)."
        return True, None
    except Exception as e:
        logger.warning("Upload validation failed: %s", e)
        return False, "Could not validate file content."


def _sanitize_filename(name: str, max_len: int = 200) -> str:
    """Remove path separators and control chars; limit length for Content-Disposition."""
    if not name or not isinstance(name, str):
        return "download"
    name = os.path.basename(name)
    name = re.sub(r"[\x00-\x1f\x7f/\\]", "", name)
    return name[:max_len] if len(name) > max_len else name or "download"


def _safe_content_disposition(filename: str) -> str:
    """Build safe Content-Disposition value (ASCII filename only)."""
    safe = _sanitize_filename(filename)
    return f'attachment; filename="{safe}"'


class OptimizerLoginView(LoginView):
    """Enterprise login: unified auth page (Sign in | Create account)."""
    template_name = "optimizer/auth.html"
    redirect_authenticated_user = True

    def get_default_redirect_url(self):
        return _build_post_login_redirect_url()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["active_tab"] = self.request.GET.get("tab", "login")
        context["registered"] = self.request.GET.get("registered") == "1"
        context["signup_form"] = SignUpForm()
        return context


@require_http_methods(["GET", "POST"])
@csrf_protect
def signup_view(request):
    """
    Enterprise signup: POST creates user and redirects to login with success message.
    GET redirects to login with tab=signup so the unified auth page shows the signup form.
    """
    if request.user.is_authenticated:
        return redirect("optimizer:dashboard")
    if request.method == "GET":
        login_url = reverse("optimizer:login")
        if not login_url.startswith("/"):
            login_url = "/" + login_url
        return redirect(login_url + "?tab=signup")
    form = SignUpForm(request.POST)
    if not form.is_valid():
        return render(
            request,
            "optimizer/auth.html",
            {
                "form": AuthenticationForm(),
                "signup_form": form,
                "active_tab": "signup",
                "registered": False,
            },
        )
    user = form.save(commit=True)
    logger.info("User registered id=%s username=%s", user.pk, user.username)
    messages.success(request, "Account created. Please sign in with your credentials.")
    login_url = reverse("optimizer:login")
    if not login_url.startswith("/"):
        login_url = "/" + login_url
    return redirect(login_url + "?registered=1")


@require_http_methods(["GET", "POST"])
def logout_view(request):
    """Log out and redirect to login. Accepts GET so 'Log out' link works without a form."""
    auth_logout(request)
    return redirect("optimizer:login")


@require_GET
def health(request):
    """Liveness: returns 200 if the app is running. No auth required."""
    return HttpResponse("ok", content_type="text/plain")


@require_GET
def ready(request):
    """Readiness: checks DB connectivity. No auth required. Use for load balancer probes."""
    try:
        from django.db import connection
        connection.ensure_connection()
        return HttpResponse("ready", content_type="text/plain")
    except Exception as e:
        logger.warning("Ready check failed: %s", e)
        return HttpResponse("not ready", status=503, content_type="text/plain")


@require_GET
@login_required
def home(request):
    """Redirect directly to the dashboard."""
    return redirect("optimizer:dashboard")


@require_http_methods(["GET", "POST"])
@csrf_protect
@login_required
def upload(request):
    """Handle file upload, process Excel, store results in session, redirect to results or loading."""
    if request.method != "POST":
        return redirect("optimizer:dashboard")

    file_obj = request.FILES.get("excel_file")
    if not file_obj or not file_obj.name.lower().endswith((".xlsx", ".xls")):
        messages.error(request, "Please upload an Excel file (.xlsx or .xls).")
        return redirect("optimizer:dashboard")

    upload_dir = getattr(settings, "MEDIA_ROOT", None) or os.path.join(settings.BASE_DIR, "uploads")
    os.makedirs(upload_dir, exist_ok=True)
    safe_name = f"{uuid.uuid4().hex}_{_sanitize_filename(file_obj.name)}"
    file_path = os.path.join(upload_dir, safe_name)
    with open(file_path, "wb") as f:
        for chunk in file_obj.chunks():
            f.write(chunk)

    allowed, error_msg = _validate_excel_upload(file_path, file_obj.name)
    if not allowed:
        try:
            os.remove(file_path)
        except OSError:
            pass
        messages.error(request, error_msg or "Invalid file.")
        return render(request, "optimizer/home.html", {"error": error_msg, "title": "SQL License Optimizer"})

    if not request.session.session_key:
        request.session.save()

    analysis_record = AnalysisSession(
        user=request.user,
        file_name=file_obj.name,
        file_path=os.path.basename(file_path),
        status="processing",
        session_key=request.session.session_key or "",
    )
    analysis_record.save()
    request_id = getattr(request, "request_id", None)
    logger.info(
        "Analysis started analysis_id=%s user_id=%s username=%s file_name=%s uploaded_at=%s request_id=%s",
        analysis_record.id,
        request.user.id,
        request.user.username,
        file_obj.name,
        timezone.localtime(analysis_record.created_at).isoformat() if analysis_record.created_at else None,
        request_id,
    )

    result = run_analysis(file_path, file_obj.name)
    if not result["success"]:
        analysis_record.status = "failed"
        analysis_record.error_message = result["error"] or ""
        analysis_record.completed_at = timezone.now()
        analysis_record.summary_metrics = {}
        analysis_record.save()
        logger.warning(
            "Analysis failed analysis_id=%s user_id=%s username=%s file_name=%s completed_at=%s error=%s request_id=%s",
            analysis_record.id,
            request.user.id,
            request.user.username,
            file_obj.name,
            timezone.localtime(analysis_record.completed_at).isoformat() if analysis_record.completed_at else None,
            analysis_record.error_message,
            request_id,
        )
        try:
            os.remove(file_path)
        except OSError:
            pass
        messages.error(request, result["error"] or "Analysis failed.")
        return render(request, "optimizer/home.html", {"error": result["error"], "title": "SQL License Optimizer"})

    context = result["context"]
    summary_metrics = build_analysis_summary_metrics(context)
    analysis_record.status = "completed"
    analysis_record.result_data = _make_json_serializable(context)
    analysis_record.summary_metrics = _make_json_serializable(summary_metrics)
    analysis_record.completed_at = timezone.now()
    analysis_record.save()

    request.session["optimizer_analysis_id"] = analysis_record.id
    request.session.pop("optimizer_results", None)
    request.session.pop("optimizer_file_path", None)
    request.session.pop("optimizer_file_name", None)
    logger.info(
        "Analysis completed analysis_id=%s user_id=%s username=%s file_name=%s completed_at=%s summary_metrics=%s request_id=%s",
        analysis_record.id,
        request.user.id,
        request.user.username,
        file_obj.name,
        timezone.localtime(analysis_record.completed_at).isoformat() if analysis_record.completed_at else None,
        summary_metrics,
        request_id,
    )
    return redirect("optimizer:results")


@require_GET
@login_required
def results(request):
    """Unified results/dashboard: reads live data from Server, CPUUtilisation, USU tables."""
    from optimizer.services.db_analysis_service import compute_live_db_metrics

    context = compute_live_db_metrics()

    render_context = build_dashboard_context(context, getattr(request, "request_id", None))
    render_context["rule_results"] = context["rule_results"]
    render_context["license_metrics"] = context["license_metrics"]
    render_context["file_name"] = ""
    render_context["sheet_names_used"] = {}
    render_context["analysis_id"] = None
    render_context["analysis_status"] = "completed"
    render_context["analysis_source_file_name"] = ""
    render_context["analysis_sheet_names"] = {}
    render_context["analysis_created_at"] = context.get("data_refreshed_at")
    render_context["data_source"] = "database"
    render_context["total_license_cost_eu"] = _eu_currency(render_context.get("total_license_cost", 0))
    render_context["azure_payg_savings_eu"] = _eu_currency(render_context.get("azure_payg_savings", 0))
    render_context["azure_payg_total_cost_eu"] = _eu_currency(
        context.get("rule_results", {}).get("azure_payg_total_cost_eur", 0)
    )
    render_context["retired_devices_savings_eu"] = _eu_currency(render_context.get("retired_devices_savings", 0))
    render_context["rightsizing_cpu_savings_eu"] = _eu_currency(render_context.get("rightsizing_cpu_savings", 0))
    render_context["rightsizing_savings_eu"] = _eu_currency(render_context.get("rightsizing_savings", 0))
    render_context["potential_savings_eu"] = _eu_currency(render_context.get("potential_savings", 0))

    # Add EU-formatted cost fields to each price_distribution item
    for item in render_context.get("price_distribution", []):
        item["total_cost_eu"] = _eu_currency(item.get("total_cost", 0))
        item["avg_price_eu"] = _eu_currency(item.get("avg_price", 0))

    # Pagination for Rule 1 and Rule 2 raw data (6 rows per page, no scrolling)
    per_page = 6
    rr = context.get("rule_results") or {}
    render_context["rr"] = rr
    render_context["azure_payg_count"] = int(rr.get("azure_payg_count") or render_context.get("azure_payg_count", 0) or 0)
    render_context["retired_count"] = int(rr.get("retired_count") or render_context.get("retired_count", 0) or 0)
    azure_full = rr.get("azure_payg") or []
    retired_full = rr.get("retired_devices") or []
    requested_rule1_page = _get_page_number(request, "rule1_page")
    requested_rule2_page = _get_page_number(request, "rule2_page")
    total_rule1_pages = max(1, (len(azure_full) + per_page - 1) // per_page)
    total_rule2_pages = max(1, (len(retired_full) + per_page - 1) // per_page)
    rule1_page = min(requested_rule1_page, total_rule1_pages)
    rule2_page = min(requested_rule2_page, total_rule2_pages)
    RULE1_DISPLAY_COLS = [
        "server_name",
        "topology_type",
        "cpu_core_count",
        "cpu_socket_count",
        "manufacturer",
        "product_family",
        "product_edition",
        "product_description",
        "license_metric",
        "no_license_required",
        "install_status",
        "environment",
        "u_hosting_zone",
        "cloud_provider",
        "is_cloud_device",
        "inventory_status_standard",
        "Actual_Line_Cost",
    ]
    RULE2_DISPLAY_COLS = [
        "server_name",
        "topology_type",
        "cpu_core_count",
        "cpu_socket_count",
        "manufacturer",
        "product_family",
        "product_edition",
        "product_description",
        "license_metric",
        "no_license_required",
        "install_status",
        "environment",
        "u_hosting_zone",
        "cloud_provider",
        "is_cloud_device",
        "inventory_status_standard",
        "Actual_Line_Cost",
    ]
    _RULE_HIDDEN_COLS = {"product_group"}
    all_rule1_keys = list(azure_full[0].keys()) if azure_full else []
    all_rule2_keys = list(retired_full[0].keys()) if retired_full else []
    rule1_keys = [c for c in RULE1_DISPLAY_COLS if c in all_rule1_keys] + \
                 [c for c in all_rule1_keys if c not in RULE1_DISPLAY_COLS and c not in _RULE_HIDDEN_COLS]
    rule2_keys = [c for c in RULE2_DISPLAY_COLS if c in all_rule2_keys] + \
                 [c for c in all_rule2_keys if c not in RULE2_DISPLAY_COLS and c not in _RULE_HIDDEN_COLS]
    azure_slice = azure_full[(rule1_page - 1) * per_page : rule1_page * per_page]
    retired_slice = retired_full[(rule2_page - 1) * per_page : rule2_page * per_page]

    render_context["azure_payg_prod_candidates_count"] = int(rr.get("azure_payg_prod_candidates_count") or 0)
    render_context["azure_payg_nonprod_candidates_count"] = int(rr.get("azure_payg_nonprod_candidates_count") or 0)

    render_context["azure_payg_page"] = [[r.get(k) for k in rule1_keys] for r in azure_slice]
    render_context["retired_devices_page"] = [[r.get(k) for k in rule2_keys] for r in retired_slice]
    render_context["rule1_page"] = rule1_page
    render_context["rule2_page"] = rule2_page
    render_context["total_rule1_pages"] = total_rule1_pages
    render_context["total_rule2_pages"] = total_rule2_pages
    render_context["rule1_keys"] = rule1_keys
    render_context["rule2_keys"] = rule2_keys
    render_context["rule1_prev"] = max(1, rule1_page - 1)
    render_context["rule1_next"] = min(total_rule1_pages, rule1_page + 1)
    render_context["rule2_prev"] = max(1, rule2_page - 1)
    render_context["rule2_next"] = min(total_rule2_pages, rule2_page + 1)
    # Full data as JSON for client-side pagination (no full page reload)
    render_context["rule1_data_json"] = [
        [r.get(k) for k in rule1_keys] for r in azure_full
    ]
    render_context["rule2_data_json"] = [
        [r.get(k) for k in rule2_keys] for r in retired_full
    ]

    # ── Strategy 3: CPU & RAM Right-Sizing pagination ─────────────────────────
    rs = context.get("rightsizing") or {}
    cpu_full = rs.get("cpu_optimizations") or rs.get("cpu_candidates") or []
    ram_full = rs.get("ram_optimizations") or rs.get("ram_candidates") or []

    rs3_workload_options = ["ALL"] + [str(option).upper() for option in (rs.get("workload_options") or ["CPU", "RAM"])]
    rs3_workload = str(
        request.GET.get("rs3_workload")
        or RS3_WORKLOAD_DEFAULT
    ).upper()
    if rs3_workload not in rs3_workload_options:
        rs3_workload = rs3_workload_options[0] if rs3_workload_options else RS3_WORKLOAD_DEFAULT

    requested_rs3_filter = request.GET.get("rs3_filter") or ""
    rs3_filter_workload = rs3_workload if rs3_workload in ("CPU", "RAM") else _get_rs3_workload_for_filter(requested_rs3_filter or _get_rs3_default_filter(rs, "CPU"))
    rs3_filter_options = _get_rs3_filter_options(rs, rs3_filter_workload)
    rs3_default_filter = _get_rs3_default_filter(rs, rs3_filter_workload)
    rs3_filter = _normalize_rs3_filter_value(rs3_filter_workload, requested_rs3_filter or rs3_default_filter)
    if rs3_filter not in rs3_filter_options:
        rs3_filter = rs3_default_filter

    rs3_source_records = ram_full if rs3_workload == "RAM" else cpu_full
    rs3_filtered_records = _filter_rs3_records(rs3_source_records, rs3_filter)
    rs3_page = _get_page_number(request, "rs3_page")
    total_rs3_pages = max(1, (len(rs3_filtered_records) + per_page - 1) // per_page)
    rs3_page = min(rs3_page, total_rs3_pages)

    rs3_columns = _get_rs3_columns(rs3_workload, rs3_filter)
    rs3_keys = [key for key in rs3_columns if any(key in row for row in rs3_filtered_records)]
    rs3_slice = rs3_filtered_records[(rs3_page - 1) * per_page : rs3_page * per_page]
    rs3_summary = _get_rs3_summary(rs, rs3_workload, rs3_filter, rs3_source_records)

    rs_cpu_page = _get_page_number(request, "rs3_cpu_page")
    rs_ram_page = _get_page_number(request, "rs3_ram_page")
    total_rs_cpu_pages = max(1, (len(cpu_full) + per_page - 1) // per_page)
    total_rs_ram_pages = max(1, (len(ram_full) + per_page - 1) // per_page)
    rs_cpu_page = min(rs_cpu_page, total_rs_cpu_pages)
    rs_ram_page = min(rs_ram_page, total_rs_ram_pages)

    cpu_initial_filter = _normalize_rs3_filter_value("CPU", rs3_filter if rs3_workload in ("ALL", "CPU") else _get_rs3_default_filter(rs, "CPU"))
    cpu_initial_records = _filter_rs3_records(cpu_full, cpu_initial_filter)
    cpu_initial_columns = _get_rs3_columns("CPU", cpu_initial_filter)
    ram_initial_filter = _normalize_rs3_filter_value("RAM", rs3_filter if rs3_workload in ("ALL", "RAM") else _get_rs3_default_filter(rs, "RAM"))
    cpu_keys = [key for key in cpu_initial_columns if any(key in row for row in cpu_initial_records)]
    ram_initial_records = _filter_rs3_records(ram_full, ram_initial_filter)
    ram_initial_columns = _get_rs3_columns("RAM", ram_initial_filter)
    ram_keys = [key for key in ram_initial_columns if any(key in row for row in ram_initial_records)]
    total_rs_cpu_pages = max(1, (len(cpu_initial_records) + per_page - 1) // per_page)
    total_rs_ram_pages = max(1, (len(ram_initial_records) + per_page - 1) // per_page)
    rs_cpu_page = min(rs_cpu_page, total_rs_cpu_pages)
    rs_ram_page = min(rs_ram_page, total_rs_ram_pages)
    cpu_slice = cpu_initial_records[(rs_cpu_page - 1) * per_page : rs_cpu_page * per_page]
    ram_slice = ram_initial_records[(rs_ram_page - 1) * per_page : rs_ram_page * per_page]
    _cpu_hz = str(request.GET.get("rs3_cpu_hosting_zone") or "").strip()
    if _cpu_hz:
        _cpu_hz = _normalize_rs3_hosting_zone_value(_cpu_hz)
    requested_rs3_cpu_hosting_zone = _cpu_hz if _cpu_hz in RS3_API_HOSTING_ZONE_OPTIONS else ""
    _cpu_st = _normalize_rs3_installed_status_value(request.GET.get("rs3_cpu_status"))
    requested_rs3_cpu_status = _cpu_st if _cpu_st in RS3_API_INSTALLED_STATUS_USU_OPTIONS else ""
    _ram_hz = str(request.GET.get("rs3_ram_hosting_zone") or "").strip()
    if _ram_hz:
        _ram_hz = _normalize_rs3_hosting_zone_value(_ram_hz)
    requested_rs3_ram_hosting_zone = _ram_hz if _ram_hz in RS3_API_HOSTING_ZONE_OPTIONS else ""
    _ram_st = _normalize_rs3_installed_status_value(request.GET.get("rs3_ram_status"))
    requested_rs3_ram_status = _ram_st if _ram_st in RS3_API_INSTALLED_STATUS_USU_OPTIONS else ""

    render_context.update({
        "rightsizing": rs,
        "rightsizing_cpu_page": _build_table_rows(cpu_slice, cpu_keys),
        "rightsizing_ram_page": _build_table_rows(ram_slice, ram_keys),
        "cpu_keys": cpu_keys,
        "ram_keys": ram_keys,
        "rs3_cpu_page": rs_cpu_page,
        "rs3_ram_page": rs_ram_page,
        "total_rs3_cpu_pages": total_rs_cpu_pages,
        "total_rs3_ram_pages": total_rs_ram_pages,
        "rs3_cpu_prev": max(1, rs_cpu_page - 1),
        "rs3_cpu_next": min(total_rs_cpu_pages, rs_cpu_page + 1),
        "rs3_ram_prev": max(1, rs_ram_page - 1),
        "rs3_ram_next": min(total_rs_ram_pages, rs_ram_page + 1),
        "rightsizing_selected_page": _build_table_rows(rs3_slice, rs3_keys),
        "rs3_keys": rs3_keys,
        "rs3_page": rs3_page,
        "total_rs3_pages": total_rs3_pages,
        "rs3_prev": max(1, rs3_page - 1),
        "rs3_next": min(total_rs3_pages, rs3_page + 1),
        "rs3_workload_options": rs3_workload_options,
        "rs3_selected_workload": rs3_workload,
        "rs3_filter_options": rs3_filter_options,
        "rs3_selected_filter": rs3_filter,
        "rs3_cpu_filter_options": _get_rs3_filter_options(rs, "CPU"),
        "rs3_cpu_selected_filter": cpu_initial_filter,
        "rs3_ram_filter_options": _get_rs3_filter_options(rs, "RAM"),
        "rs3_ram_selected_filter": ram_initial_filter,
        "rs3_api_hosting_options": RS3_API_HOSTING_ZONE_OPTIONS,
        "rs3_api_status_options": RS3_API_INSTALLED_STATUS_USU_OPTIONS,
        "rs3_cpu_selected_hosting_zone": requested_rs3_cpu_hosting_zone,
        "rs3_cpu_selected_status": requested_rs3_cpu_status,
        "rs3_ram_selected_hosting_zone": requested_rs3_ram_hosting_zone,
        "rs3_ram_selected_status": requested_rs3_ram_status,
        "rs3_selected_summary": rs3_summary,
        "rightsizing_cpu_data_json": cpu_full,
        "rightsizing_ram_data_json": ram_full,
        "crit_cpu_data_json": rs.get("crit_cpu_optimizations") or [],
        "crit_ram_data_json": rs.get("crit_ram_optimizations") or [],
        "lifecycle_data_json": rs.get("lifecycle_risk_flags") or [],
        "download_sheet_options": _build_rs3_download_sheet_options(rs),
    })
    # ── Data Quality: USU, Grafana, Flat Files ────────────────────────────────
    try:
        from optimizer.models import USUInstallation, GrafanaMetricSnapshot, GrafanaMetricMonthlyRollup, CPUUtilisation
        usu_install_total = USUInstallation.objects.filter(product_family="SQL Server").count()
        usu_qs = (
            USUInstallation.objects
            .select_related("server")
            .order_by("-fetched_at")[:300]
        )
        dq_usu_rows = [
            {
                "Server": i.server.server_name if i.server else "",
                "Product": i.product_description or "",
                "Edition": i.product_edition or "",
                "Device Status": i.device_status or "",
                "Inv. Status": i.inv_status_std_name or "",
                "CPU Cores": str(i.cpu_core_count) if i.cpu_core_count is not None else "",
                "License Metric": i.license_metric or "",
                "Fetched At": i.fetched_at.strftime("%Y-%m-%d") if i.fetched_at else "",
            }
            for i in usu_qs
        ]

        # Build server → product lookup for Grafana and Flat Files enrichment
        _srv_product = {}
        for _inst in USUInstallation.objects.select_related("server").filter(server__isnull=False).only(
            "server__server_name", "product_family", "product_description"
        ):
            _sn = _inst.server.server_name if _inst.server else None
            if _sn and _sn not in _srv_product:
                _srv_product[_sn] = {
                    "Product Family": _inst.product_family or "",
                    "Product Name": _inst.product_description or "",
                    "Product Description": _inst.product_description or "",
                }

        from django.db.models import Avg, Max, Min, Count

        dq_grafana_connections_by_server = []

        grafana_qs = (
            GrafanaMetricSnapshot.objects
            .select_related("server")
            .order_by("-metric_ts")[:500]
        )
        grafana_snap_list = list(grafana_qs)
        dq_grafana_is_snapshot = bool(grafana_snap_list)

        if dq_grafana_is_snapshot:
            dq_grafana_rows = [
                {
                    "Server": g.server.server_name if g.server else "",
                    "Product Family": _srv_product.get(g.server.server_name if g.server else "", {}).get("Product Family", ""),
                    "Product Name": _srv_product.get(g.server.server_name if g.server else "", {}).get("Product Name", ""),
                    "Product Description": _srv_product.get(g.server.server_name if g.server else "", {}).get("Product Description", ""),
                    "Dashboard": g.dashboard or "",
                    "Metric": g.metric_name or "",
                    "Value": str(round(float(g.metric_value), 4)) if g.metric_value is not None else "",
                    "Unit": g.metric_unit or "",
                    "Timestamp": g.metric_ts.strftime("%Y-%m-%d %H:%M") if g.metric_ts else "",
                    "Fetched At": g.fetched_at.strftime("%Y-%m-%d") if g.fetched_at else "",
                }
                for g in grafana_snap_list
            ]

            # ── Chart 1: avg value per metric type ────────────────────────────
            metric_avg_map = {}
            metric_unit_map = {}
            for g in grafana_snap_list:
                if g.metric_value is None:
                    continue
                mn = g.metric_name or "unknown"
                metric_avg_map.setdefault(mn, []).append(float(g.metric_value))
                if g.metric_unit:
                    metric_unit_map[mn] = g.metric_unit
            dq_grafana_metric_avg = [
                {"metric": mn, "avg": round(sum(vals) / len(vals), 4), "unit": metric_unit_map.get(mn, "")}
                for mn, vals in sorted(metric_avg_map.items())
            ]

            # ── Chart 2: timeline (avg per metric per timestamp bucket) ────────
            from collections import defaultdict
            timeline_map = defaultdict(lambda: defaultdict(list))
            for g in grafana_snap_list:
                if g.metric_value is None or not g.metric_ts:
                    continue
                ts_label = g.metric_ts.strftime("%Y-%m-%d %H:%M")
                timeline_map[g.metric_name or "unknown"][ts_label].append(float(g.metric_value))
            dq_grafana_timeline = {}
            for mn, ts_dict in timeline_map.items():
                pts = sorted(ts_dict.items())
                dq_grafana_timeline[mn] = {
                    "x": [p[0] for p in pts],
                    "y": [round(sum(v) / len(v), 4) for _, v in pts],
                }

            # ── Chart 3: servers by avg value (per metric) ────────────────────
            server_metric_map = defaultdict(lambda: defaultdict(list))
            for g in grafana_snap_list:
                if g.metric_value is None:
                    continue
                sn = g.server.server_name if g.server else "Unknown"
                server_metric_map[g.metric_name or "unknown"][sn].append(float(g.metric_value))
            dq_grafana_top_servers = {}
            for mn, srv_dict in server_metric_map.items():
                ranked = sorted(
                    [{"server": sn, "avg": round(sum(v) / len(v), 4)} for sn, v in srv_dict.items()],
                    key=lambda r: r["avg"], reverse=True
                )
                dq_grafana_top_servers[mn] = ranked

            connection_snapshots = (
                GrafanaMetricSnapshot.objects
                .select_related("server")
                .filter(metric_name="connections")
                .order_by("-metric_ts")
            )
            connection_server_map = defaultdict(list)
            for row in connection_snapshots:
                if row.metric_value is None:
                    continue
                server_name = row.server.server_name if row.server else "Unknown"
                connection_server_map[server_name].append(float(row.metric_value))
            dq_grafana_connections_by_server = sorted(
                [
                    {"server": server_name, "avg": round(sum(values) / len(values), 4)}
                    for server_name, values in connection_server_map.items()
                ],
                key=lambda r: r["avg"],
                reverse=True,
            )

            # ── Chart 4: KPI summary ───────────────────────────────────────────
            dq_grafana_kpis = {
                "total_records": len(grafana_snap_list),
                "unique_metrics": len(metric_avg_map),
                "unique_servers": len({(g.server.server_name if g.server else "") for g in grafana_snap_list}),
                "unique_dashboards": len({g.dashboard for g in grafana_snap_list if g.dashboard}),
            }

            dq_grafana_rollup_summary = []

        else:
            rollup_qs = (
                GrafanaMetricMonthlyRollup.objects
                .select_related("server")
                .order_by("-period_month")[:500]
            )
            rollup_list = list(rollup_qs)
            dq_grafana_rows = [
                {
                    "Server": r.server.server_name if r.server else "",
                    "Metric": r.metric_name or "",
                    "Unit": r.metric_unit or "",
                    "Period": str(r.period_month) if r.period_month else "",
                    "Avg": str(round(float(r.avg_value), 4)) if r.avg_value is not None else "",
                    "Max": str(round(float(r.max_value), 4)) if r.max_value is not None else "",
                    "Min": str(round(float(r.min_value), 4)) if r.min_value is not None else "",
                    "Samples": str(r.sample_count) if r.sample_count is not None else "",
                }
                for r in rollup_list
            ]

            # ── Chart 1: avg/max/min per metric (rollup) ──────────────────────
            rollup_metric_map = {}
            for r in rollup_list:
                mn = r.metric_name or "unknown"
                rollup_metric_map.setdefault(mn, {"avg": [], "max": [], "min": [], "unit": r.metric_unit or ""})
                if r.avg_value is not None:
                    rollup_metric_map[mn]["avg"].append(float(r.avg_value))
                if r.max_value is not None:
                    rollup_metric_map[mn]["max"].append(float(r.max_value))
                if r.min_value is not None:
                    rollup_metric_map[mn]["min"].append(float(r.min_value))
            dq_grafana_rollup_summary = [
                {
                    "metric": mn,
                    "avg": round(sum(d["avg"]) / len(d["avg"]), 4) if d["avg"] else 0,
                    "max": round(max(d["max"]), 4) if d["max"] else 0,
                    "min": round(min(d["min"]), 4) if d["min"] else 0,
                    "unit": d["unit"],
                }
                for mn, d in sorted(rollup_metric_map.items())
            ]

            # ── Chart 2: timeline (avg per metric per period) ─────────────────
            from collections import defaultdict
            period_map = defaultdict(lambda: defaultdict(list))
            for r in rollup_list:
                if r.avg_value is None or not r.period_month:
                    continue
                period_map[r.metric_name or "unknown"][str(r.period_month)].append(float(r.avg_value))
            dq_grafana_timeline = {}
            for mn, pd_dict in period_map.items():
                pts = sorted(pd_dict.items())
                dq_grafana_timeline[mn] = {
                    "x": [p[0] for p in pts],
                    "y": [round(sum(v) / len(v), 4) for _, v in pts],
                }

            # ── Chart 3: servers per metric ───────────────────────────────────
            srv_map = defaultdict(lambda: defaultdict(list))
            for r in rollup_list:
                if r.avg_value is None:
                    continue
                sn = r.server.server_name if r.server else "Unknown"
                srv_map[r.metric_name or "unknown"][sn].append(float(r.avg_value))
            dq_grafana_top_servers = {}
            for mn, sd in srv_map.items():
                ranked = sorted(
                    [{"server": sn, "avg": round(sum(v) / len(v), 4)} for sn, v in sd.items()],
                    key=lambda r: r["avg"], reverse=True
                )
                dq_grafana_top_servers[mn] = ranked

            connection_rollups = (
                GrafanaMetricMonthlyRollup.objects
                .select_related("server")
                .filter(metric_name="connections")
                .order_by("-period_month")
            )
            connection_server_map = defaultdict(list)
            for row in connection_rollups:
                if row.avg_value is None:
                    continue
                server_name = row.server.server_name if row.server else "Unknown"
                connection_server_map[server_name].append(float(row.avg_value))
            dq_grafana_connections_by_server = sorted(
                [
                    {"server": server_name, "avg": round(sum(values) / len(values), 4)}
                    for server_name, values in connection_server_map.items()
                ],
                key=lambda r: r["avg"],
                reverse=True,
            )

            dq_grafana_metric_avg = dq_grafana_rollup_summary

            dq_grafana_kpis = {
                "total_records": len(rollup_list),
                "unique_metrics": len(rollup_metric_map),
                "unique_servers": len({(r.server.server_name if r.server else "") for r in rollup_list}),
                "unique_dashboards": 0,
            }

        flatfile_qs = (
            CPUUtilisation.objects
            .select_related("server")
            .filter(source__in=["boones_public", "boones_private"])
            .order_by("-period_month")[:300]
        )
        dq_flatfile_rows = [
            {
                "Server": f.server.server_name if f.server else "",
                "Source": f.get_source_display(),
                "Period": str(f.period_month) if f.period_month else "",
                "Avg CPU%": str(round(float(f.avg_cpu_pct), 2)) if f.avg_cpu_pct is not None else "",
                "Max CPU%": str(round(float(f.max_cpu_pct), 2)) if f.max_cpu_pct is not None else "",
                "Logical CPUs": str(f.logical_cpu_count) if f.logical_cpu_count is not None else "",
                "RAM (GiB)": str(round(float(f.physical_ram_gib), 2)) if f.physical_ram_gib is not None else "",
                "Avg Free Mem%": str(round(float(f.avg_free_memory_pct), 2)) if f.avg_free_memory_pct is not None else "",
                "Min_FreeMem_12m": str(round(float(f.min_free_memory_pct), 2)) if f.min_free_memory_pct is not None else "",
            }
            for f in flatfile_qs
        ]
    except Exception as _dq_exc:
        logger.warning("Data Quality fetch failed: %s", _dq_exc)
        dq_usu_rows, dq_grafana_rows, dq_flatfile_rows = [], [], []
        dq_grafana_metric_avg, dq_grafana_timeline, dq_grafana_top_servers = [], {}, {}
        dq_grafana_connections_by_server = []
        dq_grafana_rollup_summary, dq_grafana_kpis = [], {"total_records": 0, "unique_metrics": 0, "unique_servers": 0, "unique_dashboards": 0}
        dq_grafana_is_snapshot = False
        usu_install_total = 0

    render_context["dq_usu_rows"] = dq_usu_rows
    render_context["dq_usu_keys"] = list(dq_usu_rows[0].keys()) if dq_usu_rows else []
    if usu_install_total:
        render_context["total_devices_analyzed"] = usu_install_total
    render_context["dq_grafana_rows"] = dq_grafana_rows
    render_context["dq_grafana_keys"] = list(dq_grafana_rows[0].keys()) if dq_grafana_rows else []
    render_context["dq_flatfile_rows"] = dq_flatfile_rows
    render_context["dq_flatfile_keys"] = list(dq_flatfile_rows[0].keys()) if dq_flatfile_rows else []
    render_context["dq_grafana_metric_avg"] = dq_grafana_metric_avg
    render_context["dq_grafana_timeline"] = dq_grafana_timeline
    render_context["dq_grafana_top_servers"] = dq_grafana_top_servers
    render_context["dq_grafana_connections_by_server"] = dq_grafana_connections_by_server
    render_context["dq_grafana_rollup_summary"] = dq_grafana_rollup_summary
    render_context["dq_grafana_kpis"] = dq_grafana_kpis
    render_context["dq_grafana_is_snapshot"] = dq_grafana_is_snapshot
    render_context["dq_grafana_metric_names"] = list(dq_grafana_top_servers.keys())
    render_context["dq_grafana_metric_options"] = [
        {"value": metric_name, "label": _format_metric_label(metric_name)}
        for metric_name in render_context["dq_grafana_metric_names"]
    ]

    # Interactive Plotly charts (hover tooltips, responsive)
    if get_all_plotly_specs is not None:
        try:
            render_context["chart_specs"] = get_all_plotly_specs(
                context.get("rule_results", {}),
                context.get("license_metrics", {}),
            )
        except Exception as e:
            logger.exception("Plotly chart specs failed: %s", e)
            render_context["chart_specs"] = {}
    else:
        render_context["chart_specs"] = {}

    # Fallback: static matplotlib images when Plotly not used
    if getattr(settings, "OPTIMIZER_CHARTS_ENABLED", True) and generate_all_charts is not None and not render_context.get("chart_specs"):
        try:
            charts, chart_formats = generate_all_charts(
                context.get("rule_results", {}),
                context.get("license_metrics", {}),
            )
            render_context["chart_images"] = charts
            render_context["chart_formats"] = chart_formats
        except Exception as e:
            logger.exception("Chart generation failed: %s", e)
            render_context["chart_images"] = {}
            render_context["chart_formats"] = {}
    else:
        render_context.setdefault("chart_images", {})
        render_context.setdefault("chart_formats", {})
    return render(request, "optimizer/dashboard.html", render_context)


@require_GET
@login_required
def dashboard(request):
    """Same as results: unified tabbed view."""
    return results(request)


@require_http_methods(["GET", "POST"])
@csrf_protect
@login_required
def profile_page(request):
    """Display and update the logged-in user's profile details."""
    profile = _get_or_create_user_profile(request.user)
    if request.method == "POST":
        form = UserProfileForm(request.POST, instance=profile, user=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, "Profile updated successfully.")
            return redirect("optimizer:profile")
    else:
        form = UserProfileForm(instance=profile, user=request.user)
    context = _build_profile_context(request.user, profile, form=form)
    return render(request, "optimizer/profile.html", context)


@require_GET
@login_required
def alerts(request):
    from optimizer.services.alerts import build_alert_page_context
    context = build_alert_page_context(request.GET)
    return render(request, "optimizer/alerts.html", context)


def _build_legacy_report_markdown(context):
    report_text = context.get("report_text")
    if not report_text:
        return ""
    return build_report_markdown(
        report_text,
        report_context=_build_report_render_context(context),
    )


def _resolve_report_markdown(context, agentic=None):
    """
    Resolve the best available report markdown in priority order:
      1. Live agent preview from build_live_agent_report_preview()
         — internally tries agent/liscence-optimizer/src/tools/report_generator.py first,
           then falls back to Django-native renderer
      2. Legacy fallback (Azure OpenAI text or static template) — only if step 1 completely fails

    Always uses live DB data so the Executive Summary matches the dashboard values.
    Stored AgentRun.report_markdown (from prior agent runs) is intentionally skipped here
    because it may contain stale counts and savings that diverge from the current data.
    """
    agentic = agentic or {}

    # Priority 1: live agent preview (uses agent/liscence-optimizer tool or Django fallback)
    preview_markdown = ""
    try:
        from optimizer.services.ai_report_generator import build_live_agent_report_preview

        preview = build_live_agent_report_preview(usecase_id="uc_1_2_3")
        preview_markdown = preview.get("report_markdown") or ""
        preview_summary = preview.get("summary_context") or {}

        # Accept the preview if the agent produced ANY text, even without live data —
        # the agent renderer always generates structured markdown regardless of data size.
        if preview_markdown:
            logger.info(
                "Using agent live preview report for /report/ (agent_payg=%s retired=%s cpu=%s ram=%s).",
                preview_summary.get("azure_payg_count"),
                preview_summary.get("retired_count"),
                preview_summary.get("cpu_count"),
                preview_summary.get("ram_count"),
            )
            return normalize_report_content_text(preview_markdown)
        else:
            logger.warning("build_live_agent_report_preview() returned empty markdown.")
    except Exception as exc:
        logger.exception(
            "build_live_agent_report_preview() raised an exception for /report/ — "
            "falling back to legacy report. Error: %s",
            exc,
        )

    # Priority 2: build agent-format report directly from the already-computed context
    # (the context was built by _get_db_context_for_report which calls compute_live_db_metrics,
    #  so the rule_results and rightsizing data are already available without extra DB queries)
    logger.warning(
        "build_live_agent_report_preview() failed — building agent-format report "
        "directly from pre-computed context data (all 3 strategies)."
    )
    try:
        from optimizer.services.ai_report_generator import (
            build_agent_strategy_results_payload,
            _build_local_rules_evaluation,
            _build_agent_report_summary_context,
            _render_local_agent_report_markdown,
        )
        strategy_results = build_agent_strategy_results_payload(context)
        rules_evaluation = _build_local_rules_evaluation(
            rule_results=context.get("rule_results") or {},
            rightsizing=context.get("rightsizing") or {},
        )
        summary_context = _build_agent_report_summary_context(context, strategy_results)
        md = _render_local_agent_report_markdown(
            usecase_id="uc_1_2_3",
            strategy_results=strategy_results,
            rules_evaluation=rules_evaluation,
            summary_context=summary_context,
        )
        if md:
            logger.info("Agent-format report built directly from context (%d chars).", len(md))
            return md
    except Exception as exc2:
        logger.exception("Direct agent-format report build from context also failed: %s", exc2)

    # Absolute last resort: empty string so the template shows "No report generated"
    logger.error("All report generation paths failed — returning empty report.")
    return ""


@require_GET
@login_required
def report_page(request):
    """Report page: live data from DB, rendered as structured markdown."""
    context = _get_db_context_for_report()

    # Merge in the latest agentic run data (agent report + candidates)
    from optimizer.services.db_analysis_service import get_latest_agentic_context
    agentic = get_latest_agentic_context()
    context["report_text"] = _resolve_report_markdown(context, agentic=agentic)
    context["agentic"] = agentic
    context["has_agentic_data"] = agentic.get("has_agentic_data", False)

    return render(request, "optimizer/report.html", context)


@require_GET
@login_required
def report_download(request, format_type):
    """Download report as PDF, Word, or Excel — uses agent AI report markdown."""
    normalized_format = REPORT_FORMAT_ALIASES.get(format_type, format_type)
    if normalized_format not in ALLOWED_REPORT_FORMATS:
        return HttpResponse("Invalid format.", status=400)
    context = _get_db_context_for_report()
    from optimizer.services.db_analysis_service import get_latest_agentic_context

    report_text = _resolve_report_markdown(context, agentic=get_latest_agentic_context())
    report_text = normalize_report_content_text(report_text or "")
    generated_at = timezone.localtime()
    base_name = "sql_license_optimization_report_db"
    # Pass report_context=None so export functions use _parse_report_blocks(report_text)
    # instead of the old _build_template_blocks — this preserves the agent AI report structure.
    if normalized_format == "pdf":
        content = export_pdf(report_text, generated_at=generated_at, report_context=None)
        if content is None:
            return HttpResponse("PDF export not available (install reportlab).", status=501)
        response = HttpResponse(content, content_type="application/pdf")
        response["Content-Disposition"] = _safe_content_disposition(f"{base_name}.pdf")
        return response
    if normalized_format == "docx":
        content = export_docx(report_text, generated_at=generated_at, report_context=None)
        if content is None:
            return HttpResponse("Word export not available (install python-docx).", status=501)
        response = HttpResponse(content, content_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document")
        response["Content-Disposition"] = _safe_content_disposition(f"{base_name}.docx")
        return response
    if normalized_format == "xlsx":
        content = export_xlsx(report_text, generated_at=generated_at, report_context=None)
        if content is None:
            return HttpResponse("Excel export not available (install openpyxl).", status=501)
        response = HttpResponse(
            content,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = _safe_content_disposition(f"{base_name}.xlsx")
        return response
    return HttpResponse("Invalid format.", status=400)


@require_GET
@login_required
def analysis_logs(request):
    """Return all persisted analysis logs for the authenticated user."""
    return JsonResponse({"logs": get_user_analysis_logs(request.user)})


@require_GET
@login_required
def api_strategy3_rightsizing(request):
    """
    API: Filtered Strategy 3 table data for CPU or RAM right-sizing.

    GET /api/strategy3-rightsizing/

    Query params:
    - workload=CPU|RAM (optional when screen_filter is provided)
    - screen_filter=PROD_CPU_Optimization|NONPROD_CPU_Optimization|PROD_CPU_Recommendation|...
    - hosting_zone=Public Cloud (repeat or comma-separate for multiple values)
    - installed_status_usu=Installed (repeat or comma-separate for multiple values)
    - sort_field=server_name|avg_cpu_12m|current_ram_gib|...
    - sort_order=asc|desc
    - page=1
    - page_size=25
    """
    from optimizer.services.db_analysis_service import compute_live_db_metrics

    requested_screen_filter = request.GET.get("screen_filter") or request.GET.get("filter") or ""
    requested_workload = str(request.GET.get("workload") or "").strip().upper()

    if requested_workload and requested_workload not in {"CPU", "RAM"}:
        return JsonResponse({
            "status": "failed",
            "error": "Invalid workload. Use CPU or RAM.",
        }, status=400)

    if requested_screen_filter:
        workload = _get_rs3_workload_for_filter(requested_screen_filter)
    else:
        workload = requested_workload or "CPU"
    requested_screen_filter = _normalize_rs3_filter_value(workload, requested_screen_filter)

    live_context = compute_live_db_metrics()
    rightsizing = live_context.get("rightsizing") or {}
    rightsizing_meta = live_context.get("rightsizing_meta") or {}
    filter_options = _get_rs3_filter_options(rightsizing, workload)
    screen_filter = requested_screen_filter or _get_rs3_default_filter(rightsizing, workload)
    if screen_filter not in filter_options:
        return JsonResponse({
            "status": "failed",
            "error": f"Invalid screen_filter. Allowed values: {filter_options}",
        }, status=400)

    requested_hosting_zones = _parse_rs3_multi_value_query_param(request, "hosting_zone")
    requested_installed_statuses = _parse_rs3_multi_value_query_param(request, "installed_status_usu")

    hosting_zones, invalid_hosting_zones = _canonicalize_rs3_filter_values(
        requested_hosting_zones,
        RS3_API_HOSTING_ZONE_OPTIONS,
        _normalize_rs3_hosting_zone_value,
    )
    installed_statuses, invalid_installed_statuses = _canonicalize_rs3_filter_values(
        requested_installed_statuses,
        RS3_API_INSTALLED_STATUS_USU_OPTIONS,
        _normalize_rs3_installed_status_value,
    )
    if invalid_hosting_zones or invalid_installed_statuses:
        errors = {}
        if invalid_hosting_zones:
            errors["hosting_zone"] = {
                "invalid": invalid_hosting_zones,
                "allowed": RS3_API_HOSTING_ZONE_OPTIONS,
            }
        if invalid_installed_statuses:
            errors["installed_status_usu"] = {
                "invalid": invalid_installed_statuses,
                "allowed": RS3_API_INSTALLED_STATUS_USU_OPTIONS,
            }
        return JsonResponse({
            "status": "failed",
            "error": "Invalid filter values supplied.",
            "details": errors,
        }, status=400)

    source_records = (
        rightsizing.get("ram_optimizations") or rightsizing.get("ram_candidates") or []
        if workload == "RAM"
        else rightsizing.get("cpu_optimizations") or rightsizing.get("cpu_candidates") or []
    )
    screen_records = _filter_rs3_records(source_records, screen_filter)
    filtered_records = _filter_rs3_api_records(
        screen_records,
        hosting_zones=hosting_zones,
        installed_statuses=installed_statuses,
    )
    enriched_records = _enrich_rs3_api_records_with_cost_savings(filtered_records, workload, rightsizing_meta=rightsizing_meta)
    sort_field, sort_order = _get_rs3_api_sort_params(request, workload)
    sorted_records = _sort_rs3_api_records(enriched_records, workload, sort_field=sort_field, sort_order=sort_order)

    page = _get_page_number(request, "page")
    page_size = _get_rs3_api_page_size(request)
    total_records = len(sorted_records)
    total_pages = max(1, (total_records + page_size - 1) // page_size)
    page = min(page, total_pages)
    start = (page - 1) * page_size
    end = start + page_size
    paged_records = sorted_records[start:end]

    return JsonResponse({
        "request_id": str(uuid.uuid4()),
        "status": "completed",
        "result": {
            "workload": workload,
            "screen_filter": screen_filter,
            "screen_label": _format_rs3_api_screen_label(screen_filter),
            "columns": _get_rs3_api_columns(workload),
            "filters": {
                "hosting_zone": hosting_zones,
                "installed_status_usu": installed_statuses,
            },
            "available_filters": {
                "screen_filter": filter_options,
                "hosting_zone": RS3_API_HOSTING_ZONE_OPTIONS,
                "installed_status_usu": RS3_API_INSTALLED_STATUS_USU_OPTIONS,
            },
            "summary": _build_rs3_api_summary(sorted_records, workload),
            "total": total_records,
            "sort": {
                "field": sort_field,
                "order": sort_order,
            },
            "pagination": {
                "page": page,
                "page_size": page_size,
                "total_records": total_records,
                "total_pages": total_pages,
            },
            "items": [
                _serialize_rs3_api_record(record, workload)
                for record in paged_records
            ],
            "error": rightsizing.get("error"),
        },
    })


@require_GET
@login_required
def api_savings_summary(request):
    """
    API: Potential savings summary for all three strategies.

    GET /api/savings-summary/

    Response shape:
      {
        "request_id": str,
        "status": "completed",
        "result": {
          "strategies": [
            {
              "id": "byol_to_payg",
              "name": "BYOL to PAYG",
              "candidates": int,          <- azure_payg_count
              "label": "candidates",
              "savings_eur": float
            },
            {
              "id": "retired_but_reporting",
              "name": "Retired But Reporting",
              "candidates": int,          <- retired_count
              "label": "devices",
              "savings_eur": float
            },
            {
              "id": "rightsizing",
              "name": "CPU Right-Sizing",
              "candidates": int,          <- cpu_count (servers eligible)
              "label": "servers",
              "savings_eur": float,
              "vcpu_reduction": int,      <- total vCPUs that can be removed
              "ram_reduction_gib": float, <- total GiB RAM that can be freed
              "avg_cost_per_core_pair_eur": float,
              "avg_cost_per_gib_eur": float
            }
          ],
          "total_savings_eur": float,
          "data_refreshed_at": str (ISO-8601)
        }
      }

    Intended for PowerBI and other BI consumers. Authentication required.
    """
    from optimizer.services.db_analysis_service import compute_live_db_metrics

    ctx = compute_live_db_metrics()

    rr = ctx.get("rule_results") or {}
    rws = ctx.get("rule_wise_savings") or {}
    rs_meta = ctx.get("rightsizing_meta") or {}
    rs = ctx.get("rightsizing") or {}

    refreshed_at = ctx.get("data_refreshed_at")
    refreshed_at_str = refreshed_at.isoformat() if refreshed_at else None

    strategies = [
        {
            "id": "byol_to_payg",
            "name": "BYOL to PAYG",
            "candidates": int(rr.get("azure_payg_count") or 0),
            "label": "candidates",
            "savings_eur": float(rws.get("azure_payg") or 0),
        },
        {
            "id": "retired_but_reporting",
            "name": "Retired But Reporting",
            "candidates": int(rr.get("retired_count") or 0),
            "label": "devices",
            "savings_eur": float(rws.get("retired_devices") or 0),
        },
        {
            "id": "rightsizing",
            "name": "CPU Right-Sizing",
            "candidates": int(rs_meta.get("cpu_count") or rs.get("cpu_count") or 0),
            "label": "servers",
            "savings_eur": float(rws.get("rightsizing") or 0),
            "vcpu_reduction": int(rs_meta.get("total_vcpu_reduction") or rs.get("total_vcpu_reduction") or 0),
            "ram_reduction_gib": float(rs_meta.get("total_ram_reduction_gib") or rs.get("total_ram_reduction_gib") or 0),
            "avg_cost_per_core_pair_eur": float(rs_meta.get("avg_cost_per_core_pair_eur") or rs.get("avg_cost_per_core_pair_eur") or 0),
            "avg_cost_per_gib_eur": float(rs_meta.get("avg_cost_per_gib_eur") or rs.get("avg_cost_per_gib_eur") or 0),
        },
    ]

    total_savings_eur = round(sum(s["savings_eur"] for s in strategies), 2)

    return JsonResponse({
        "request_id": str(uuid.uuid4()),
        "status": "completed",
        "result": {
            "strategies": strategies,
            "total_savings_eur": total_savings_eur,
            "data_refreshed_at": refreshed_at_str,
        },
    })


@require_GET
@login_required
def api_oracle_data(request):
    """
    API: USU Installations and Demand Details — MSSQL and/or Oracle (Java) product families.

    GET /api/usu-data/

    Query params (all optional):
      family    — "mssql" | "oracle" | "all"  (default: "all")
      type      — "installations" | "demand" | "all"  (default: "all")
      page      — 1-based page number  (default: 1)
      page_size — rows per page, max 500  (default: 100)
      hosting    — filter by NormalizedHostingZone (e.g. "Public Cloud")
      status     — filter by install_status, installations only (e.g. "Installed", "Retired")
      sort_field — field to sort by (server_name, product_description, product_edition,
                   product_family, install_status, inv_status_std_name, cpu_core_count,
                   hosting_zone, environment, manufacturer, inventory_date)
      sort_order — "asc" | "desc"  (default: "asc")

    Response shape (family=all):
      {
        "request_id": str,
        "status": "completed",
        "family": "all",
        "result": { "installations": {...}, "demand_details": {...} }
      }

    Response shape (family=mssql or family=oracle):
      {
        "request_id": str,
        "status": "completed",
        "family": "mssql",               # or "oracle"
        "product_family": "SQL Server",  # or "Java"
        "label": "MSSQL Server Data",
        "result": { "installations": {...}, "demand_details": {...} }
      }
    """
    from math import ceil
    from optimizer.models import USUInstallation, USUDemandDetail

    # family key → (DB product_family value, human label)
    FAMILY_MAP = {
        "mssql":  ("SQL Server", "MSSQL Server Data"),
        "oracle": ("Java",       "Oracle Server Data"),
    }

    # ── Parse query params ────────────────────────────────────────────────────
    family_param = request.GET.get("family", "all").lower().strip()
    if family_param not in ("mssql", "oracle", "all"):
        return JsonResponse(
            {"error": "Invalid family. Use 'mssql', 'oracle', or 'all'."},
            status=400,
        )

    try:
        page = max(1, int(request.GET.get("page", 1)))
    except (ValueError, TypeError):
        page = 1

    try:
        page_size = min(max(1, int(request.GET.get("page_size", 100))), 500)
    except (ValueError, TypeError):
        page_size = 100

    data_type      = request.GET.get("type", "all").lower()
    hosting_filter = request.GET.get("hosting", "").strip()
    status_filter  = request.GET.get("status", "").strip()
    sort_field     = request.GET.get("sort_field", "").strip()
    sort_order     = request.GET.get("sort_order", "asc").lower().strip()
    if sort_order not in ("asc", "desc"):
        sort_order = "asc"
    skip = (page - 1) * page_size

    # Allowed sort fields for installations (maps API key → item dict key)
    INST_SORTABLE_FIELDS = {
        "server_name", "product_description", "product_edition", "product_family",
        "install_status", "inv_status_std_name", "cpu_core_count",
        "hosting_zone", "environment", "manufacturer", "inventory_date",
    }

    def _normalize_hosting(zone):
        z = str(zone or "").strip().lower()
        if not z:
            return ""
        if "public" in z:
            return "Public Cloud"
        if "avs" in z or ("private" in z and "cloud" in z):
            return "Private Cloud AVS"
        return str(zone or "").strip()

    def _fetch_installations(db_pf):
        if data_type not in ("all", "installations"):
            return {"total": 0, "page": page, "page_size": page_size, "total_pages": 0, "items": []}

        qs = USUInstallation.objects.select_related("server").order_by("server__server_name")
        if db_pf is not None:
            qs = qs.filter(product_family=db_pf)
        if status_filter:
            qs = qs.filter(device_status__iexact=status_filter)

        rows = list(qs.values(
            "server__server_name", "server__hosting_zone", "server__environment",
            "server__is_cloud_device", "server__cloud_provider",
            "device_status", "no_license_required", "product_description",
            "product_edition", "product_family", "product_group", "manufacturer",
            "inv_status_std_name", "cpu_core_count", "cpu_socket_count",
            "topology_type", "inventory_date",
        ))

        if hosting_filter:
            rows = [
                r for r in rows
                if _normalize_hosting(r["server__hosting_zone"]).lower() == hosting_filter.lower()
            ]

        # Build mapped items first so we can sort by the public field names
        all_items = [
            {
                "server_name":         r["server__server_name"],
                "hosting_zone":        _normalize_hosting(r["server__hosting_zone"]),
                "environment":         r["server__environment"],
                "install_status":      r["device_status"],
                "no_license_required": r["no_license_required"],
                "product_description": r["product_description"],
                "product_edition":     r["product_edition"],
                "product_family":      r["product_family"],
                "product_group":       r["product_group"],
                "manufacturer":        r["manufacturer"],
                "inv_status_std_name": r["inv_status_std_name"],
                "cpu_core_count":      float(r["cpu_core_count"]) if r["cpu_core_count"] is not None else None,
                "cpu_socket_count":    r["cpu_socket_count"],
                "topology_type":       r["topology_type"],
                "inventory_date":      r["inventory_date"].isoformat() if r["inventory_date"] else None,
                "is_cloud_device":     r["server__is_cloud_device"],
                "cloud_provider":      r["server__cloud_provider"],
            }
            for r in rows
        ]

        if sort_field and sort_field in INST_SORTABLE_FIELDS:
            reverse = sort_order == "desc"
            all_items.sort(
                key=lambda item: (
                    item.get(sort_field) is None or item.get(sort_field) == "",
                    str(item.get(sort_field) or "").lower(),
                ),
                reverse=reverse,
            )

        total = len(all_items)
        return {
            "total": total,
            "page": page,
            "page_size": page_size,
            "total_pages": ceil(total / page_size) if total else 0,
            "sort_field": sort_field,
            "sort_order": sort_order,
            "items": all_items[skip: skip + page_size],
        }

    def _fetch_demand(db_pf):
        if data_type not in ("all", "demand"):
            return {"total": 0, "page": page, "page_size": page_size, "total_pages": 0, "items": []}

        qs = USUDemandDetail.objects.select_related("server").order_by("server__server_name")
        if db_pf is not None:
            qs = qs.filter(product_family=db_pf)

        rows = list(qs.values(
            "server__server_name", "server__hosting_zone", "server__environment",
            "server__is_cloud_device", "server__cloud_provider",
            "product_description", "product_edition", "product_family", "product_group",
            "manufacturer", "eff_quantity", "no_license_required",
            "device_purpose", "cpu_core_count", "topology_type", "virt_type",
        ))

        if hosting_filter:
            rows = [
                r for r in rows
                if _normalize_hosting(r["server__hosting_zone"]).lower() == hosting_filter.lower()
            ]

        total = len(rows)
        return {
            "total": total,
            "page": page,
            "page_size": page_size,
            "total_pages": ceil(total / page_size) if total else 0,
            "items": [
                {
                    "server_name":         r["server__server_name"],
                    "hosting_zone":        _normalize_hosting(r["server__hosting_zone"]),
                    "environment":         r["server__environment"],
                    "product_description": r["product_description"],
                    "product_edition":     r["product_edition"],
                    "product_family":      r["product_family"],
                    "product_group":       r["product_group"],
                    "manufacturer":        r["manufacturer"],
                    "eff_quantity":        float(r["eff_quantity"]) if r["eff_quantity"] is not None else None,
                    "no_license_required": r["no_license_required"],
                    "device_purpose":      r["device_purpose"],
                    "cpu_core_count":      float(r["cpu_core_count"]) if r["cpu_core_count"] is not None else None,
                    "topology_type":       r["topology_type"],
                    "virt_type":           r["virt_type"],
                    "is_cloud_device":     r["server__is_cloud_device"],
                    "cloud_provider":      r["server__cloud_provider"],
                }
                for r in rows[skip: skip + page_size]
            ],
        }

    # ── Build response ────────────────────────────────────────────────────────
    if family_param == "all":
        return JsonResponse({
            "request_id": str(uuid.uuid4()),
            "status":     "completed",
            "family":     "all",
            "result": {
                "installations":  _fetch_installations(None),
                "demand_details": _fetch_demand(None),
            },
        })

    # Single family
    db_pf, label = FAMILY_MAP[family_param]
    return JsonResponse({
        "request_id":     str(uuid.uuid4()),
        "status":         "completed",
        "family":         family_param,
        "product_family": db_pf,
        "label":          label,
        "result": {
            "installations":  _fetch_installations(db_pf),
            "demand_details": _fetch_demand(db_pf),
        },
    })


@require_GET
@login_required
def api_agent_runs(request):
    """
    API: List recent agent runs.

    GET /api/agent-runs/
    Response shape (A2A-style):
      {
        "request_id": str,
        "status": "completed",
        "result": {
          "runs": [...],
          "total": int
        }
      }
    """
    from optimizer.services.db_analysis_service import get_agent_run_list
    limit = min(int(request.GET.get("limit", 20)), 100)
    runs = get_agent_run_list(limit=limit)
    return JsonResponse({
        "request_id": str(uuid.uuid4()),
        "status": "completed",
        "result": {
            "runs": runs,
            "total": len(runs),
        },
    })


@require_GET
@login_required
def api_agent_run_detail(request, run_id):
    """
    API: Get a single agent run with its candidates.

    GET /api/agent-runs/<run_id>/
    Response shape:
      {
        "request_id": str,
        "status": "completed",
        "result": { agent_run, candidates, ... }
      }
    """
    from optimizer.models import AgentRun, OptimizationCandidate, OptimizationDecision
    try:
        run = AgentRun.objects.get(pk=run_id)
    except (AgentRun.DoesNotExist, Exception):
        return JsonResponse({"status": "failed", "error": "Agent run not found."}, status=404)

    candidates_qs = (
        OptimizationCandidate.objects.filter(agent_run=run)
        .select_related("server", "rule")
        .order_by("-estimated_saving_eur")
    )
    candidates = []
    for c in candidates_qs:
        decision = None
        try:
            decision = {
                "decision": c.decision.decision,
                "decided_by_email": c.decision.decided_by_email or "",
                "decided_at": c.decision.decided_at.isoformat() if c.decision.decided_at else None,
                "decision_notes": c.decision.decision_notes or "",
            }
        except OptimizationDecision.DoesNotExist:
            pass
        candidates.append({
            "id": str(c.id),
            "use_case": c.use_case,
            "server_name": c.server.server_name if c.server else "",
            "rule_name": c.rule.rule_name if c.rule else "",
            "rule_code": c.rule.rule_code if c.rule else "",
            "recommendation": c.recommendation,
            "rationale": c.rationale,
            "estimated_saving_eur": float(c.estimated_saving_eur) if c.estimated_saving_eur is not None else None,
            "status": c.status,
            "detected_on": c.detected_on.isoformat() if c.detected_on else None,
            "decision": decision,
        })

    return JsonResponse({
        "request_id": str(uuid.uuid4()),
        "status": "completed",
        "result": {
            "agent_run": {
                "id": str(run.id),
                "run_label": run.run_label,
                "status": run.status,
                "triggered_by": run.triggered_by,
                "servers_evaluated": run.servers_evaluated,
                "candidates_found": run.candidates_found,
                "llm_model": run.llm_model,
                "llm_tokens_used": run.llm_tokens_used,
                "llm_used": run.llm_used,
                "run_duration_sec": float(run.run_duration_sec) if run.run_duration_sec else None,
                "agent_endpoint": run.agent_endpoint,
                "has_report": bool(run.report_markdown),
                "report_markdown": run.report_markdown,
                "started_at": run.started_at.isoformat() if run.started_at else None,
                "finished_at": run.finished_at.isoformat() if run.finished_at else None,
                "error_detail": run.error_detail,
            },
            "candidates": candidates,
        },
    })


@require_http_methods(["POST"])
@csrf_protect
@login_required
def api_trigger_agent_run(request):
    """
    API: Trigger a new agent run using live DB data.

    POST /api/agent-runs/trigger/
    Body (optional JSON):
      { "usecase_id": str, "notes": str }

    Response shape (A2A-style):
      {
        "request_id": str,
        "workflow_id": str,   <- agent_run UUID
        "status": "completed" | "failed",
        "result": { "agent_run_id", "report_markdown", "candidates_created", ... },
        "execution_time_ms": int
      }
    """
    import json as _json
    import time

    from optimizer.services.db_analysis_service import _build_installations_df, compute_db_metrics
    from optimizer.services.ai_report_generator import (
        build_agent_strategy_results_payload,
        generate_and_store_agentic_report,
    )

    # Parse optional body
    body = {}
    if request.content_type and "application/json" in request.content_type:
        try:
            body = _json.loads(request.body.decode("utf-8"))
        except Exception:
            pass

    usecase_id = body.get("usecase_id") or "uc_1_2_3"
    notes = body.get("notes") or ""

    started = time.time()

    # Build normalized records from live DB (same pipeline as db_analysis_service)
    try:
        df = _build_installations_df()
        records = df.to_dict("records") if not df.empty else []
    except Exception as exc:
        logger.exception("Failed to build installation records for agent run: %s", exc)
        records = []

    # Also add the strategy outputs used by the agent report.
    try:
        strategy_results = build_agent_strategy_results_payload(compute_db_metrics())
    except Exception:
        strategy_results = {}

    result = generate_and_store_agentic_report(
        records=records,
        usecase_id=usecase_id,
        strategy_results=strategy_results,
        triggered_by=request.user.email or request.user.username,
    )

    elapsed_ms = int((time.time() - started) * 1000)
    http_status = 200 if result.get("success") else 500

    return JsonResponse(
        {
            "request_id": str(uuid.uuid4()),
            "workflow_id": result.get("agent_run_id", ""),
            "status": "completed" if result.get("success") else "failed",
            "result": result,
            "execution_time_ms": elapsed_ms,
            "metadata": {
                "triggered_by": request.user.email or request.user.username,
                "usecase_id": usecase_id,
            },
        },
        status=http_status,
    )


@require_http_methods(["POST"])
@csrf_protect
@login_required
def api_candidate_decision(request, candidate_id):
    """
    API: Accept or reject an OptimizationCandidate.

    POST /api/candidates/<candidate_id>/decision/
    Body (JSON):
      {
        "decision": "accepted" | "rejected",
        "decision_notes": str   (optional)
      }

    Response:
      { "success": bool, "candidate_id": str, "decision": str }
    """
    import json as _json
    from optimizer.models import OptimizationCandidate, OptimizationDecision

    try:
        candidate = OptimizationCandidate.objects.get(pk=candidate_id)
    except (OptimizationCandidate.DoesNotExist, Exception):
        return JsonResponse({"success": False, "error": "Candidate not found."}, status=404)

    try:
        body = _json.loads(request.body.decode("utf-8"))
    except Exception:
        return JsonResponse({"success": False, "error": "Invalid JSON body."}, status=400)

    decision_value = body.get("decision", "").strip().lower()
    if decision_value not in (OptimizationDecision.DECISION_ACCEPTED, OptimizationDecision.DECISION_REJECTED):
        return JsonResponse(
            {"success": False, "error": "decision must be 'accepted' or 'rejected'."}, status=400
        )

    decision_notes = str(body.get("decision_notes") or "").strip()[:500]

    # Upsert decision — overwrite an existing decision if present
    OptimizationDecision.objects.update_or_create(
        candidate=candidate,
        defaults={
            "tenant": candidate.tenant,
            "decision": decision_value,
            "decided_by": request.user.get_full_name() or request.user.username,
            "decided_by_email": request.user.email or None,
            "decision_notes": decision_notes,
        },
    )

    return JsonResponse({
        "success": True,
        "candidate_id": str(candidate.id),
        "decision": decision_value,
        "server_name": candidate.server.server_name if candidate.server else "",
    })


@require_GET
@login_required
def api_rule1_data(request):
    """API: Rule 1 raw candidate data with optional sorting and pagination.

    GET /api/rule1-data/
    Query params:
      sort_field  – column key to sort by (must match a key in the data)
      sort_order  – asc | desc  (default: asc)
      page        – 1-based page number (default: 1)
      page_size   – rows per page, max 200 (default: 25)
    """
    from optimizer.services.db_analysis_service import compute_live_db_metrics
    context = compute_live_db_metrics()
    rr = context.get("rule_results", {})
    data = list(rr.get("azure_payg") or [])

    keys = list(data[0].keys()) if data else []

    sort_field = str(request.GET.get("sort_field", "")).strip()
    sort_order = str(request.GET.get("sort_order", "asc")).strip().lower()
    if sort_order not in ("asc", "desc"):
        sort_order = "asc"

    if sort_field and sort_field in keys:
        reverse_sort = sort_order == "desc"
        data = sorted(
            data,
            key=lambda r: (r.get(sort_field) is None or r.get(sort_field) == "", str(r.get(sort_field) or "").lower()),
            reverse=reverse_sort,
        )

    try:
        page = max(1, int(request.GET.get("page", 1)))
    except (TypeError, ValueError):
        page = 1
    try:
        page_size = min(200, max(1, int(request.GET.get("page_size", 25))))
    except (TypeError, ValueError):
        page_size = 25

    total = len(data)
    total_pages = max(1, (total + page_size - 1) // page_size)
    page = min(page, total_pages)
    start = (page - 1) * page_size
    page_data = data[start: start + page_size]

    return JsonResponse({
        "keys": keys,
        "rows": [[r.get(k) for k in keys] for r in page_data],
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": total_pages,
        "sort_field": sort_field,
        "sort_order": sort_order,
    })


@require_GET
@login_required
def api_rule2_data(request):
    """API: Rule 2 retired-devices data with optional sorting and pagination.

    GET /api/rule2-data/
    Query params:
      sort_field  – column key to sort by (must match a key in the data)
      sort_order  – asc | desc  (default: asc)
      page        – 1-based page number (default: 1)
      page_size   – rows per page, max 200 (default: 25)
    """
    from optimizer.services.db_analysis_service import compute_live_db_metrics
    context = compute_live_db_metrics()
    rr = context.get("rule_results", {})
    data = list(rr.get("retired_devices") or [])

    keys = list(data[0].keys()) if data else []

    sort_field = str(request.GET.get("sort_field", "")).strip()
    sort_order = str(request.GET.get("sort_order", "asc")).strip().lower()
    if sort_order not in ("asc", "desc"):
        sort_order = "asc"

    if sort_field and sort_field in keys:
        reverse_sort = sort_order == "desc"
        data = sorted(
            data,
            key=lambda r: (r.get(sort_field) is None or r.get(sort_field) == "", str(r.get(sort_field) or "").lower()),
            reverse=reverse_sort,
        )

    try:
        page = max(1, int(request.GET.get("page", 1)))
    except (TypeError, ValueError):
        page = 1
    try:
        page_size = min(200, max(1, int(request.GET.get("page_size", 25))))
    except (TypeError, ValueError):
        page_size = 25

    total = len(data)
    total_pages = max(1, (total + page_size - 1) // page_size)
    page = min(page, total_pages)
    start = (page - 1) * page_size
    page_data = data[start: start + page_size]

    return JsonResponse({
        "keys": keys,
        "rows": [[r.get(k) for k in keys] for r in page_data],
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": total_pages,
        "sort_field": sort_field,
        "sort_order": sort_order,
    })


@require_GET
@login_required
def download_demand_data(request):
    """Export all Total Demand License rows with Actual_Line_Cost as Excel."""
    from optimizer.models import USUDemandDetail
    from optimizer.services.db_analysis_service import (
        _get_rightsizing_cpu_license_cost_eur,
        SHOWCASE_ONLY_PRODUCT_FAMILIES,
        _normalize_hosting_zone,
    )
    from io import BytesIO

    rows = USUDemandDetail.objects.exclude(
        product_family__in=SHOWCASE_ONLY_PRODUCT_FAMILIES
    ).select_related("server").values(
        "server__server_name",
        "server__hosting_zone",
        "server__environment",
        "product_description",
        "product_edition",
        "product_family",
        "eff_quantity",
        "cpu_core_count",
        "no_license_required",
    ).order_by("server__server_name", "product_description")

    records = []
    for r in rows:
        eff_qty = float(r["eff_quantity"] or 0)
        price = _get_rightsizing_cpu_license_cost_eur(str(r["product_edition"] or ""))
        actual_line_cost = round((price * eff_qty) / 2, 2)
        records.append({
            "Server Name":              r["server__server_name"] or "",
            "Hosting Zone":             _normalize_hosting_zone(r["server__hosting_zone"] or ""),
            "Environment":              r["server__environment"] or "",
            "Product Description":      r["product_description"] or "",
            "Product Edition":          r["product_edition"] or "",
            "Product Family":           r["product_family"] or "",
            "Eff. Quantity":            eff_qty,
            "CPU Core Count":           float(r["cpu_core_count"] or 0),
            "No License Required":      int(r["no_license_required"] or 0),
            "Price EUR (per core pair)": price,
            "Actual Line Cost (EUR)":   actual_line_cost,
        })

    df = pd.DataFrame(records)
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Total Demand Licenses")
    buf.seek(0)
    response = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="total_demand_licenses.xlsx"'
    return response


@require_GET
@login_required
def download_rule_data(request, rule_id):
    """Download Rule 1 or Rule 2 data as Excel. rule_id whitelisted to rule1, rule2."""
    if rule_id not in ALLOWED_RULE_IDS:
        return HttpResponse("Invalid rule.", status=400)
    from optimizer.services.db_analysis_service import (
        compute_live_db_metrics,
        _build_raw_installations_df,
        _build_raw_rule1_df,
    )
    context = compute_live_db_metrics()
    rr = context.get("rule_results", {})
    if rule_id == "rule1":
        data = rr.get("azure_payg", [])
        filename = "azure_payg_candidates.xlsx"
    else:
        data = rr.get("retired_devices", [])
        filename = "retired_devices_with_installations.xlsx"
    if not data:
        return HttpResponse("No data to download.", status=404)
    from io import BytesIO
    buf = BytesIO()
    if rule_id == "rule2":
        raw_df = _build_raw_installations_df()
        results_df = pd.DataFrame(data)
        data_sources_df = pd.DataFrame([
            {
                "Source Table": "usu_installation",
                "Django Model": "USUInstallation",
                "Key Columns Used": "device_status, no_license_required, manufacturer, product_family, product_group, product_description, product_edition, license_metric, cpu_core_count, cpu_socket_count, topology_type, inv_status_std_name",
                "Filter Applied (Rule 2)": "device_status = 'retired' AND no_license_required = 0",
            },
            {
                "Source Table": "server",
                "Django Model": "Server",
                "Key Columns Used": "server_name, hosting_zone, environment, cloud_provider, is_cloud_device, installed_status_usu, installed_status_boones, is_active",
                "Filter Applied (Rule 2)": "is_active = True (joined via FK to USUInstallation)",
            },
        ])
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            results_df.to_excel(writer, index=False, sheet_name="Rule2 Retired Devices")
            if not raw_df.empty:
                raw_df.to_excel(writer, index=False, sheet_name="Raw Input Data")
            data_sources_df.to_excel(writer, index=False, sheet_name="Data Sources")
    else:  # rule1
        raw_df = _build_raw_rule1_df()
        results_df = pd.DataFrame(data)
        data_sources_df = pd.DataFrame([
            {
                "Source Table": "usu_installation",
                "Django Model": "USUInstallation",
                "Key Columns Used": "inv_status_std_name, no_license_required, product_family, manufacturer, product_description, device_status",
                "Filter Applied (Rule 1)": "hosting_zone normalized to 'Public Cloud'/'Private Cloud AVS' AND inv_status_std_name != 'License Included' AND no_license_required = 0",
            },
            {
                "Source Table": "server",
                "Django Model": "Server",
                "Key Columns Used": "server_name, hosting_zone, environment, cloud_provider, is_cloud_device, installed_status_usu, installed_status_boones, is_active",
                "Filter Applied (Rule 1)": "hosting_zone in Public Cloud / Private Cloud AVS (after normalization); is_active = True",
            },
        ])
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            results_df.to_excel(writer, index=False, sheet_name="Rule1 PAYG Candidates")
            if not raw_df.empty:
                raw_df.to_excel(writer, index=False, sheet_name="Raw Input Data")
            data_sources_df.to_excel(writer, index=False, sheet_name="Data Sources")
    buf.seek(0)
    response = HttpResponse(buf.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = _safe_content_disposition(filename)
    return response


@require_GET
@login_required
def download_rightsizing_sheet(request, sheet_key):
    """Download a specific Strategy 3 sheet as a single-sheet Excel workbook."""
    normalized_sheet_key = str(sheet_key or "").strip()
    if normalized_sheet_key not in RS3_DOWNLOAD_SHEET_KEYS:
        return HttpResponse("Invalid sheet.", status=400)

    from io import BytesIO
    from optimizer.services.db_analysis_service import build_rightsizing_sheet_export

    df = build_rightsizing_sheet_export(normalized_sheet_key)
    workbook_name = f"{normalized_sheet_key.lower()}.xlsx"
    sheet_name = normalized_sheet_key[:31] or "Sheet1"

    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = _safe_content_disposition(workbook_name)
    return response


@require_GET
@login_required
def download_uc1_input_data(request):
    """
    Export the full UC1 input dataset (all rows, all columns) plus a filter-funnel
    summary sheet matching the UC1 analysis breakdown.

    Sheet 1 – "UC1 Input Data": every row fed into the UC1 (Azure PAYG) rule,
              before any filter is applied (active servers, Java excluded).
    Sheet 2 – "UC1 Summary": row-count funnel for each UC1 filter step and a
              hosting_zone breakdown of the full input.
    """
    from io import BytesIO
    from optimizer.services.db_analysis_service import _build_installations_df

    installations_df = _build_installations_df()
    if installations_df.empty:
        return HttpResponse("No UC1 input data found in the database.", status=404)

    # ── Compute filter-funnel counts ──────────────────────────────────────────
    total_input = len(installations_df)

    # Filter 1: hosting_zone in Public Cloud / Private Cloud AVS
    mask_f1 = installations_df["u_hosting_zone"].astype(str).str.strip().isin(
        ["Public Cloud", "Private Cloud AVS"]
    )
    count_f1 = int(mask_f1.sum())

    # Filter 2: inv_status_std_name != "License Included" (case-insensitive), on F1 survivors
    inv_norm = installations_df.loc[mask_f1, "inventory_status_standard"].astype(str).str.strip().str.lower()
    mask_f2_local = inv_norm != "license included"
    count_f2 = int(mask_f2_local.sum())

    # Filter 3: no_license_required == 0, on F1+F2 survivors
    f2_idx = installations_df.loc[mask_f1].index[mask_f2_local]
    from optimizer.rules.column_utils import find_no_license_required_column, no_license_required_is_zero
    no_lic_col = find_no_license_required_column(installations_df)
    if no_lic_col:
        mask_f3_local = no_license_required_is_zero(installations_df.loc[f2_idx, no_lic_col])
        count_f3 = int(mask_f3_local.sum())
    else:
        count_f3 = 0

    # ── Hosting zone breakdown (all input rows) ───────────────────────────────
    zone_counts = (
        installations_df["u_hosting_zone"]
        .astype(str)
        .value_counts()
        .reset_index()
    )
    zone_counts.columns = ["hosting_zone", "count"]

    # ── Build Summary DataFrame ───────────────────────────────────────────────
    summary_rows = [
        {"Description": "Total rows in DB (UC1 input)", "Count": total_input, "Note": "Active servers, Java excluded"},
        {"Description": "", "Count": None, "Note": ""},
        {"Description": "UC1 Filter 1: hosting_zone = Public Cloud / AVS", "Count": count_f1, "Note": ""},
        {"Description": "UC1 Filter 2: inv_status_std_name != Lic. Included", "Count": count_f2, "Note": ""},
        {"Description": "UC1 Filter 3: no_license_required = False", "Count": count_f3, "Note": ""},
        {"Description": "", "Count": None, "Note": ""},
        {"Description": "UC1 FINAL CANDIDATES", "Count": count_f3, "Note": "Rows the rule will action"},
        {"Description": "", "Count": None, "Note": ""},
        {"Description": "--- hosting_zone breakdown (all input rows) ---", "Count": None, "Note": ""},
    ]
    for _, zone_row in zone_counts.iterrows():
        summary_rows.append({
            "Description": f" {zone_row['hosting_zone']}",
            "Count": int(zone_row["count"]),
            "Note": "",
        })
    summary_df = pd.DataFrame(summary_rows, columns=["Description", "Count", "Note"])

    # ── Write workbook ────────────────────────────────────────────────────────
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        installations_df.to_excel(writer, index=False, sheet_name="UC1 Input Data")
        summary_df.to_excel(writer, index=False, sheet_name="UC1 Summary")

        # Style the Summary sheet to match the screenshot
        try:
            from openpyxl.styles import Font, PatternFill, Alignment
            ws = writer.sheets["UC1 Summary"]

            # Header row
            header_fill = PatternFill("solid", fgColor="1F3864")
            header_font = Font(bold=True, color="FFFFFF")
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            # Highlight "UC1 FINAL CANDIDATES" row
            for row in ws.iter_rows(min_row=2):
                desc = str(row[0].value or "")
                if "FINAL CANDIDATES" in desc:
                    cand_fill = PatternFill("solid", fgColor="BDD7EE")
                    cand_font = Font(bold=True)
                    for cell in row:
                        cell.fill = cand_fill
                        cell.font = cand_font

            # Auto-width columns
            for col in ws.columns:
                max_len = max((len(str(c.value or "")) for c in col), default=10)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)
        except Exception:
            pass  # styling is best-effort

    buf.seek(0)
    response = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = _safe_content_disposition("uc1_input_data.xlsx")
    return response


@require_GET
@login_required
def download_uc3_ram_input_data(request):
    """
    Export the full UC3.2 RAM Rightsizing input dataset (all rows, all columns) plus
    a filter-funnel summary sheet matching the UC3.2 analysis breakdown.

    Sheet 1 – "UC3 RAM Input Data": every server row fed into the RAM rightsizing rule
              (active servers with utilisation data, before any filter).
    Sheet 2 – "UC3 RAM Summary": PROD and NON-PROD filter funnels, final candidate
              counts, and an Environment breakdown of all input rows.
    """
    from io import BytesIO
    from optimizer.services.db_analysis_service import _build_rightsizing_df
    from optimizer.rules.rightsizing import NON_PROD_ENVS, find_ram_rightsizing_optimizations

    df = _build_rightsizing_df()
    if df.empty:
        return HttpResponse("No UC3 RAM rightsizing input data found in the database.", status=404)

    total_input = len(df)

    # ── PROD filter funnel (cumulative intermediate counts) ───────────────────
    prod_df = df[~df["Environment"].isin(NON_PROD_ENVS)].copy()
    prod_total = len(prod_df)

    avg_free = pd.to_numeric(prod_df.get("Avg_FreeMem_12m", pd.Series(dtype=float)), errors="coerce")
    min_free = pd.to_numeric(prod_df.get("Min_FreeMem_12m", pd.Series(dtype=float)), errors="coerce")
    ram_col  = pd.to_numeric(prod_df.get("Current_RAM_GiB",  pd.Series(dtype=float)), errors="coerce")

    mask_prod_f1 = avg_free >= 35
    prod_after_f1 = int(mask_prod_f1.sum())

    mask_prod_f2 = mask_prod_f1 & (min_free >= 20)
    prod_after_f2 = int(mask_prod_f2.sum())

    mask_prod_f3 = mask_prod_f2 & (ram_col > 8)
    prod_eligible = int(mask_prod_f3.sum())

    # ── NON-PROD filter funnel (cumulative intermediate counts) ───────────────
    nonprod_df = df[df["Environment"].isin(NON_PROD_ENVS)].copy()
    nonprod_total = len(nonprod_df)

    avg_free_np = pd.to_numeric(nonprod_df.get("Avg_FreeMem_12m", pd.Series(dtype=float)), errors="coerce")
    min_free_np = pd.to_numeric(nonprod_df.get("Min_FreeMem_12m", pd.Series(dtype=float)), errors="coerce")
    ram_col_np  = pd.to_numeric(nonprod_df.get("Current_RAM_GiB",  pd.Series(dtype=float)), errors="coerce")

    mask_np_f1 = avg_free_np >= 30
    nonprod_after_f1 = int(mask_np_f1.sum())

    mask_np_f2 = mask_np_f1 & (min_free_np >= 15)
    nonprod_after_f2 = int(mask_np_f2.sum())

    mask_np_f3 = mask_np_f2 & (ram_col_np > 4)
    nonprod_eligible = int(mask_np_f3.sum())

    # ── Actual final candidates: run the full rule (includes recommendation-band filter)
    final_df = find_ram_rightsizing_optimizations(df)
    prod_candidates   = int((final_df["Env_Type"] == "PROD").sum())
    nonprod_candidates = int((final_df["Env_Type"] == "NON-PROD").sum())
    total_candidates  = prod_candidates + nonprod_candidates

    # ── Environment breakdown (all input rows) ────────────────────────────────
    env_counts = (
        df["Environment"]
        .astype(str)
        .value_counts()
        .reset_index()
    )
    env_counts.columns = ["environment", "count"]

    # ── Build Summary DataFrame ───────────────────────────────────────────────
    summary_rows = [
        {"Description": "Total rows in DB (UC3.2 RAM input)", "Count": total_input, "Note": "Active servers with utilisation data"},
        {"Description": "", "Count": None, "Note": ""},
        {"Description": "--- PROD filters (Environment not in NON-PROD list) ---", "Count": prod_total, "Note": ""},
        {"Description": "UC3.2 PROD Filter 1: Avg_FreeMem_12m >= 35%", "Count": prod_after_f1, "Note": ""},
        {"Description": "UC3.2 PROD Filter 2: Min_FreeMem_12m >= 20%", "Count": prod_after_f2, "Note": ""},
        {"Description": "UC3.2 PROD Filter 3: Current_RAM_GiB > 8", "Count": prod_eligible, "Note": ""},
        {"Description": "UC3.2 PROD Filter 4: Recommendation band applied", "Count": prod_candidates, "Note": ""},
        {"Description": "", "Count": None, "Note": ""},
        {"Description": "--- NON-PROD filters (Development, Test, QA, UAT, DR) ---", "Count": nonprod_total, "Note": ""},
        {"Description": "UC3.2 NON-PROD Filter 1: Avg_FreeMem_12m >= 30%", "Count": nonprod_after_f1, "Note": ""},
        {"Description": "UC3.2 NON-PROD Filter 2: Min_FreeMem_12m >= 15%", "Count": nonprod_after_f2, "Note": ""},
        {"Description": "UC3.2 NON-PROD Filter 3: Current_RAM_GiB > 4", "Count": nonprod_eligible, "Note": ""},
        {"Description": "UC3.2 NON-PROD Filter 4: Recommendation band applied", "Count": nonprod_candidates, "Note": ""},
        {"Description": "", "Count": None, "Note": ""},
        {"Description": "UC3.2 RAM FINAL CANDIDATES", "Count": total_candidates, "Note": f"{prod_candidates} PROD, {nonprod_candidates} NON-PROD — Rows the rule will action"},
        {"Description": "", "Count": None, "Note": ""},
        {"Description": "--- Environment breakdown (all input rows) ---", "Count": None, "Note": ""},
    ]
    for _, env_row in env_counts.iterrows():
        summary_rows.append({
            "Description": f" {env_row['environment']}",
            "Count": int(env_row["count"]),
            "Note": "NON-PROD" if env_row["environment"] in NON_PROD_ENVS else "PROD",
        })
    summary_df = pd.DataFrame(summary_rows, columns=["Description", "Count", "Note"])

    # ── Write workbook ────────────────────────────────────────────────────────
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="UC3 RAM Input Data")
        summary_df.to_excel(writer, index=False, sheet_name="UC3 RAM Summary")

        try:
            from openpyxl.styles import Font, PatternFill, Alignment
            ws = writer.sheets["UC3 RAM Summary"]

            header_fill = PatternFill("solid", fgColor="1F3864")
            header_font = Font(bold=True, color="FFFFFF")
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            for row in ws.iter_rows(min_row=2):
                desc = str(row[0].value or "")
                if "FINAL CANDIDATES" in desc:
                    cand_fill = PatternFill("solid", fgColor="BDD7EE")
                    cand_font = Font(bold=True)
                    for cell in row:
                        cell.fill = cand_fill
                        cell.font = cand_font
                elif desc.startswith("---"):
                    section_fill = PatternFill("solid", fgColor="D9E1F2")
                    for cell in row:
                        cell.fill = section_fill
                        cell.font = Font(bold=True)

            for col in ws.columns:
                max_len = max((len(str(c.value or "")) for c in col), default=10)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)
        except Exception:
            pass

    buf.seek(0)
    response = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = _safe_content_disposition("uc3_ram_input_data.xlsx")
    return response


def _apply_summary_styles(ws):
    """Apply consistent header/highlight styles to a UC summary worksheet."""
    try:
        from openpyxl.styles import Font, PatternFill, Alignment
        header_fill = PatternFill("solid", fgColor="1F3864")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        for row in ws.iter_rows(min_row=2):
            desc = str(row[0].value or "")
            if "FINAL CANDIDATES" in desc or "FINAL FLAGS" in desc:
                for cell in row:
                    cell.fill = PatternFill("solid", fgColor="BDD7EE")
                    cell.font = Font(bold=True)
            elif desc.startswith("---"):
                for cell in row:
                    cell.fill = PatternFill("solid", fgColor="D9E1F2")
                    cell.font = Font(bold=True)
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)
    except Exception:
        pass


@require_GET
@login_required
def download_uc3_cpu_input_data(request):
    """
    Export the full UC3.1 CPU Rightsizing input dataset plus a filter-funnel summary.

    Sheet 1 – "UC3 CPU Input Data": all rows from _build_rightsizing_df().
    Sheet 2 – "UC3 CPU Summary": PROD / NON-PROD filter funnels + Environment breakdown.
    """
    from io import BytesIO
    from optimizer.services.db_analysis_service import _build_rightsizing_df
    from optimizer.rules.rightsizing import NON_PROD_ENVS, find_cpu_rightsizing_optimizations

    df = _build_rightsizing_df()
    if df.empty:
        return HttpResponse("No UC3 CPU rightsizing input data found in the database.", status=404)

    total_input = len(df)

    # ── PROD filter funnel ────────────────────────────────────────────────────
    prod_df = df[~df["Environment"].isin(NON_PROD_ENVS)].copy()
    prod_total = len(prod_df)

    avg_cpu  = pd.to_numeric(prod_df.get("Avg_CPU_12m",  pd.Series(dtype=float)), errors="coerce")
    peak_cpu = pd.to_numeric(prod_df.get("Peak_CPU_12m", pd.Series(dtype=float)), errors="coerce")
    vcpu     = pd.to_numeric(prod_df.get("Current_vCPU", pd.Series(dtype=float)), errors="coerce")

    mask_p1 = avg_cpu < 15
    prod_after_f1 = int(mask_p1.sum())
    mask_p2 = mask_p1 & (peak_cpu <= 70)
    prod_after_f2 = int(mask_p2.sum())
    mask_p3 = mask_p2 & (vcpu >= 4)
    prod_eligible = int(mask_p3.sum())

    # ── NON-PROD filter funnel ────────────────────────────────────────────────
    nonprod_df = df[df["Environment"].isin(NON_PROD_ENVS)].copy()
    nonprod_total = len(nonprod_df)

    avg_cpu_np  = pd.to_numeric(nonprod_df.get("Avg_CPU_12m",  pd.Series(dtype=float)), errors="coerce")
    peak_cpu_np = pd.to_numeric(nonprod_df.get("Peak_CPU_12m", pd.Series(dtype=float)), errors="coerce")
    vcpu_np     = pd.to_numeric(nonprod_df.get("Current_vCPU", pd.Series(dtype=float)), errors="coerce")

    mask_np1 = avg_cpu_np < 25
    nonprod_after_f1 = int(mask_np1.sum())
    mask_np2 = mask_np1 & (peak_cpu_np <= 80)
    nonprod_after_f2 = int(mask_np2.sum())
    mask_np3 = mask_np2 & (vcpu_np >= 4)
    nonprod_eligible = int(mask_np3.sum())

    # ── Actual final candidates via full rule ─────────────────────────────────
    final_df = find_cpu_rightsizing_optimizations(df)
    prod_candidates   = int((final_df["Env_Type"] == "PROD").sum())
    nonprod_candidates = int((final_df["Env_Type"] == "NON-PROD").sum())
    total_candidates  = prod_candidates + nonprod_candidates

    # ── Environment breakdown ─────────────────────────────────────────────────
    env_counts = df["Environment"].astype(str).value_counts().reset_index()
    env_counts.columns = ["environment", "count"]

    summary_rows = [
        {"Description": "Total rows in DB (UC3.1 CPU input)", "Count": total_input, "Note": "Active servers with utilisation data"},
        {"Description": "", "Count": None, "Note": ""},
        {"Description": "--- PROD filters (Environment not in NON-PROD list) ---", "Count": prod_total, "Note": ""},
        {"Description": "UC3.1 PROD Filter 1: Avg_CPU_12m < 15%", "Count": prod_after_f1, "Note": ""},
        {"Description": "UC3.1 PROD Filter 2: Peak_CPU_12m <= 70%", "Count": prod_after_f2, "Note": ""},
        {"Description": "UC3.1 PROD Filter 3: Current_vCPU >= 4", "Count": prod_eligible, "Note": ""},
        {"Description": "UC3.1 PROD Filter 4: Recommendation band applied", "Count": prod_candidates, "Note": ""},
        {"Description": "", "Count": None, "Note": ""},
        {"Description": "--- NON-PROD filters (Development, Test, QA, UAT, DR) ---", "Count": nonprod_total, "Note": ""},
        {"Description": "UC3.1 NON-PROD Filter 1: Avg_CPU_12m < 25%", "Count": nonprod_after_f1, "Note": ""},
        {"Description": "UC3.1 NON-PROD Filter 2: Peak_CPU_12m <= 80%", "Count": nonprod_after_f2, "Note": ""},
        {"Description": "UC3.1 NON-PROD Filter 3: Current_vCPU >= 4", "Count": nonprod_eligible, "Note": ""},
        {"Description": "UC3.1 NON-PROD Filter 4: Recommendation band applied", "Count": nonprod_candidates, "Note": ""},
        {"Description": "", "Count": None, "Note": ""},
        {"Description": "UC3.1 CPU FINAL CANDIDATES", "Count": total_candidates, "Note": f"{prod_candidates} PROD, {nonprod_candidates} NON-PROD — Rows the rule will action"},
        {"Description": "", "Count": None, "Note": ""},
        {"Description": "--- Environment breakdown (all input rows) ---", "Count": None, "Note": ""},
    ]
    for _, row in env_counts.iterrows():
        summary_rows.append({"Description": f" {row['environment']}", "Count": int(row["count"]), "Note": "NON-PROD" if row["environment"] in NON_PROD_ENVS else "PROD"})

    summary_df = pd.DataFrame(summary_rows, columns=["Description", "Count", "Note"])

    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="UC3 CPU Input Data")
        summary_df.to_excel(writer, index=False, sheet_name="UC3 CPU Summary")
        _apply_summary_styles(writer.sheets["UC3 CPU Summary"])

    buf.seek(0)
    response = HttpResponse(buf.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = _safe_content_disposition("uc3_cpu_input_data.xlsx")
    return response


@require_GET
@login_required
def download_uc3_crit_cpu_input_data(request):
    """
    Export the full UC3.3 Criticality-Aware CPU input dataset plus a filter-funnel summary.

    UC3.3a Downsize: Criticality in ALL_CRITICAL_VALS AND Avg_CPU_12m < 10%
    UC3.3b Upsize:   Criticality in CRITICAL_VALS AND Avg_CPU_12m > 80%

    Sheet 1 – "UC3 Crit CPU Input Data": all rows from _build_rightsizing_df().
    Sheet 2 – "UC3 Crit CPU Summary": downsize / upsize filter funnels + Criticality breakdown.
    """
    from io import BytesIO
    from optimizer.services.db_analysis_service import _build_rightsizing_df
    from optimizer.rules.rightsizing import (
        ALL_CRITICAL_VALS, CRITICAL_VALS,
        find_criticality_cpu_optimizations,
        find_criticality_cpu_downsize_optimizations,
        find_criticality_cpu_upsize_optimizations,
    )

    df = _build_rightsizing_df()
    if df.empty:
        return HttpResponse("No UC3 Criticality CPU input data found in the database.", status=404)

    total_input = len(df)

    # ── UC3.3a Downsize filter funnel ─────────────────────────────────────────
    crit_col = "Criticality" if "Criticality" in df.columns else None
    if crit_col:
        mask_all_crit = df[crit_col].isin(ALL_CRITICAL_VALS)
        total_critical = int(mask_all_crit.sum())
        avg_cpu = pd.to_numeric(df.get("Avg_CPU_12m", pd.Series(dtype=float)), errors="coerce")
        mask_down = mask_all_crit & (avg_cpu.fillna(100) < 10)
        downsize_eligible = int(mask_down.sum())

        # ── UC3.3b Upsize filter funnel ───────────────────────────────────────
        mask_bc_mc = df[crit_col].isin(CRITICAL_VALS)
        total_bc_mc = int(mask_bc_mc.sum())
        mask_up = mask_bc_mc & (avg_cpu.fillna(0) > 80)
        upsize_eligible = int(mask_up.sum())
    else:
        total_critical = downsize_eligible = total_bc_mc = upsize_eligible = 0

    # ── Actual final candidates via full rule ─────────────────────────────────
    final_df   = find_criticality_cpu_optimizations(df)
    downsize_df = find_criticality_cpu_downsize_optimizations(df)
    upsize_df   = find_criticality_cpu_upsize_optimizations(df)
    downsize_final = len(downsize_df)
    upsize_final   = len(upsize_df)
    total_candidates = len(final_df)

    # ── Criticality breakdown ─────────────────────────────────────────────────
    crit_counts = df["Criticality"].astype(str).value_counts().reset_index() if crit_col else pd.DataFrame(columns=["criticality", "count"])
    crit_counts.columns = ["criticality", "count"]

    summary_rows = [
        {"Description": "Total rows in DB (UC3.3 Crit CPU input)", "Count": total_input, "Note": "Active servers with utilisation data"},
        {"Description": "", "Count": None, "Note": ""},
        {"Description": "--- UC3.3a Downsize filters ---", "Count": None, "Note": ""},
        {"Description": "UC3.3a Filter 1: Criticality = Business/Mission/Manufacturing Critical", "Count": total_critical, "Note": f"Values: {', '.join(ALL_CRITICAL_VALS)}"},
        {"Description": "UC3.3a Filter 2: Avg_CPU_12m < 10%", "Count": downsize_eligible, "Note": ""},
        {"Description": "UC3.3a DOWNSIZE CANDIDATES", "Count": downsize_final, "Note": "Human Intervention Required"},
        {"Description": "", "Count": None, "Note": ""},
        {"Description": "--- UC3.3b Upsize filters ---", "Count": None, "Note": ""},
        {"Description": "UC3.3b Filter 1: Criticality = Business Critical / Mission Critical", "Count": total_bc_mc, "Note": f"Values: {', '.join(CRITICAL_VALS)}"},
        {"Description": "UC3.3b Filter 2: Avg_CPU_12m > 80%", "Count": upsize_eligible, "Note": ""},
        {"Description": "UC3.3b UPSIZE CANDIDATES", "Count": upsize_final, "Note": "Flag Only — Human Intervention Required"},
        {"Description": "", "Count": None, "Note": ""},
        {"Description": "UC3.3 Crit CPU FINAL FLAGS", "Count": total_candidates, "Note": f"{downsize_final} Downsize, {upsize_final} Upsize — Rows the rule will action"},
        {"Description": "", "Count": None, "Note": ""},
        {"Description": "--- Criticality breakdown (all input rows) ---", "Count": None, "Note": ""},
    ]
    for _, row in crit_counts.iterrows():
        summary_rows.append({"Description": f" {row['criticality']}", "Count": int(row["count"]), "Note": ""})

    summary_df = pd.DataFrame(summary_rows, columns=["Description", "Count", "Note"])

    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="UC3 Crit CPU Input Data")
        summary_df.to_excel(writer, index=False, sheet_name="UC3 Crit CPU Summary")
        _apply_summary_styles(writer.sheets["UC3 Crit CPU Summary"])

    buf.seek(0)
    response = HttpResponse(buf.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = _safe_content_disposition("uc3_crit_cpu_input_data.xlsx")
    return response


@require_GET
@login_required
def download_uc3_crit_ram_input_data(request):
    """
    Export the full UC3.4 Criticality-Aware RAM input dataset plus a filter-funnel summary.

    UC3.4a Downsize: Criticality in ALL_CRITICAL_VALS AND Avg_FreeMem_12m > 80%
    UC3.4b Upsize:   Criticality in CRITICAL_VALS AND Avg_FreeMem_12m < 20%

    Sheet 1 – "UC3 Crit RAM Input Data": all rows from _build_rightsizing_df().
    Sheet 2 – "UC3 Crit RAM Summary": downsize / upsize filter funnels + Criticality breakdown.
    """
    from io import BytesIO
    from optimizer.services.db_analysis_service import _build_rightsizing_df
    from optimizer.rules.rightsizing import (
        ALL_CRITICAL_VALS, CRITICAL_VALS,
        find_criticality_ram_optimizations,
        find_criticality_ram_downsize_optimizations,
        find_criticality_ram_upsize_optimizations,
    )

    df = _build_rightsizing_df()
    if df.empty:
        return HttpResponse("No UC3 Criticality RAM input data found in the database.", status=404)

    total_input = len(df)

    crit_col = "Criticality" if "Criticality" in df.columns else None
    avg_free = pd.to_numeric(df.get("Avg_FreeMem_12m", pd.Series(dtype=float)), errors="coerce")

    if crit_col:
        mask_all_crit = df[crit_col].isin(ALL_CRITICAL_VALS)
        total_critical = int(mask_all_crit.sum())
        mask_down = mask_all_crit & (avg_free.fillna(0) > 80)
        downsize_eligible = int(mask_down.sum())

        mask_bc_mc = df[crit_col].isin(CRITICAL_VALS)
        total_bc_mc = int(mask_bc_mc.sum())
        mask_up = mask_bc_mc & (avg_free.fillna(100) < 20)
        upsize_eligible = int(mask_up.sum())
    else:
        total_critical = downsize_eligible = total_bc_mc = upsize_eligible = 0

    final_df    = find_criticality_ram_optimizations(df)
    downsize_df = find_criticality_ram_downsize_optimizations(df)
    upsize_df   = find_criticality_ram_upsize_optimizations(df)
    downsize_final   = len(downsize_df)
    upsize_final     = len(upsize_df)
    total_candidates = len(final_df)

    crit_counts = df["Criticality"].astype(str).value_counts().reset_index() if crit_col else pd.DataFrame(columns=["criticality", "count"])
    crit_counts.columns = ["criticality", "count"]

    summary_rows = [
        {"Description": "Total rows in DB (UC3.4 Crit RAM input)", "Count": total_input, "Note": "Active servers with utilisation data"},
        {"Description": "", "Count": None, "Note": ""},
        {"Description": "--- UC3.4a Downsize filters ---", "Count": None, "Note": ""},
        {"Description": "UC3.4a Filter 1: Criticality = Business/Mission/Manufacturing Critical", "Count": total_critical, "Note": f"Values: {', '.join(ALL_CRITICAL_VALS)}"},
        {"Description": "UC3.4a Filter 2: Avg_FreeMem_12m > 80%", "Count": downsize_eligible, "Note": ""},
        {"Description": "UC3.4a DOWNSIZE CANDIDATES", "Count": downsize_final, "Note": "Downsize RAM by ~25% — Human Intervention Required"},
        {"Description": "", "Count": None, "Note": ""},
        {"Description": "--- UC3.4b Upsize filters ---", "Count": None, "Note": ""},
        {"Description": "UC3.4b Filter 1: Criticality = Business Critical / Mission Critical", "Count": total_bc_mc, "Note": f"Values: {', '.join(CRITICAL_VALS)}"},
        {"Description": "UC3.4b Filter 2: Avg_FreeMem_12m < 20%", "Count": upsize_eligible, "Note": ""},
        {"Description": "UC3.4b UPSIZE CANDIDATES", "Count": upsize_final, "Note": "Flag Only — Human Intervention Required"},
        {"Description": "", "Count": None, "Note": ""},
        {"Description": "UC3.4 Crit RAM FINAL FLAGS", "Count": total_candidates, "Note": f"{downsize_final} Downsize, {upsize_final} Upsize — Rows the rule will action"},
        {"Description": "", "Count": None, "Note": ""},
        {"Description": "--- Criticality breakdown (all input rows) ---", "Count": None, "Note": ""},
    ]
    for _, row in crit_counts.iterrows():
        summary_rows.append({"Description": f" {row['criticality']}", "Count": int(row["count"]), "Note": ""})

    summary_df = pd.DataFrame(summary_rows, columns=["Description", "Count", "Note"])

    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="UC3 Crit RAM Input Data")
        summary_df.to_excel(writer, index=False, sheet_name="UC3 Crit RAM Summary")
        _apply_summary_styles(writer.sheets["UC3 Crit RAM Summary"])

    buf.seek(0)
    response = HttpResponse(buf.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = _safe_content_disposition("uc3_crit_ram_input_data.xlsx")
    return response


@require_GET
@login_required
def download_uc3_physical_input_data(request):
    """
    Export the full Physical Systems input dataset plus a filter-funnel summary.

    Filter: is_virtual == False (case-insensitive) → Physical → flagged for human review.

    Sheet 1 – "Physical Input Data": all rows from _build_rightsizing_df().
    Sheet 2 – "Physical Summary": virtual/physical split + Environment breakdown.
    """
    from io import BytesIO
    from optimizer.services.db_analysis_service import _build_rightsizing_df
    from optimizer.rules.rightsizing import find_physical_systems_flags

    df = _build_rightsizing_df()
    if df.empty:
        return HttpResponse("No Physical Systems input data found in the database.", status=404)

    total_input = len(df)

    # ── Detect is_virtual column ──────────────────────────────────────────────
    virt_col = next((c for c in ("Is Virtual?", "is_virtual", "IsVirtual") if c in df.columns), None)
    if virt_col:
        is_physical_mask = df[virt_col].astype(str).str.strip().str.lower() == "false"
        total_physical = int(is_physical_mask.sum())
        total_virtual  = total_input - total_physical
    else:
        total_physical = total_virtual = 0

    final_df         = find_physical_systems_flags(df)
    total_candidates = len(final_df)

    # ── Environment breakdown ─────────────────────────────────────────────────
    env_counts = df["Environment"].astype(str).value_counts().reset_index()
    env_counts.columns = ["environment", "count"]

    from optimizer.rules.rightsizing import NON_PROD_ENVS

    summary_rows = [
        {"Description": "Total rows in DB (Physical Systems input)", "Count": total_input, "Note": "Active servers with utilisation data"},
        {"Description": "", "Count": None, "Note": ""},
        {"Description": "--- Physical Systems filter ---", "Count": None, "Note": ""},
        {"Description": "Physical Systems Filter 1: is_virtual = False", "Count": total_physical, "Note": "Physical servers only"},
        {"Description": "Virtual servers (excluded)", "Count": total_virtual, "Note": "is_virtual = True or blank"},
        {"Description": "", "Count": None, "Note": ""},
        {"Description": "Physical Systems FINAL FLAGS", "Count": total_candidates, "Note": "Human review required before any rightsizing action"},
        {"Description": "", "Count": None, "Note": ""},
        {"Description": "--- Environment breakdown (all input rows) ---", "Count": None, "Note": ""},
    ]
    for _, row in env_counts.iterrows():
        summary_rows.append({"Description": f" {row['environment']}", "Count": int(row["count"]), "Note": "NON-PROD" if row["environment"] in NON_PROD_ENVS else "PROD"})

    summary_df = pd.DataFrame(summary_rows, columns=["Description", "Count", "Note"])

    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Physical Input Data")
        summary_df.to_excel(writer, index=False, sheet_name="Physical Summary")
        _apply_summary_styles(writer.sheets["Physical Summary"])

    buf.seek(0)
    response = HttpResponse(buf.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = _safe_content_disposition("uc3_physical_input_data.xlsx")
    return response


def download_uc3_lifecycle_input_data(request):
    """
    Export the full UC3.5 Lifecycle Risk Flags input dataset plus a filter-funnel summary.

    Filters (sequential AND):
      Filter 1: Criticality in [Business Critical, Mission Critical]
      Filter 2: Peak_CPU_12m > 95%
      Filter 3: Min_FreeMem_12m < 5%

    Sheet 1 – "Lifecycle Input Data": all rows from _build_rightsizing_df().
    Sheet 2 – "Lifecycle Summary": sequential filter funnel + Criticality breakdown.
    """
    from io import BytesIO
    from optimizer.services.db_analysis_service import _build_rightsizing_df
    from optimizer.rules.rightsizing import find_lifecycle_risk_flags, LC_CRITICAL_VALS, COL_CRITICALITY

    df = _build_rightsizing_df()
    if df.empty:
        return HttpResponse("No Lifecycle Risk Flags input data found in the database.", status=404)

    total_input = len(df)

    # ── Sequential filter counts ──────────────────────────────────────────────
    has_criticality = COL_CRITICALITY in df.columns
    if has_criticality:
        step1 = df[df[COL_CRITICALITY].isin(LC_CRITICAL_VALS)]
        after_f1 = len(step1)
        step2 = step1[step1["Peak_CPU_12m"].fillna(0) > 95]
        after_f2 = len(step2)
        step3 = step2[step2["Min_FreeMem_12m"].fillna(100) < 5]
        after_f3 = len(step3)
    else:
        after_f1 = after_f2 = after_f3 = 0

    final_df         = find_lifecycle_risk_flags(df)
    total_candidates = len(final_df)

    # ── Criticality breakdown on final candidates ─────────────────────────────
    if not final_df.empty and has_criticality:
        crit_counts = final_df[COL_CRITICALITY].astype(str).value_counts().reset_index()
        crit_counts.columns = ["criticality", "count"]
    else:
        crit_counts = pd.DataFrame(columns=["criticality", "count"])

    summary_rows = [
        {"Description": "Total rows in DB (UC3.5 Lifecycle input)", "Count": total_input, "Note": "Active servers with utilisation data"},
        {"Description": "", "Count": None, "Note": ""},
        {"Description": "--- UC3.5 Lifecycle Risk Flag filters (sequential AND) ---", "Count": None, "Note": ""},
        {"Description": "UC3.5 Filter 1: Criticality = Business Critical OR Mission Critical", "Count": after_f1, "Note": f"of {total_input} input rows"},
        {"Description": "UC3.5 Filter 2: Peak_CPU_12m > 95%", "Count": after_f2, "Note": f"of {after_f1} after Filter 1"},
        {"Description": "UC3.5 Filter 3: Min_FreeMem_12m < 5%", "Count": after_f3, "Note": f"of {after_f2} after Filter 2"},
        {"Description": "", "Count": None, "Note": ""},
        {"Description": "UC3.5 Lifecycle Risk FINAL FLAGS", "Count": total_candidates, "Note": "Human review required — all 3 filters passed"},
        {"Description": "", "Count": None, "Note": ""},
        {"Description": "--- Criticality breakdown (final flagged systems) ---", "Count": None, "Note": ""},
    ]
    for _, row in crit_counts.iterrows():
        summary_rows.append({"Description": f" {row['criticality']}", "Count": int(row["count"]), "Note": ""})

    summary_df = pd.DataFrame(summary_rows, columns=["Description", "Count", "Note"])

    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Lifecycle Input Data")
        summary_df.to_excel(writer, index=False, sheet_name="Lifecycle Summary")
        _apply_summary_styles(writer.sheets["Lifecycle Summary"])

    buf.seek(0)
    response = HttpResponse(buf.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = _safe_content_disposition("uc3_lifecycle_input_data.xlsx")
    return response
